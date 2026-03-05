"""
Timetable Generator V4.0 — Streamlit Cloud App
Converted from tkinter (timetable_generator_v16.py)
All computation logic preserved exactly; only UI layer changed to Streamlit.
"""

import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import random
import math
import io
from datetime import datetime
from collections import defaultdict
import copy

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Timetable Generator V4.0",
    page_icon="📋",
    layout="wide",
)

# ─────────────────────────────────────────────────────────────────────────────
# Lightweight Variable wrapper  (mimics tk.StringVar / IntVar)
# Lets all original computation code call .get() and .set() unchanged.
# ─────────────────────────────────────────────────────────────────────────────
class Var:
    def __init__(self, value=None):
        self._v = value
    def get(self):
        return self._v
    def set(self, value):
        self._v = value

# ─────────────────────────────────────────────────────────────────────────────
# Global CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.step-header{font-size:1.4rem;font-weight:700;margin-bottom:0.3rem}
.step-sub{font-size:0.85rem;color:#555;margin-bottom:0.8rem}
.section-card{background:#f8f9fa;border:1px solid #dee2e6;border-radius:8px;padding:14px;margin-bottom:12px}
.tt-table{border-collapse:collapse;font-size:0.8rem;width:100%}
.tt-table th,.tt-table td{border:1px solid #aaa;padding:4px 6px;text-align:center;vertical-align:middle;min-width:100px}
.tt-hdr{background:#2c3e50;color:white;font-weight:700}
.tt-day{background:#34495e;color:white;font-weight:700;min-width:55px}
.tt-ct{background:#d5e8d4}
.tt-comb{background:#dae8fc}
.tt-par{background:#ffe6cc}
.tt-cpar{background:#f8cecc}
.tt-free{background:#f5f5f5;color:#999}
.tt-normal{background:#ffffff}
.card-overloaded{background:#fdecea;border:1px solid #e74c3c;border-radius:6px;padding:8px;margin:4px 0}
.card-ok{background:#ffffff;border:1px solid #ddd;border-radius:6px;padding:8px;margin:4px 0}
.card-skipped{background:#e8f5e9;border:1px solid #27ae60;border-radius:6px;padding:8px;margin:4px 0}
.card-resolved{background:#fff8e1;border:1px solid #f39c12;border-radius:6px;padding:8px;margin:4px 0}
.badge-overloaded{color:#c0392b;font-weight:700}
.badge-ok{color:#27ae60;font-weight:700}
.badge-skipped{color:#27ae60;font-weight:700}
.period-counter-ok{color:#1a7a1a;font-weight:700;font-size:0.95rem}
.period-counter-err{color:#c0392b;font-weight:700;font-size:0.95rem}
.combine-card{background:#e8f8f5;border:1px solid #27ae60;border-radius:6px;padding:8px;margin:4px 0}
.unavail-card{background:#fff3cd;border:1px solid #f39c12;border-radius:6px;padding:8px;margin:4px 0}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Session State initialization
# ─────────────────────────────────────────────────────────────────────────────
def _init():
    ss = st.session_state
    defaults = {
        'step': 1,
        'configuration': {},
        'class_config_data': {},
        'step3_data': {},
        'step3_unavailability': {},
        '_gen': None,
        '_timetable': None,
        '_last_allocation': {},
        '_last_all_rows': [],
        '_last_group_slots': {},
        '_relaxed_consec_keys': set(),
        '_relaxed_main_keys': set(),
        '_step3_selected_teacher': None,
        '_gen_stage': 0,
        'step2_active_class': None,
        'step2_editing_idx': None,
        's1_ppd': 7,
        's1_wdays': 6,
        's1_fh': 4,
        's1_sh': 3,
        's1_teacher_names': [],
        's1_class_sections': {c: 4 for c in range(6, 13)},
        '_ta_alloc_done': False,
        '_force_fill_result': None,
        's4_substep': 'main',  # 'main', 'ta', 'final'
        's3_unavail_edit_teacher': None,
        '_s1_cfg_upload_status': '',
        '_s2_cfg_upload_status': '',
        '_s3_cfg_upload_status': '',
        's2_form': {},  # ephemeral form values
        'validation_report': None,
        '_stage1_issues': [],
        'show_validation_report': False,
    }
    for k, v in defaults.items():
        if k not in ss:
            ss[k] = v

_init()
ss = st.session_state

# ─────────────────────────────────────────────────────────────────────────────
# Helper: access ss as "self" for computation methods
# We use a simple proxy object that reads/writes from st.session_state
# ─────────────────────────────────────────────────────────────────────────────
class App:
    """Proxy so computation methods can use self.* pattern unchanged."""

    def __getattr__(self, name):
        return st.session_state.get(name)

    def __setattr__(self, name, value):
        st.session_state[name] = value

    # ─────────────────────────────────────────────────────────────────────────
    #  Workload / combine helpers  (identical to original)
    # ─────────────────────────────────────────────────────────────────────────
    def _compute_teacher_workload(self):
        cfg   = self.configuration
        result = {}
        def _add(t, entry):
            if not t: return
            result.setdefault(t, {'total': 0, 'entries': []})
            result[t]['entries'].append(entry)
            result[t]['total'] += entry['periods']
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd = self.class_config_data[cn]
                ct = cd['teacher_var'].get().strip()
                ct_per = cd['teacher_period_var'].get()
                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t:
                        is_ct_subject = (t == ct)
                        ct_note = "  [incl. CT Period {}]".format(ct_per) if is_ct_subject else ""
                        _add(t, {
                            'class':    cn,
                            'subject':  s['name'],
                            'label':    "'{}' in {}  x{}/wk{}".format(
                                s['name'], cn, s['periods'], ct_note),
                            'periods':  s['periods'],
                            'is_ct':    is_ct_subject,
                            'ct_period': ct_per if is_ct_subject else None,
                        })
                    pt = s['parallel_teacher'].strip() if s.get('parallel') else ''
                    if pt:
                        _add(pt, {
                            'class':   cn,
                            'subject': s.get('parallel_subject', '?'),
                            'label':   "Parallel '{}' in {}  x{}/wk".format(
                                s.get('parallel_subject', '?'), cn, s['periods']),
                            'periods': s['periods'],
                            'is_ct':   False,
                            'ct_period': None,
                        })
        return result

    def _effective_total(self, teacher):
        wl    = getattr(self, '_step3_teacher_wl', {}) or {}
        total = wl.get(teacher, {}).get('total', 0)
        for cb in (self.step3_data or {}).get(teacher, {}).get('combines', []):
            n   = len(cb.get('entry_indices', []))
            per = cb.get('periods_each', 0)
            if n > 1:
                total -= (n - 1) * per
        return total

    def _get_class_ct_info(self, cn, teacher, teacher_subject):
        cd  = self.class_config_data.get(cn, {})
        ct_var = cd.get('teacher_var')
        ct  = ct_var.get().strip() if ct_var else ''
        subjs = cd.get('subjects', [])
        ct_subjects = [s['name'] for s in subjs if s.get('teacher', '').strip() == ct]
        is_parallel_with_ct = False
        parallel_ct_subject = ''
        for s in subjs:
            if (s.get('teacher', '').strip() == teacher
                    and s['name'] == teacher_subject
                    and s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s.get('parallel_subject', '')
                break
            if (s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == teacher
                    and s.get('parallel_subject', '') == teacher_subject
                    and s.get('teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s['name']
                break
        return {
            'ct': ct, 'ct_subjects': ct_subjects,
            'is_parallel_with_ct': is_parallel_with_ct,
            'parallel_ct_subject': parallel_ct_subject,
        }

    def _check_unavailability_feasible(self, teacher, blocked_days, blocked_periods):
        cfg       = self.configuration
        ppd       = cfg['periods_per_day']
        wdays     = cfg['working_days']
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]
        blocked_days_set    = set(blocked_days)
        blocked_periods_set = set(int(p) for p in blocked_periods)
        slot_conflicts = []
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd  = self.class_config_data[cn]
                ct  = cd['teacher_var'].get().strip()
                ct_per = cd['teacher_period_var'].get()
                if ct == teacher and ct_per in blocked_periods_set:
                    conflict_days = [d for d in day_names if d in blocked_days_set]
                    if conflict_days:
                        slot_conflicts.append(
                            "Class Teacher of {} (Period {}) conflicts on: {}".format(
                                cn, ct_per, ', '.join(conflict_days)))
                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t != teacher:
                        continue
                    s_periods = s.get('periods_pref', [])
                    s_days    = s.get('days_pref', [])
                    if not s_periods and not s_days:
                        continue
                    relevant_days = set(s_days) if s_days else set(day_names)
                    conflict_days = relevant_days & blocked_days_set
                    if not conflict_days:
                        continue
                    if s_periods:
                        bad_periods = set(s_periods) & blocked_periods_set
                        if bad_periods:
                            slot_conflicts.append(
                                "'{}' in {} — Period(s) {} on {} are both preferred and blocked".format(
                                    s['name'], cn, sorted(bad_periods), ', '.join(sorted(conflict_days))))
                    else:
                        slot_conflicts.append(
                            "'{}' in {} — preferred days {} overlap with blocked days".format(
                                s['name'], cn, ', '.join(sorted(conflict_days))))
        total_week    = ppd * wdays
        blocked_total = len(blocked_days_set) * len(blocked_periods_set)
        available     = total_week - blocked_total
        wl = getattr(self, '_step3_teacher_wl', {}) or {}
        assigned = self._effective_total(teacher) if teacher in wl else 0
        slot_ok  = available >= assigned
        free     = available - assigned
        parts = []
        if slot_conflicts:
            parts.append("SLOT CONFLICTS ({}):\n{}".format(
                len(slot_conflicts), "\n".join("  • " + c for c in slot_conflicts)))
        parts.append("CAPACITY: {} assigned, {} available after blocking ({} blocked, {} free).".format(
            assigned, available, blocked_total, free))
        message = "\n".join(parts)
        ok = (not slot_conflicts) and slot_ok
        if not ok:
            if slot_conflicts and not slot_ok:
                message = "Slot conflicts AND capacity problem.\n" + message
            elif slot_conflicts:
                message = "Slot conflicts found (capacity OK).\n" + message
            else:
                message = "Capacity problem.\n" + message
        return (ok, message)

    # ─────────────────────────────────────────────────────────────────────────
    #  Validation  (identical to original logic)
    # ─────────────────────────────────────────────────────────────────────────
    def _validate_and_complete(self):
        """Run all Step 2 validations. Returns a report dict."""
        cfg      = self.configuration
        ppd      = cfg['periods_per_day']
        wdays    = cfg['working_days']
        required = ppd * wdays
        days_all = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][:wdays]

        period_ok     = []
        period_errors = []
        hard_conflicts = []
        within_class_conflicts = []

        # ── 1. Period counts ──────────────────────────────────────────────
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd      = self.class_config_data[cn]
                subjects = cd.get('subjects', [])
                total = sum(s['periods'] for s in subjects)
                if total == required:
                    period_ok.append((cn, "{}/{} ✓".format(total, required)))
                else:
                    diff = total - required
                    period_errors.append((cn, "{}/{} ({}{})".format(
                        total, required, "+" if diff > 0 else "", diff)))

        # ── 2. Teacher-level hard conflicts ─────────────────────────────
        teacher_slots = defaultdict(list)
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd = self.class_config_data[cn]
                for s in cd.get('subjects', []):
                    t = s['teacher'].strip()
                    if not t:
                        continue
                    p_pref = s.get('periods_pref', [])
                    d_pref = s.get('days_pref', [])
                    if not p_pref and not d_pref:
                        continue
                    slot_key = "Class {} '{}' (Period(s) {}, Day(s) {})".format(
                        cn, s['name'],
                        sorted(p_pref) if p_pref else 'any',
                        sorted(d_pref) if d_pref else 'any')
                    teacher_slots[t].append({
                        'class': cn, 'subject': s['name'],
                        'periods': s['periods'],
                        'p_pref': set(p_pref), 'd_pref': set(d_pref),
                        'slot_key': slot_key,
                    })

        for teacher, slots in teacher_slots.items():
            for i in range(len(slots)):
                for j in range(i+1, len(slots)):
                    a, b = slots[i], slots[j]
                    if a['class'] == b['class']:
                        continue
                    if not a['p_pref'] or not b['p_pref']:
                        continue
                    a_days = a['d_pref'] if a['d_pref'] else set(days_all)
                    b_days = b['d_pref'] if b['d_pref'] else set(days_all)
                    common_days = a_days & b_days
                    if not common_days:
                        continue
                    common_periods = a['p_pref'] & b['p_pref']
                    if not common_periods:
                        continue
                    needed_a = math.ceil(a['periods'] / len(a_days))
                    needed_b = math.ceil(b['periods'] / len(b_days))
                    if needed_a + needed_b > len(common_periods):
                        for day in sorted(common_days):
                            hard_conflicts.append({
                                'teacher': teacher,
                                'slot_a': a['slot_key'],
                                'slot_b': b['slot_key'],
                                'reason': ("On {}, teacher {} is assigned to both "
                                           "'{}' in class {} (needs Period(s) {}) "
                                           "AND '{}' in class {} (needs Period(s) {}) "
                                           "— overlapping hard-fixed slots.").format(
                                    day, teacher,
                                    a['subject'], a['class'], sorted(a['p_pref']),
                                    b['subject'], b['class'], sorted(b['p_pref']))
                            })

        # ── 3. Within-class slot conflicts ───────────────────────────────
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd = self.class_config_data[cn]
                items = []
                for s in cd.get('subjects', []):
                    period_set = set(s.get('periods_pref', []))
                    day_set    = set(s.get('days_pref', []))
                    if not period_set and not day_set:
                        continue
                    if not period_set:
                        period_set = set(range(1, ppd+1))
                    if not day_set:
                        day_set = set(days_all)
                    n_periods  = s['periods']
                    n_days     = len(day_set)
                    need_per_day = math.ceil(n_periods / n_days) if n_days > 0 else 1
                    items.append({
                        'label': "Subject '{}' (Period(s) {}, day(s) {}, teacher: {})".format(
                            s['name'], sorted(set(s.get('periods_pref', []))),
                            sorted(day_set) if s['days_pref'] else 'any', s['teacher'].strip()),
                        'period_set': period_set,
                        'day_set': day_set,
                        'need_per_day': need_per_day,
                        'teacher': s['teacher'].strip(),
                    })
                if len(items) < 2:
                    continue
                for day in days_all:
                    active = [it for it in items if day in it['day_set']]
                    if len(active) < 2:
                        continue
                    for i in range(len(active)):
                        for j in range(i+1, len(active)):
                            a, b = active[i], active[j]
                            if a['teacher'] == b['teacher']:
                                continue
                            combined_slots = a['period_set'] | b['period_set']
                            combined_need  = a['need_per_day'] + b['need_per_day']
                            if combined_need > len(combined_slots):
                                contested = a['period_set'] & b['period_set']
                                within_class_conflicts.append({
                                    'class': cn, 'day': day,
                                    'item_a': a['label'], 'item_b': b['label'],
                                    'reason': ("On {}, two subjects by DIFFERENT teachers both "
                                               "need Period(s) {} in class {} — combined demand "
                                               "({} + {} = {}) exceeds available slots {}.").format(
                                        day, sorted(contested) if contested else sorted(combined_slots),
                                        cn, a['need_per_day'], b['need_per_day'], combined_need,
                                        sorted(combined_slots))
                                })
        return {
            'period_ok': period_ok,
            'period_errors': period_errors,
            'hard_conflicts': hard_conflicts,
            'within_class_conflicts': within_class_conflicts,
            'required': required,
        }

    # ─────────────────────────────────────────────────────────────────────────
    #  Generation Engine  (all unchanged from original)
    # ─────────────────────────────────────────────────────────────────────────
    def _init_gen_state(self):
        cfg   = self.configuration
        ppd   = cfg['periods_per_day']
        wdays = cfg['working_days']
        half1 = cfg['periods_first_half']
        DAYS  = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]
        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                all_classes.append("{}{}".format(cls, chr(65 + si)))
        grid    = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        task_at = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        t_busy  = {}
        def t_free(t, d, p):
            return not t or (d, p) not in t_busy.get(t, set())
        def t_mark(t, d, p):
            if t: t_busy.setdefault(t, set()).add((d, p))
        def t_unmark(t, d, p):
            if t: t_busy.get(t, set()).discard((d, p))
        unavail = self.step3_unavailability or {}
        def t_unavail(t, d, p):
            u = unavail.get(t, {})
            if not u: return False
            return DAYS[d] in u.get('days', []) and (p+1) in u.get('periods', [])
        s3 = self.step3_data or {}
        cn_subj_combined = {}
        for _teacher, s3d in s3.items():
            for cb in s3d.get('combines', []):
                classes  = sorted(cb.get('classes', []))
                subjects = cb.get('subjects', [])
                if len(classes) >= 2 and subjects:
                    for cn in classes:
                        cn_subj_combined[(cn, subjects[0])] = classes
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                primary_subj = s.get('name', '').strip()
                par_subj     = (s.get('parallel_subject') or '').strip()
                if par_subj and (cn, par_subj) in cn_subj_combined:
                    if (cn, primary_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, primary_subj)] = cn_subj_combined[(cn, par_subj)]
                if (cn, primary_subj) in cn_subj_combined and par_subj:
                    if (cn, par_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, par_subj)] = cn_subj_combined[(cn, primary_subj)]
        tasks = []
        seen_combined = set()
        for cn in all_classes:
            if cn not in self.class_config_data:
                continue
            cd     = self.class_config_data[cn]
            ct     = cd['teacher_var'].get().strip()
            ct_per = cd['teacher_period_var'].get()
            ct_subject_assigned = False
            for s in cd['subjects']:
                subj = s['name']
                t    = s['teacher'].strip()
                n    = s['periods']
                cn_list = cn_subj_combined.get((cn, subj), [cn])
                if len(cn_list) > 1:
                    key = (frozenset(cn_list), subj)
                    if key in seen_combined:
                        continue
                    seen_combined.add(key)
                if t == ct and not ct_subject_assigned:
                    is_ct = True
                    ct_subject_assigned = True
                else:
                    is_ct = False
                par    = bool(s.get('parallel', False))
                pt     = s.get('parallel_teacher', '').strip() if par else ''
                ps     = s.get('parallel_subject', '').strip() if par else ''
                consec = (s.get('consecutive', 'No') == 'Yes')
                if consec and (cn, subj) in (self._relaxed_consec_keys or set()):
                    consec = False
                p_pref = list(s.get('periods_pref', []))
                d_pref = list(s.get('days_pref', []))
                if len(cn_list) > 1 and par:
                    ttype = 'combined_parallel'
                elif len(cn_list) > 1:
                    ttype = 'combined'
                elif par:
                    ttype = 'parallel'
                else:
                    ttype = 'normal'
                if is_ct:
                    priority = 'HC1'
                elif p_pref or d_pref:
                    priority = 'HC2'
                elif consec:
                    priority = 'SC1'
                elif n >= wdays:
                    priority = 'SC2'
                else:
                    priority = 'filler'
                tasks.append({
                    'idx':       len(tasks),
                    'cn_list':   cn_list,
                    'subject':   subj,
                    'teacher':   t,
                    'par_subj':  ps,
                    'par_teach': pt,
                    'periods':   n,
                    'remaining': n,
                    'is_ct':     is_ct,
                    'ct_period': ct_per if is_ct else None,
                    'p_pref':    p_pref,
                    'd_pref':    d_pref,
                    'consec':    consec,
                    'daily':     (n >= wdays),
                    'priority':  priority,
                    'type':      ttype,
                    'rx_sc1':    False,
                    'rx_sc3':    False,
                    'rx_sc2':    False,
                })
        total_atoms = sum(t['periods'] for t in tasks)
        self._gen = {
            'cfg': cfg, 'ppd': ppd, 'wdays': wdays, 'half1': half1,
            'DAYS': DAYS, 'all_classes': all_classes,
            'grid': grid, 'task_at': task_at, 't_busy': t_busy,
            'tasks': tasks, 'total_atoms': total_atoms,
            't_free': t_free, 't_mark': t_mark, 't_unmark': t_unmark,
            't_unavail': t_unavail,
        }

    def _gen_can_place(self, task, d, p, ignore_sc1=False, ignore_sc3=False, ignore_sc2=False):
        g      = self._gen
        DAYS   = g['DAYS']; ppd = g['ppd']
        grid   = g['grid']
        t_free = g['t_free']; t_unavail = g['t_unavail']
        t      = task['teacher']; pt = task['par_teach']
        p1     = p + 1
        if task['is_ct'] and p1 != task['ct_period']:
            return False
        if task['p_pref'] and not task['is_ct']:
            if p1 not in task['p_pref']:
                return False
        if task['d_pref']:
            if DAYS[d] not in task['d_pref']:
                return False
        for cn in task['cn_list']:
            if grid[cn][d][p] is not None:
                return False
        if not t_free(t, d, p): return False
        if pt and not t_free(pt, d, p): return False
        if not (ignore_sc3 or task['rx_sc3']):
            if t_unavail(t, d, p): return False
            if pt and t_unavail(pt, d, p): return False
        if task['consec'] and not (ignore_sc1 or task['rx_sc1']):
            if p >= ppd - 1:
                return False
            for cn in task['cn_list']:
                if grid[cn][d][p + 1] is not None:
                    return False
            if not t_free(t, d, p + 1): return False
            if pt and not t_free(pt, d, p + 1): return False
            if not (ignore_sc3 or task['rx_sc3']):
                if t_unavail(t, d, p + 1): return False
                if pt and t_unavail(pt, d, p + 1): return False
        if not task['consec']:
            if not (ignore_sc2 or task['rx_sc2']):
                for cn in task['cn_list']:
                    for pp in range(ppd):
                        e = grid[cn][d][pp]
                        if e and e.get('subject') == task['subject']:
                            return False
        return True

    def _gen_count_valid_slots(self, task, ignore_sc1=False, ignore_sc3=False, ignore_sc2=False):
        g = self._gen
        return sum(1 for d in range(g['wdays']) for p in range(g['ppd'])
                   if self._gen_can_place(task, d, p, ignore_sc1, ignore_sc3, ignore_sc2))

    def _gen_make_cell(self, task):
        return {
            'type':      task['type'],
            'subject':   task['subject'],
            'teacher':   task['teacher'],
            'par_subj':  task['par_subj'],
            'par_teach': task['par_teach'],
            'combined_classes': task['cn_list'] if len(task['cn_list']) > 1 else [],
            'is_ct':     task['is_ct'],
        }

    def _gen_place(self, task, d, p):
        g = self._gen
        cell = self._gen_make_cell(task)
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = cell
            g['task_at'][cn][d][p] = task['idx']
        g['t_mark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_mark'](task['par_teach'], d, p)
        task['remaining'] -= 1

    def _gen_unplace(self, task, d, p):
        g = self._gen
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = None
            g['task_at'][cn][d][p] = None
        g['t_unmark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_unmark'](task['par_teach'], d, p)
        task['remaining'] += 1

    def _gen_prog(self, msg, extra_pct=0):
        # In Streamlit we don't update UI live; just store status
        st.session_state['_gen_status_msg'] = msg

    def _gen_snapshot_tt(self):
        g = self._gen
        unplaced = sum(t['remaining'] for t in g['tasks'])
        return {
            'grid':        g['grid'],
            'days':        g['DAYS'],
            'ppd':         g['ppd'],
            'half1':       g['half1'],
            'all_classes': g['all_classes'],
            'tasks':       g['tasks'],
            'unplaced':    unplaced,
        }

    # ── Stage 1 (identical to original, _gen_prog simplified above) ────────
    def _run_stage1_phases(self):
        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']
        s1_issues = []

        self._gen_prog("Stage 1 · Phase 1 — Placing Class Teacher subject periods…")
        for task in tasks:
            if task['priority'] != 'HC1':
                continue
            p_idx = task['ct_period'] - 1
            for d in range(wdays):
                if task['remaining'] <= 0:
                    break
                blocked_by = None
                for cn in task['cn_list']:
                    existing = grid[cn][d][p_idx]
                    if existing is not None:
                        blocked_by = (cn, existing)
                        break
                if blocked_by is None:
                    self._gen_place(task, d, p_idx)
                else:
                    cn_blk, cell_blk = blocked_by
                    s1_issues.append(
                        "HC1 — CT subject '{}' (teacher: {}, class: {}) "
                        "could NOT be placed on {} at Period {} — "
                        "cell already occupied by subject '{}' (teacher: {}).".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            DAYS[d], p_idx + 1,
                            cell_blk.get('subject', '?'),
                            cell_blk.get('teacher', '?')))

        self._gen_prog("Stage 1 · Phase 2 — Placing preference-constrained subjects…")
        hc2_tasks = sorted(
            [t for t in tasks if t['priority'] == 'HC2'],
            key=lambda t: (len(t['p_pref']) or ppd) * (len(t['d_pref']) or wdays))

        for task in hc2_tasks:
            if task['remaining'] <= 0:
                continue
            pref_p = [x - 1 for x in task['p_pref']] if task['p_pref'] else list(range(ppd))
            pref_d = ([DAYS.index(x) for x in task['d_pref'] if x in DAYS]
                      if task['d_pref'] else list(range(wdays)))
            slots = [(d, p) for d in pref_d for p in pref_p]
            blocked_slots = []
            for d, p in slots:
                if task['remaining'] <= 0:
                    break
                blocked_by = None
                for cn in task['cn_list']:
                    existing = grid[cn][d][p]
                    if existing is not None:
                        blocked_by = (cn, existing)
                        break
                t_ok = (g['t_free'](task['teacher'], d, p) and
                        not g['t_unavail'](task['teacher'], d, p))
                pt   = task['par_teach']
                if pt and pt not in ('', '—', '?'):
                    t_ok = t_ok and (g['t_free'](pt, d, p) and
                                     not g['t_unavail'](pt, d, p))
                if blocked_by is None and t_ok:
                    self._gen_place(task, d, p)
                else:
                    cn_blk, cell_blk = blocked_by if blocked_by else (task['cn_list'][0], {})
                    blocked_slots.append(
                        "{} P{}: occupied by '{}' or teacher busy".format(
                            DAYS[d], p + 1, cell_blk.get('subject', 'teacher conflict')))
            if task['remaining'] > 0 and blocked_slots:
                s1_issues.append(
                    "HC2 — '{}' in {} has {} period(s) unplaced after Stage 1. "
                    "Blocked slots: {}".format(
                        task['subject'], '+'.join(task['cn_list']),
                        task['remaining'], '; '.join(blocked_slots[:3])))

        st.session_state['_stage1_issues'] = s1_issues
        self._gen_prog("Stage 1 complete.")

    # ── Stage 2 phases (Stage 3 in original naming) ────────────────────────
    def _run_stage2_phases(self):
        """Fill remaining periods after Stage 1 + task-analysis allocation."""
        g      = self._gen
        tasks  = g['tasks']
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']
        DAYS   = g['DAYS']

        if not hasattr(self, '_relaxed_main_keys') or self._relaxed_main_keys is None:
            self._relaxed_main_keys = set()

        relax_level = 0
        for rep in range(80):
            remaining_tasks = [t for t in tasks if t["remaining"] > 0]
            if not remaining_tasks:
                break
            self._gen_prog("Stage 3 · Repair {}: {} unplaced, relax={}".format(
                rep + 1, sum(t["remaining"] for t in remaining_tasks), relax_level))
            ix_sc1 = relax_level >= 1
            ix_sc3 = relax_level >= 2
            ix_sc2 = relax_level >= 3
            if ix_sc1:
                for t in tasks: t["rx_sc1"] = True
            if ix_sc3:
                for t in tasks: t["rx_sc3"] = True
            if ix_sc2:
                for t in tasks: t["rx_sc2"] = True
            progress = False
            for task in sorted(remaining_tasks, key=lambda t: -t["remaining"]):
                if task["remaining"] <= 0:
                    continue
                pt = task["par_teach"]
                for d in range(wdays):
                    if task["remaining"] <= 0: break
                    for p in range(ppd):
                        if task["remaining"] <= 0: break
                        if self._gen_can_place(task, d, p, ix_sc1, ix_sc3, ix_sc2):
                            self._gen_place(task, d, p)
                            progress = True
                if task["remaining"] <= 0:
                    continue
                swap_done = False
                for d in range(wdays):
                    if task["remaining"] <= 0 or swap_done: break
                    for p in range(ppd):
                        if task["remaining"] <= 0 or swap_done: break
                        if not (g["t_free"](task["teacher"], d, p) and
                                (ix_sc3 or not g["t_unavail"](task["teacher"], d, p))):
                            continue
                        if pt and not (g["t_free"](pt, d, p) and
                                       (ix_sc3 or not g["t_unavail"](pt, d, p))):
                            continue
                        if task["is_ct"] and (p + 1) != task["ct_period"]:
                            continue
                        if task["p_pref"] and not task["is_ct"]:
                            if (p + 1) not in task["p_pref"]: continue
                        if task["d_pref"] and DAYS[d] not in task["d_pref"]:
                            continue
                        if not task["consec"]:
                            dup = False
                            for cn in task["cn_list"]:
                                for pp in range(ppd):
                                    e = g["grid"][cn][d][pp]
                                    if e and e.get("subject") == task["subject"]:
                                        dup = True; break
                                if dup: break
                            if dup: continue
                        blocking_idx = None
                        for cn in task["cn_list"]:
                            if g["grid"][cn][d][p] is not None:
                                blocking_idx = g["task_at"][cn][d][p]
                                break
                        if blocking_idx is None:
                            self._gen_place(task, d, p)
                            progress = True
                            swap_done = True
                            break
                        if blocking_idx >= len(tasks): continue
                        blocker = tasks[blocking_idx]
                        if blocker["priority"] in ("HC1", "HC2"): continue
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p): continue
                                if not self._gen_can_place(blocker, d2, p2, ix_sc1, ix_sc3, ix_sc2):
                                    continue
                                self._gen_unplace(blocker, d, p)
                                slot_clear = all(g["grid"][cn][d][p] is None for cn in task["cn_list"])
                                t_now = (g["t_free"](task["teacher"], d, p) and
                                         (not pt or g["t_free"](pt, d, p)))
                                if slot_clear and t_now:
                                    self._gen_place(task, d, p)
                                    self._gen_place(blocker, d2, p2)
                                    progress = True
                                    swap_done = True
                                    moved = True
                                    break
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved or swap_done: break
                        if swap_done: break
            if not progress:
                relax_level += 1
                if relax_level > 4:
                    break

    # ── Force Fill ──────────────────────────────────────────────────────────
    def _force_fill_backtrack(self, progress_cb=None):
        def _prog(msg):
            if progress_cb: progress_cb(msg)
        g      = self._gen
        tasks  = g['tasks']
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']
        if not hasattr(self, '_relaxed_main_keys') or self._relaxed_main_keys is None:
            self._relaxed_main_keys = set()
        if not hasattr(self, '_relaxed_consec_keys') or self._relaxed_consec_keys is None:
            self._relaxed_consec_keys = set()
        relaxed_notes = []
        PRIO_W = {'HC1': 0, 'HC2': 1, 'SC1': 2, 'SC2': 3, 'filler': 4}
        def _prio(t):
            return PRIO_W.get(t['priority'], 4)
        def _unplaced():
            return sum(t['remaining'] for t in tasks)
        def _can(task, d, p, ign_sc1=False, ign_sc3=False):
            return self._gen_can_place(task, d, p, ignore_sc1=ign_sc1, ignore_sc3=ign_sc3)

        def _greedy_pass(ign_sc1=False, ign_sc3=False):
            remaining_tasks = [t for t in tasks if t['remaining'] > 0]
            remaining_tasks.sort(key=lambda t: sum(
                1 for d in range(wdays) for p in range(ppd) if _can(t, d, p, ign_sc1, ign_sc3)))
            for task in remaining_tasks:
                for d in range(wdays):
                    if task['remaining'] <= 0: break
                    for p in range(ppd):
                        if task['remaining'] <= 0: break
                        if _can(task, d, p, ign_sc1, ign_sc3):
                            self._gen_place(task, d, p)

        def _swap_pass(ign_sc1=False, ign_sc3=False):
            for task in sorted(tasks, key=lambda t: -t['remaining']):
                if task['remaining'] <= 0 or _prio(task) == 0:
                    continue
                for d in range(wdays):
                    if task['remaining'] <= 0: break
                    for p in range(ppd):
                        if task['remaining'] <= 0: break
                        tname = task['teacher']
                        pt    = task.get('par_teach', '')
                        t_ok  = g['t_free'](tname, d, p)
                        if not ign_sc3:
                            t_ok = t_ok and not g['t_unavail'](tname, d, p)
                        if pt and pt not in ('', '—', '?'):
                            t_ok = t_ok and g['t_free'](pt, d, p)
                        if not t_ok: continue
                        bidx = None
                        for cn in task['cn_list']:
                            if grid[cn][d][p] is not None:
                                bidx = g['task_at'][cn][d][p]
                                break
                        if bidx is None:
                            if _can(task, d, p, ign_sc1, ign_sc3):
                                self._gen_place(task, d, p)
                            continue
                        blocker = tasks[bidx]
                        if _prio(blocker) <= _prio(task): continue
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p): continue
                                if not _can(blocker, d2, p2, ign_sc1, ign_sc3): continue
                                self._gen_unplace(blocker, d, p)
                                clr = all(grid[cn][d][p] is None for cn in task['cn_list'])
                                tok = (g['t_free'](tname, d, p) and
                                       (not pt or pt in ('','—','?') or g['t_free'](pt, d, p)))
                                if clr and tok:
                                    self._gen_place(blocker, d2, p2)
                                    if _can(task, d, p, ign_sc1, ign_sc3):
                                        self._gen_place(task, d, p)
                                        moved = True
                                        break
                                    else:
                                        self._gen_unplace(blocker, d2, p2)
                                        self._gen_place(blocker, d, p)
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved: break

        def _run_stage_a(ign_sc1=False, ign_sc3=False):
            for _ in range(4):
                if _unplaced() == 0: return
                _greedy_pass(ign_sc1, ign_sc3)
            for _ in range(4):
                if _unplaced() == 0: return
                _swap_pass(ign_sc1, ign_sc3)
                _greedy_pass(ign_sc1, ign_sc3)

        _prog("Stage A — greedy placement…")
        _run_stage_a()
        if _unplaced() == 0:
            _prog(""); return None

        _prog("Stage A — relaxing consecutive…")
        consec_items = []
        for t in tasks:
            if t['consec'] and t['remaining'] > 0:
                t['rx_sc1'] = True
                for cn_i in t['cn_list']:
                    self._relaxed_consec_keys.add((cn_i, t['subject']))
                consec_items.append("  • {} — {}".format('+'.join(t['cn_list']), t['subject']))
        if consec_items:
            relaxed_notes.append("Consecutive constraint relaxed for:\n" + '\n'.join(consec_items))
        _run_stage_a(ign_sc1=True)
        if _unplaced() == 0:
            _prog(""); return '\n\n'.join(relaxed_notes)

        _prog("Stage A — relaxing unavailability…")
        unav_set = set()
        for t in tasks:
            if t['remaining'] > 0:
                t['rx_sc3'] = True
                if t['teacher']: unav_set.add(t['teacher'])
                pt = t.get('par_teach', '')
                if pt and pt not in ('', '—', '?'): unav_set.add(pt)
        if unav_set:
            relaxed_notes.append("Teacher unavailability bypassed for:\n" +
                '\n'.join("  • {}".format(x) for x in sorted(unav_set)))
        _run_stage_a(ign_sc1=True, ign_sc3=True)
        if _unplaced() == 0:
            _prog(""); return '\n\n'.join(relaxed_notes)

        _prog("Stage A — relaxing preferences…")
        main_items = []
        for t in tasks:
            if t['remaining'] == 0 or t.get('is_ct'): continue
            if t['p_pref'] or t['d_pref'] or t.get('daily') or t['priority'] == 'SC2':
                t['p_pref'] = []; t['d_pref'] = []
                t['daily'] = False; t['priority'] = 'filler'
                self._relaxed_main_keys.add((frozenset(t['cn_list']), t['subject']))
                main_items.append("  • {} — {}".format('+'.join(t['cn_list']), t['subject']))
        if main_items:
            relaxed_notes.append("Period/day preferences relaxed for:\n" + '\n'.join(main_items))
        _run_stage_a(ign_sc1=True, ign_sc3=True)
        if _unplaced() == 0:
            _prog(""); return '\n\n'.join(relaxed_notes)

        # Stage B: Min-Conflicts
        relaxed_notes.append("Min-Conflicts solver applied: soft constraints overridden.")
        MAX_ITER = 1500; RESTART = 150
        _prog("Stage B — Min-Conflicts CSP solver…")
        for iteration in range(MAX_ITER):
            if _unplaced() == 0: break
            if iteration % RESTART == 0 and iteration > 0:
                # Random restart: unplace some low-priority assigned tasks
                low_pri = [t for t in tasks if t['remaining'] == 0 and _prio(t) >= 3]
                if low_pri:
                    sel = random.choice(low_pri)
                    placed_cells = []
                    for d in range(wdays):
                        for p in range(ppd):
                            for cn in sel['cn_list']:
                                if (self._gen and
                                        g['grid'].get(cn, [[]])[d][p] is not None and
                                        g['task_at'].get(cn, [[]])[d][p] == sel['idx']):
                                    placed_cells.append((d, p))
                                    break
                    for (d, p) in placed_cells:
                        self._gen_unplace(sel, d, p)

            conflict_tasks = [t for t in tasks if t['remaining'] > 0]
            if not conflict_tasks: break
            task = min(conflict_tasks,
                       key=lambda t: self._gen_count_valid_slots(t, True, True))
            placed = False
            for d in range(wdays):
                if placed: break
                for p in range(ppd):
                    if self._gen_can_place(task, d, p, True, True):
                        self._gen_place(task, d, p)
                        placed = True; break
            if not placed:
                # Try displacing
                for d in range(wdays):
                    if placed: break
                    for p in range(ppd):
                        if placed: break
                        for cn in task['cn_list']:
                            bidx = g['task_at'][cn][d][p]
                            if bidx is None: continue
                            blocker = tasks[bidx]
                            if _prio(blocker) <= _prio(task): continue
                            self._gen_unplace(blocker, d, p)
                            if self._gen_can_place(task, d, p, True, True):
                                self._gen_place(task, d, p)
                                placed = True
                                # Try to re-place blocker
                                for d2 in range(wdays):
                                    for p2 in range(ppd):
                                        if self._gen_can_place(blocker, d2, p2, True, True):
                                            self._gen_place(blocker, d2, p2)
                                            break
                                    if blocker['remaining'] < blocker['periods']: break
                                break
                            else:
                                self._gen_place(blocker, d, p)
                        if placed: break
        _prog("")
        return '\n\n'.join(relaxed_notes) if relaxed_notes else None

    # ── Task analysis helpers ───────────────────────────────────────────────
    def _calculate_group_slots(self, all_rows):
        if not self._gen:
            return {}
        g = self._gen
        result = {}
        DAYS = g['DAYS']
        ppd  = g['ppd']
        wdays = g['wdays']
        groups = {}
        for r in all_rows:
            groups.setdefault(r['group'], r)

        for gn, row in groups.items():
            sec = row.get('section', 'A')
            if sec == 'A':
                classes  = [r['class'] for r in all_rows if r['group'] == gn]
                subjects = [r['subject'] for r in all_rows if r['group'] == gn]
                if not classes: continue
                s0 = subjects[0] if subjects else ''
                total = 0
                for cn in classes:
                    cd = self.class_config_data.get(cn, {})
                    for s in cd.get('subjects', []):
                        if s['name'] == s0:
                            total = s['periods']
                            break
                    if total: break
                result[gn] = {'needed': total, 'section': 'A'}
            elif sec == 'B':
                cn = row['class']
                subj = row['subject']
                total = 0
                cd = self.class_config_data.get(cn, {})
                for s in cd.get('subjects', []):
                    if s['name'] == subj:
                        total = s['periods']
                        break
                result[gn] = {'needed': total, 'section': 'B'}
            elif sec == 'C':
                cn = row['class']
                subj = row['subject']
                total = 0
                cd = self.class_config_data.get(cn, {})
                for s in cd.get('subjects', []):
                    if s['name'] == subj:
                        total = s['periods']
                        break
                result[gn] = {'needed': total, 'section': 'C'}
        return result

    def _run_ta2_allocation(self):
        if not self._gen:
            return {}
        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']
        if not hasattr(self, '_relaxed_main_keys') or self._relaxed_main_keys is None:
            self._relaxed_main_keys = set()
        results = {}
        def slot_free(task, d, p):
            for cn in task['cn_list']:
                if grid.get(cn, [[]])[d][p] is not None:
                    return False
            t  = task['teacher']
            pt = task.get('par_teach', '')
            if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                return False
            if pt and pt not in ('', '—', '?'):
                if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                    return False
            return True

        def place_slot(task, extra_par, d, p, class_info_map):
            cell = self._gen_make_cell(task)
            for cn in task['cn_list']:
                grid[cn][d][p] = cell
                g['task_at'][cn][d][p] = task['idx']
            g['t_mark'](task['teacher'], d, p)
            if task.get('par_teach') and task['par_teach'] not in ('', '—', '?'):
                g['t_mark'](task['par_teach'], d, p)
            task['remaining'] -= 1

        def slot_is_free_for_classes(cn_list, d, p):
            return all(grid.get(cn, [[]])[d][p] is None for cn in cn_list)

        def all_teachers_free(teachers, d, p):
            return all(g['t_free'](t, d, p) and not g['t_unavail'](t, d, p)
                       for t in teachers if t and t not in ('—', '?', ''))

        s3 = self.step3_data or {}
        cfg = self.configuration
        all_classes = g['all_classes']

        def _find_parallel(cn, sn):
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if s['name'] == sn and s.get('parallel'):
                    return (s.get('parallel_subject','').strip() or '?',
                            s.get('parallel_teacher','').strip() or '—')
                if (s.get('parallel') and
                        s.get('parallel_subject','').strip() == sn and s['name'] != sn):
                    return (s['name'], s.get('teacher','').strip() or '—')
            return ('—', '—')

        all_rows_local = []
        group_no = 0
        covered = set()
        for teacher, s3d in sorted(s3.items()):
            for cb in s3d.get('combines', []):
                classes  = cb.get('classes', [])
                subjects = cb.get('subjects', [])
                if not classes: continue
                group_no += 1
                for j, cn in enumerate(classes):
                    tsub = subjects[j] if j < len(subjects) else (subjects[0] if subjects else '?')
                    all_rows_local.append({'group': group_no, 'class': cn, 'subject': tsub,
                                           'teacher': teacher, 'section': 'A'})
                    covered.add((cn, tsub))
        seen_pairs = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if not s.get('parallel'): continue
                sn = s['name']; st_ = s.get('teacher','').strip()
                ps = s.get('parallel_subject','').strip()
                pt_ = s.get('parallel_teacher','').strip()
                if not ps: continue
                if (cn, sn) in covered or (cn, ps) in covered: continue
                pk = frozenset([(cn, sn), (cn, ps)])
                if pk in seen_pairs: continue
                seen_pairs.add(pk)
                group_no += 1
                all_rows_local.append({'group': group_no, 'class': cn, 'subject': sn,
                                       'teacher': st_, 'section': 'B'})
        consec_covered = set(covered)
        for r in all_rows_local:
            if r['section'] == 'B':
                consec_covered.add((r['class'], r['subject']))
        seen_consec = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if s.get('consecutive', 'No') != 'Yes': continue
                sn = s['name']; st_ = s.get('teacher','').strip()
                key = (cn, sn)
                if key in seen_consec: continue
                seen_consec.add(key)
                group_no += 1
                all_rows_local.append({'group': group_no, 'class': cn, 'subject': sn,
                                       'teacher': st_, 'section': 'C'})

        grp_map = {}
        for r in all_rows_local:
            grp_map.setdefault(r['group'], []).append(r)

        for gn, rows in grp_map.items():
            sec = rows[0].get('section', 'A')
            all_cn = list(dict.fromkeys(r['class'] for r in rows))
            teacher = rows[0]['teacher']
            subj0   = rows[0]['subject']

            # Find task
            task = None
            for t in tasks:
                if (subj0 in t['cn_list'] or subj0 == t['subject']) and \
                        t['subject'] == subj0 and set(t['cn_list']) == set(all_cn):
                    task = t; break
            if task is None:
                for t in tasks:
                    if t['subject'] == subj0 and t['teacher'] == teacher:
                        task = t; break
            if task is None:
                results[gn] = {'ok': False, 'total': 0, 's1_placed': 0,
                               'new_placed': 0, 'slots': [],
                               'reason': "Task not found in engine for '{}' / {}".format(subj0, all_cn)}
                continue

            total_periods = task['periods']
            s1_placed     = total_periods - task['remaining']
            remaining     = task['remaining']

            if remaining == 0:
                results[gn] = {'ok': True, 'total': total_periods,
                               's1_placed': s1_placed, 'new_placed': 0, 'slots': []}
                continue

            all_teachers_needed = [teacher]
            for r in rows:
                pt_ = _find_parallel(r['class'], r['subject'])[1]
                if pt_ and pt_ not in ('—','?','') and pt_ not in all_teachers_needed:
                    all_teachers_needed.append(pt_)

            placed_slots = []
            last_fail_why = ''
            class_info_map = {}

            extra_par = None

            is_consec = (sec == 'C' and task.get('consec', False) and
                         (frozenset(all_cn), subj0) not in (self._relaxed_main_keys or set()))

            if is_consec:
                for d in range(wdays):
                    if len(placed_slots) >= remaining: break
                    for p1 in range(ppd - 1):
                        if len(placed_slots) >= remaining: break
                        p2 = p1 + 1
                        cls_ok = (slot_is_free_for_classes(all_cn, d, p1) and
                                  slot_is_free_for_classes(all_cn, d, p2))
                        tch_ok = (all_teachers_free(all_teachers_needed, d, p1) and
                                  all_teachers_free(all_teachers_needed, d, p2))
                        if cls_ok and tch_ok:
                            if remaining - len(placed_slots) >= 2:
                                place_slot(task, extra_par, d, p1, class_info_map)
                                place_slot(task, extra_par, d, p2, class_info_map)
                                placed_slots.extend([(d, p1), (d, p2)])
                            else:
                                place_slot(task, extra_par, d, p1, class_info_map)
                                placed_slots.append((d, p1))
                        else:
                            last_fail_why = '{} P{}-P{}: {}'.format(
                                DAYS[d], p1+1, p2+1,
                                'class occupied' if not cls_ok else 'teacher busy')
            else:
                for p in range(ppd - 1, -1, -1):
                    if len(placed_slots) >= remaining: break
                    for d in range(wdays):
                        if len(placed_slots) >= remaining: break
                        cls_ok = slot_is_free_for_classes(all_cn, d, p)
                        tch_ok = all_teachers_free(all_teachers_needed, d, p)
                        if cls_ok and tch_ok:
                            place_slot(task, extra_par, d, p, class_info_map)
                            placed_slots.append((d, p))
                        else:
                            last_fail_why = '{} P{}: {}'.format(
                                DAYS[d], p+1,
                                'class occupied' if not cls_ok else 'teacher busy')

            new_placed = len(placed_slots)
            if new_placed >= remaining:
                results[gn] = {'ok': True, 'total': total_periods,
                               's1_placed': s1_placed, 'new_placed': new_placed,
                               'slots': placed_slots}
            else:
                still_short = remaining - new_placed
                reason = ('{} slot(s) still unplaced. Last conflict: {}'.format(
                    still_short, last_fail_why or 'Unknown')
                    if new_placed > 0 else
                    'No free slots found. ' + (last_fail_why or 'All slots occupied'))
                results[gn] = {'ok': False, 'total': total_periods,
                               's1_placed': s1_placed, 'new_placed': new_placed,
                               'slots': placed_slots, 'reason': reason}
        return results

    def _run_task_analysis_allocation(self):
        s3  = self.step3_data or {}
        cfg = self.configuration
        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'].get(cls, 0)):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        def _find_parallel(cn, sn):
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if s['name'] == sn and s.get('parallel'):
                    return (s.get('parallel_subject','').strip() or '?',
                            s.get('parallel_teacher','').strip() or '—')
                if (s.get('parallel') and
                        s.get('parallel_subject','').strip() == sn and s['name'] != sn):
                    return (s['name'], s.get('teacher','').strip() or '—')
            return ('—','—')

        all_rows = []
        group_no = 0; covered = set()
        for teacher, s3d in sorted(s3.items()):
            for cb in s3d.get('combines', []):
                classes  = cb.get('classes', [])
                subjects = cb.get('subjects', [])
                if not classes: continue
                group_no += 1
                for j, cn in enumerate(classes):
                    tsub = subjects[j] if j < len(subjects) else (subjects[0] if subjects else '?')
                    ps, pt_ = _find_parallel(cn, tsub)
                    all_rows.append({'group': group_no, 'class': cn, 'subject': tsub,
                                     'teacher': teacher, 'par_subj': ps, 'par_teacher': pt_,
                                     'section': 'A'})
                    covered.add((cn, tsub))
                    if ps not in ('—','?'): covered.add((cn, ps))

        seen_pairs = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if not s.get('parallel'): continue
                sn = s['name']; st_ = s.get('teacher','').strip()
                ps = s.get('parallel_subject','').strip()
                pt_ = s.get('parallel_teacher','').strip()
                if not ps: continue
                if (cn, sn) in covered or (cn, ps) in covered: continue
                pk = frozenset([(cn, sn), (cn, ps)])
                if pk in seen_pairs: continue
                seen_pairs.add(pk)
                group_no += 1
                all_rows.append({'group': group_no, 'class': cn, 'subject': sn,
                                 'teacher': st_, 'par_subj': ps or '?',
                                 'par_teacher': pt_ or '—', 'section': 'B'})

        consec_covered = set(covered)
        for r in all_rows:
            if r.get('section') == 'B':
                consec_covered.add((r['class'], r['subject']))
                if r.get('par_subj') not in ('—','?',''): consec_covered.add((r['class'], r['par_subj']))
        seen_consec = set()
        for cn in all_classes:
            for s in self.class_config_data.get(cn, {}).get('subjects', []):
                if s.get('consecutive','No') != 'Yes': continue
                sn = s['name']; st_ = s.get('teacher','').strip()
                key = (cn, sn)
                if key in seen_consec: continue
                seen_consec.add(key)
                group_no += 1
                all_rows.append({'group': group_no, 'class': cn, 'subject': sn,
                                 'teacher': st_, 'par_subj':'—', 'par_teacher':'—',
                                 'section': 'C', 'periods': s.get('periods','')})

        group_slots = self._calculate_group_slots(all_rows)
        group_allocation = self._run_ta2_allocation()
        return group_slots, group_allocation, all_rows

    def _allocation_suggestion(self, reason, rows, sec):
        reason_l = (reason or '').lower()
        teachers = list(dict.fromkeys(
            t for r in rows for t in [r.get('teacher',''), r.get('par_teacher','')]
            if t and t not in ('—','?','')))
        classes = list(dict.fromkeys(r['class'] for r in rows))
        if 'stage 1 not run' in reason_l:
            return "Run Stage 1 first before allocating periods."
        if 'not found in engine' in reason_l or 'not in task list' in reason_l:
            return "The subject name in Step 2 may not exactly match what Step 3 recorded."
        if 'teacher' in reason_l and 'busy' in reason_l:
            t_str = ', '.join(teachers)
            return ("Teacher {} is fully occupied. Options: (a) Reduce periods in Step 2, "
                    "(b) Remove unavailability block in Step 3, (c) Reassign subject.".format(t_str))
        if 'occupied' in reason_l or 'class' in reason_l:
            c_str = ', '.join(classes)
            return ("Class {} has no free slots. Options: (a) Reduce period count for another "
                    "subject, (b) Split the combine into smaller groups.".format(c_str))
        if sec == 'C':
            return "No two adjacent free periods found. Try reducing period count or disable Consecutive flag."
        if sec == 'B':
            return "Both primary and parallel teachers must be free at the same period."
        if sec == 'A':
            return ("All {} classes AND all teachers must be free at the same slot. "
                    "Reduce classes in combine or reduce period counts.".format(len(classes)))
        return "Check that all teachers have free periods available."

    # ── Display helper ──────────────────────────────────────────────────────
    def _get_combined_par_display(self, cn, e):
        s3 = self.step3_data or {}
        cc = e.get('combined_classes', [])
        combined_teacher = ''; combined_subj = ''
        for _t, s3d in s3.items():
            for cb in s3d.get('combines', []):
                if set(cb.get('classes', [])) == set(cc):
                    combined_teacher = _t
                    combined_subj = cb.get('subjects', [''])[0] if cb.get('subjects') else ''
                    break
            if combined_teacher: break
        class_teacher = ''; class_subj = ''
        if combined_subj and cn in self.class_config_data:
            for _s in self.class_config_data[cn].get('subjects', []):
                sname = _s.get('name', '').strip()
                pname = (_s.get('parallel_subject') or '').strip()
                if sname == combined_subj:
                    class_subj = pname; class_teacher = (_s.get('parallel_teacher') or '').strip(); break
                elif pname == combined_subj:
                    class_subj = sname; class_teacher = _s.get('teacher', '').strip(); break
        if not combined_teacher:
            combined_teacher = e.get('teacher', '')
            combined_subj    = e.get('subject', '')
            class_teacher    = e.get('par_teach', '')
            class_subj       = e.get('par_subj', '')
        line1 = "{}\n{}".format(combined_subj, combined_teacher) if combined_teacher else e.get('subject','')
        line2 = "{}\n{}".format(class_subj,    class_teacher)    if class_teacher else ''
        return (line1, line2)

    # ── Excel export ────────────────────────────────────────────────────────
    def _write_excel_bytes(self, mode):
        tt          = self._timetable
        days        = tt['days']
        ppd         = tt['ppd']
        half1       = tt['half1']
        grid        = tt['grid']
        all_classes = tt['all_classes']

        def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
        def _font(bold=False, sz=9, col="000000"):
            return Font(bold=bold, size=sz, color=col.lstrip("#"), name="Arial")
        def _border():
            s = Side(style="thin", color="AAAAAA")
            return Border(left=s, right=s, top=s, bottom=s)
        def _align(h="center", wrap=True):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
        def _sv(val):
            if hasattr(val, 'get'): return val.get()
            return val or ''

        HDR_F = _fill("#2c3e50"); HDR_N = _font(True, 10, "FFFFFF")
        DAY_F = _fill("#34495e"); DAY_N = _font(True,  9, "FFFFFF")
        SUB_F = _fill("#d5e8d4"); COMB_F= _fill("#dae8fc")
        PAR_F = _fill("#ffe6cc"); CPAF  = _fill("#f8cecc")
        FREE_F= _fill("#f5f5f5"); WHT_F = _fill("#FFFFFF")
        SUM_F = _fill("#eaf2ff"); CT_H_F= _fill("#1a5276")
        WRN_F = _fill("#fdebd0")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def _build_tg():
            tg = {}
            for cn in all_classes:
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] if d < len(grid.get(cn,[])) else None
                        if not e: continue
                        etype = e.get('type','normal')
                        cc    = e.get('combined_classes', [])
                        is_cp = bool(cc) and etype == 'combined_parallel'
                        is_c  = bool(cc) and etype == 'combined'
                        def _add(tname, tcls, tsubj, tct):
                            if not tname: return
                            tg.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                            tg[tname][d][p] = {'class': tcls, 'subject': tsubj, 'is_ct': tct}
                        if is_cp:
                            if not cc or cn == cc[0]:
                                _add(e.get('teacher'), '+'.join(cc), e.get('subject',''), False)
                            pt_ = e.get('par_teach','')
                            if pt_ and pt_ not in ('—','?',''):
                                _add(pt_, cn, e.get('par_subj',''), e.get('is_ct',False))
                        elif is_c:
                            if not cc or cn == cc[0]:
                                _add(e.get('teacher'), '+'.join(cc), e.get('subject',''), e.get('is_ct',False))
                        else:
                            _add(e.get('teacher'), cn, e.get('subject',''), e.get('is_ct',False))
                            pt_ = e.get('par_teach','')
                            if pt_ and pt_ not in ('—','?',''):
                                _add(pt_, cn, e.get('par_subj',''), False)
            return tg

        def _ct_map():
            ct = {}
            for cn in all_classes:
                cfg = self.class_config_data.get(cn, {})
                t = _sv(cfg.get('teacher_var', '')).strip()
                if t: ct.setdefault(t, []).append(cn)
            return ct

        if mode == "class":
            for cn in all_classes:
                ws = wb.create_sheet(cn)
                cfg     = self.class_config_data.get(cn, {})
                ct_name = _sv(cfg.get('teacher_var', '')).strip()
                ct_per  = _sv(cfg.get('teacher_period_var', ''))
                hdr_txt = "Class: {}   |   Class Teacher: {}{}".format(
                    cn, ct_name or '—', "   |   CT Period: {}".format(ct_per) if ct_per else '')
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ppd+1)
                c = ws.cell(1, 1, hdr_txt); c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[1].height = 20
                ws.cell(2,1,"Day"); ws.cell(2,1).fill = HDR_F; ws.cell(2,1).font = HDR_N
                ws.cell(2,1).alignment = _align(); ws.cell(2,1).border = _border()
                for p in range(ppd):
                    h = ws.cell(2, p+2, "P{} {}".format(p+1, "①" if p < half1 else "②"))
                    h.fill = HDR_F; h.font = HDR_N; h.alignment = _align(); h.border = _border()
                ws.row_dimensions[2].height = 16
                for d, dname in enumerate(days):
                    r = 3 + d; ws.row_dimensions[r].height = 48
                    dc = ws.cell(r, 1, dname); dc.fill = DAY_F; dc.font = DAY_N
                    dc.alignment = _align(); dc.border = _border()
                    for p in range(ppd):
                        e = grid.get(cn,[[]])[d][p] if d < len(grid.get(cn,[])) else None
                        if e is None:
                            txt = "FREE"; fill = FREE_F
                        else:
                            etype = e.get('type','normal')
                            if etype == 'combined_parallel':
                                l1,l2 = self._get_combined_par_display(cn, e)
                                txt = "{}\n{}".format(l1,l2); fill = CPAF
                            elif etype == 'parallel':
                                txt = "{}/{}\n{}/{}".format(e['subject'],e.get('par_subj',''),e['teacher'],e.get('par_teach','')); fill = PAR_F
                            elif etype == 'combined':
                                cc2 = e.get('combined_classes',[])
                                mark = " ★" if e.get('is_ct') else ""
                                txt = "{}{}[{}]\n{}".format(e['subject'],mark,'+'.join(cc2),e['teacher']); fill = COMB_F
                            else:
                                mark = " ★" if e.get('is_ct') else ""
                                txt = "{}{}\n{}".format(e['subject'],mark,e['teacher'])
                                fill = SUB_F if e.get('is_ct') else WHT_F
                        c = ws.cell(r, p+2, txt); c.fill = fill; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=8)
                sr = 3 + len(days) + 1
                ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=ppd+1)
                c = ws.cell(sr, 1, "Summary — {}".format(cn))
                c.fill = HDR_F; c.font = HDR_N; c.alignment = _align("left"); c.border = _border()
                smry = defaultdict(int)
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn,[[]])[d][p] if d < len(grid.get(cn,[])) else None
                        if not e: continue
                        etype = e.get('type','normal')
                        if etype == 'combined_parallel':
                            l1,l2 = self._get_combined_par_display(cn, e)
                            for ln in (l1,l2):
                                pts = ln.split('\n')
                                smry[(pts[0].strip(), pts[1].strip() if len(pts)>1 else '')] += 1
                        elif etype == 'parallel':
                            smry[(e['subject'],e['teacher'])] += 1
                            smry[(e.get('par_subj',''),e.get('par_teach',''))] += 1
                        else:
                            smry[(e['subject'],e['teacher'])] += 1
                hdr_r = sr + 1
                for col, txt in enumerate(["Subject","Teacher","Periods/Week"], 1):
                    c = ws.cell(hdr_r, col, txt); c.fill = HDR_F; c.font = HDR_N
                    c.alignment = _align(); c.border = _border()
                for i, ((subj, teach), cnt) in enumerate(sorted(smry.items())):
                    row = hdr_r + 1 + i
                    for col, val in enumerate([subj, teach, cnt], 1):
                        c = ws.cell(row, col, val)
                        c.fill = SUM_F if i%2==0 else WHT_F; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=9)
                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        elif mode == "teacher":
            tg = _build_tg(); ct_mp = _ct_map()
            for teacher in sorted(tg.keys()):
                ws = wb.create_sheet(teacher[:31]); tdata = tg[teacher]
                ctc = ct_mp.get(teacher, [])
                hdr_txt = "Teacher: {}   |   Class Teacher of: {}".format(
                    teacher, ', '.join(ctc) if ctc else '—')
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ppd+1)
                c = ws.cell(1,1,hdr_txt); c.fill = CT_H_F; c.font = _font(True,11,"FFFFFF")
                c.alignment = _align(); c.border = _border(); ws.row_dimensions[1].height = 20
                ws.cell(2,1,"Day"); ws.cell(2,1).fill=HDR_F; ws.cell(2,1).font=HDR_N
                ws.cell(2,1).alignment=_align(); ws.cell(2,1).border=_border()
                for p in range(ppd):
                    h = ws.cell(2,p+2,"P{} {}".format(p+1,"①" if p<half1 else "②"))
                    h.fill=HDR_F; h.font=HDR_N; h.alignment=_align(); h.border=_border()
                ws.row_dimensions[2].height = 16
                for d, dname in enumerate(days):
                    r = 3+d; ws.row_dimensions[r].height = 48
                    dc = ws.cell(r,1,dname); dc.fill=DAY_F; dc.font=DAY_N
                    dc.alignment=_align(); dc.border=_border()
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e is None:
                            txt="FREE"; fill=FREE_F
                        else:
                            txt = "{}\n{}".format(e['class'],e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        c = ws.cell(r,p+2,txt); c.fill=fill; c.alignment=_align()
                        c.border=_border(); c.font=_font(sz=8)
                sr = 3+len(days)+1
                ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=ppd+1)
                c = ws.cell(sr,1,"Summary — {}".format(teacher))
                c.fill=HDR_F; c.font=HDR_N; c.alignment=_align("left"); c.border=_border()
                smry = defaultdict(lambda: defaultdict(int)); total = 0
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d<len(tdata) else None
                        if e: smry[e['class']][e['subject']] += 1; total += 1
                hdr_r = sr+1
                for col, txt in enumerate(["Class","Subject","Periods/Week"],1):
                    c = ws.cell(hdr_r,col,txt); c.fill=HDR_F; c.font=HDR_N
                    c.alignment=_align(); c.border=_border()
                row = hdr_r+1
                for cls in sorted(smry.keys()):
                    for subj, cnt in sorted(smry[cls].items()):
                        for col, val in enumerate([cls,subj,cnt],1):
                            c = ws.cell(row,col,val)
                            c.fill=SUM_F if row%2==0 else WHT_F; c.alignment=_align()
                            c.border=_border(); c.font=_font(sz=9)
                        row += 1
                for col, val in enumerate(["","TOTAL",total],1):
                    c=ws.cell(row,col,val); c.fill=_fill("#d4e6f1"); c.font=_font(True,9)
                    c.alignment=_align(); c.border=_border()
                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        elif mode == "ct_list":
            ws = wb.create_sheet("Class Teacher List")
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = "Class Teacher List"
            c.fill=HDR_F; c.font=_font(True,13,"FFFFFF"); c.alignment=_align(); c.border=_border()
            ws.row_dimensions[1].height = 22
            for col, txt in enumerate(["Class","Class Teacher","CT Period"],1):
                c = ws.cell(2,col,txt); c.fill=DAY_F; c.font=DAY_N
                c.alignment=_align(); c.border=_border()
            for i, cn in enumerate(all_classes):
                cfg = self.class_config_data.get(cn,{})
                ct_name = _sv(cfg.get('teacher_var','')).strip() or '—'
                ct_per  = _sv(cfg.get('teacher_period_var','')) or '—'
                row = 3+i
                for col, val in enumerate([cn,ct_name,ct_per],1):
                    c = ws.cell(row,col,val)
                    c.fill=SUM_F if i%2==0 else WHT_F; c.alignment=_align()
                    c.border=_border(); c.font=_font(sz=10)
            ws.column_dimensions["A"].width=14; ws.column_dimensions["B"].width=28
            ws.column_dimensions["C"].width=12

        elif mode == "workload":
            tg = _build_tg(); ct_mp = _ct_map()
            ws = wb.create_sheet("Teacher Workload")
            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = "Teacher Workload List"
            c.fill=HDR_F; c.font=_font(True,13,"FFFFFF"); c.alignment=_align(); c.border=_border()
            ws.row_dimensions[1].height = 22
            for col, txt in enumerate(["Teacher","Subject","Class","Periods/Week","Total Periods"],1):
                c = ws.cell(2,col,txt); c.fill=DAY_F; c.font=DAY_N; c.alignment=_align(); c.border=_border()
            row = 3; grand_total = 0
            for teacher in sorted(tg.keys()):
                tdata = tg[teacher]
                smry = defaultdict(lambda: defaultdict(int))
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d<len(tdata) else None
                        if e: smry[e['subject']][e['class']] += 1
                total = sum(c2 for cd in smry.values() for c2 in cd.values())
                grand_total += total; ctc = ct_mp.get(teacher,[])
                start_row = row
                for si, subj in enumerate(sorted(smry.keys())):
                    for cls, cnt in sorted(smry[subj].items()):
                        fill = SUM_F if row%2==0 else WHT_F
                        c = ws.cell(row,1, teacher if row==start_row else "")
                        c.fill=WRN_F if ctc else fill
                        c.font=_font(True if row==start_row else False, 9)
                        c.alignment=_align(); c.border=_border()
                        for col, val in enumerate([subj,cls,cnt],2):
                            c2=ws.cell(row,col,val); c2.fill=fill
                            c2.alignment=_align(); c2.border=_border(); c2.font=_font(sz=9)
                        c5=ws.cell(row,5,total if row==start_row else "")
                        c5.fill=_fill("#d4e6f1") if row==start_row else fill
                        c5.font=_font(True if row==start_row else False,9)
                        c5.alignment=_align(); c5.border=_border()
                        row += 1
                span = row - start_row
                if span > 1:
                    ws.merge_cells(start_row=start_row,start_column=1,end_row=row-1,end_column=1)
            for col, val in enumerate(["","","","GRAND TOTAL",grand_total],1):
                c=ws.cell(row,col,val); c.fill=HDR_F; c.font=_font(True,10,"FFFFFF")
                c.alignment=_align(); c.border=_border()
            ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=22
            ws.column_dimensions["C"].width=16; ws.column_dimensions["D"].width=16
            ws.column_dimensions["E"].width=16

        elif mode == "one_sheet":
            tg = _build_tg()
            ws = wb.create_sheet("Teacherwise Timetable")
            ws.row_dimensions[1].height = 18
            ws.cell(1,1,"Teacher"); ws.cell(1,2,"Day")
            for col in (1,2):
                ws.cell(1,col).fill=HDR_F; ws.cell(1,col).font=HDR_N
                ws.cell(1,col).alignment=_align(); ws.cell(1,col).border=_border()
            for p in range(ppd):
                c = ws.cell(1,p+3,str(p+1))
                c.fill=HDR_F; c.font=HDR_N; c.alignment=_align(); c.border=_border()
            row = 2
            for teacher in sorted(tg.keys()):
                tdata = tg[teacher]; t_start = row
                for d, dname in enumerate(days):
                    c = ws.cell(row,1, teacher if d==0 else "")
                    c.fill=WRN_F; c.alignment=_align()
                    c.font=_font(True if d==0 else False,9); c.border=_border()
                    c2 = ws.cell(row,2,dname); c2.fill=DAY_F; c2.font=DAY_N
                    c2.alignment=_align(); c2.border=_border()
                    for p in range(ppd):
                        e = tdata[d][p] if d<len(tdata) else None
                        if e:
                            txt  = "{}/{}".format(e['class'],e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        else:
                            txt = ""; fill = FREE_F
                        c3 = ws.cell(row,p+3,txt); c3.fill=fill; c3.alignment=_align()
                        c3.border=_border(); c3.font=_font(sz=8)
                    row += 1
                if len(days) > 1:
                    ws.merge_cells(start_row=t_start,start_column=1,end_row=row-1,end_column=1)
            ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=10
            for p in range(ppd):
                ws.column_dimensions[get_column_letter(p+3)].width=18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


# Create the global app proxy
app = App()


###############################################################################
#  UTILITY FUNCTIONS
###############################################################################

def _all_classes():
    cfg = ss.configuration
    if not cfg: return []
    result = []
    for cls in range(6, 13):
        for si in range(cfg.get('classes', {}).get(cls, 0)):
            result.append("{}{}".format(cls, chr(65 + si)))
    return result


def _required_periods():
    cfg = ss.configuration
    return cfg.get('periods_per_day', 0) * cfg.get('working_days', 0)


def _day_names():
    wdays = ss.configuration.get('working_days', 6)
    return ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][:wdays]


def _ensure_class_data(cn):
    if cn not in ss.class_config_data:
        ss.class_config_data[cn] = {
            'teacher_var': Var(''),
            'teacher_period_var': Var(1),
            'subjects': [],
        }


def _period_count(cn):
    cd = ss.class_config_data.get(cn, {})
    return sum(s['periods'] for s in cd.get('subjects', []))


###############################################################################
#  STEP 1: School Configuration
###############################################################################

def render_step1():
    st.markdown('<div class="step-header">📋 Step 1: School Configuration</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-sub">Configure the basic schedule, upload teacher list, and set class sections.</div>', unsafe_allow_html=True)

    # ── Config save / load ──────────────────────────────────────────────────
    with st.expander("💾 Save / Load Configuration", expanded=False):
        col_save, col_load = st.columns([1,1])
        with col_save:
            st.markdown("**Save current configuration**")
            cfg_name = st.text_input("Config name", value="Config_{}".format(datetime.now().strftime("%Y%m%d_%H%M%S")), key="s1_save_name")
            if st.button("💾 Download Config as JSON", key="s1_download_btn"):
                if not ss.s1_teacher_names:
                    st.error("Upload teacher file first!")
                else:
                    data = {
                        "periods_per_day": ss.s1_ppd,
                        "working_days": ss.s1_wdays,
                        "periods_first_half": ss.s1_fh,
                        "periods_second_half": ss.s1_sh,
                        "teacher_names": ss.s1_teacher_names,
                        "classes": {str(k): v for k, v in ss.s1_class_sections.items()},
                        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "step": 1,
                    }
                    fname = "{}.json".format(cfg_name.strip() or "config")
                    st.download_button("⬇️ Click to Save", data=json.dumps(data, indent=2),
                                       file_name=fname, mime="application/json", key="s1_dl_btn2")
        with col_load:
            st.markdown("**Upload a saved configuration**")
            uploaded = st.file_uploader("Upload JSON config", type=["json"], key="s1_upload_cfg")
            if uploaded:
                try:
                    d = json.load(uploaded)
                    if d.get('step') != 1 and 'periods_per_day' not in d:
                        st.error("This does not appear to be a Step 1 config.")
                    else:
                        ss.s1_ppd   = d.get("periods_per_day", 7)
                        ss.s1_wdays = d.get("working_days", 6)
                        ss.s1_fh    = d.get("periods_first_half", 4)
                        ss.s1_sh    = d.get("periods_second_half", 3)
                        ss.s1_teacher_names = d.get("teacher_names", [])
                        cl = d.get("classes", {})
                        ss.s1_class_sections = {int(k): v for k, v in cl.items()} if cl else ss.s1_class_sections
                        st.success("✓ Configuration loaded!")
                        st.rerun()
                except Exception as e:
                    st.error("Error loading config: {}".format(e))

    # ── Section 1: Basic Schedule ───────────────────────────────────────────
    st.markdown("### 1. Basic Schedule Configuration")
    c1, c2 = st.columns(2)
    with c1:
        ss.s1_ppd = st.number_input("Periods per day", min_value=1, max_value=15, value=ss.s1_ppd, key="inp_ppd")
    with c2:
        ss.s1_wdays = st.number_input("Working days in week", min_value=1, max_value=7, value=ss.s1_wdays, key="inp_wdays")

    # ── Section 2: Period Division ──────────────────────────────────────────
    st.markdown("### 2. Division of Periods")
    c3, c4 = st.columns(2)
    with c3:
        ss.s1_fh = st.number_input("Periods in first half", min_value=1, max_value=15, value=ss.s1_fh, key="inp_fh")
    with c4:
        ss.s1_sh = st.number_input("Periods in second half", min_value=1, max_value=15, value=ss.s1_sh, key="inp_sh")

    fh_sh = ss.s1_fh + ss.s1_sh
    if fh_sh == ss.s1_ppd:
        st.success("✓ Valid: {} + {} = {}".format(ss.s1_fh, ss.s1_sh, ss.s1_ppd))
    else:
        st.error("✗ Invalid: {} + {} = {}, need {}!".format(ss.s1_fh, ss.s1_sh, fh_sh, ss.s1_ppd))

    # ── Section 3: Teachers ─────────────────────────────────────────────────
    st.markdown("### 3. Teachers (Upload Excel)")
    with st.expander("📋 Expected Excel Format", expanded=False):
        st.markdown("""
**Column A only:**
| A |
|---|
| Teacher Name |
| Mr. Sharma |
| Mrs. Patel |
| ... |

✓ One name per row  ✓ No duplicates  ✓ No data in other columns
        """)

    teacher_file = st.file_uploader("Upload Teacher Excel (.xlsx / .xls)", type=["xlsx","xls"], key="teacher_file_up")
    if teacher_file:
        try:
            wb = openpyxl.load_workbook(teacher_file)
            ws = wb.active
            max_col = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for ci, cell in enumerate(row, 1):
                    if cell.value is not None:
                        max_col = max(max_col, ci)
            if max_col > 1:
                st.error("❌ WRONG FORMAT! Found data in {} columns. Use ONLY Column A!".format(max_col))
            else:
                names = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                    v = row[0].value
                    if v:
                        n = str(v).strip()
                        if n and n.lower() != "teacher name":
                            names.append(n)
                if not names:
                    st.error("❌ No teacher names found!")
                else:
                    seen, dups = set(), []
                    for n in names:
                        if n in seen: dups.append(n)
                        seen.add(n)
                    if dups:
                        st.error("❌ DUPLICATE NAMES:\n" + "\n".join(set(dups)))
                    else:
                        names.sort()
                        ss.s1_teacher_names = names
                        st.success("✓ Loaded {} teachers (sorted A→Z)".format(len(names)))
        except Exception as e:
            st.error("Error reading file: {}".format(e))

    if ss.s1_teacher_names:
        st.info("✓ {} teachers loaded: {}{}".format(
            len(ss.s1_teacher_names),
            ", ".join(ss.s1_teacher_names[:5]),
            "..." if len(ss.s1_teacher_names) > 5 else ""))

    # ── Section 4: Classes ──────────────────────────────────────────────────
    st.markdown("### 4. Classes (6–12) — Number of Sections")
    cols = st.columns(7)
    for i, cls in enumerate(range(6, 13)):
        with cols[i]:
            ss.s1_class_sections[cls] = st.number_input(
                "Class {}".format(cls), min_value=0, max_value=26,
                value=ss.s1_class_sections.get(cls, 4), key="cls_{}".format(cls))

    # ── Buttons ─────────────────────────────────────────────────────────────
    st.markdown("---")
    col_btn1, col_btn2 = st.columns([1,5])
    with col_btn1:
        if st.button("⟲ Reset", key="s1_reset"):
            ss.s1_ppd = 7; ss.s1_wdays = 6; ss.s1_fh = 4; ss.s1_sh = 3
            ss.s1_teacher_names = []
            ss.s1_class_sections = {c: 4 for c in range(6, 13)}
            st.rerun()
    with col_btn2:
        if st.button("✓ Continue to Step 2 →", type="primary", key="s1_continue"):
            # Validate
            if ss.s1_fh + ss.s1_sh != ss.s1_ppd:
                st.error("Period halves don't match total periods!")
            elif not ss.s1_teacher_names:
                st.error("Please upload a teacher file!")
            elif all(ss.s1_class_sections.get(c, 0) == 0 for c in range(6, 13)):
                st.error("At least one class must have sections > 0!")
            else:
                ss.configuration = {
                    "periods_per_day":      ss.s1_ppd,
                    "working_days":         ss.s1_wdays,
                    "periods_first_half":   ss.s1_fh,
                    "periods_second_half":  ss.s1_sh,
                    "teacher_names":        ss.s1_teacher_names,
                    "classes":              dict(ss.s1_class_sections),
                }
                # Initialise class_config_data for new classes
                for cn in _all_classes():
                    _ensure_class_data(cn)
                ss.step = 2
                st.rerun()


###############################################################################
#  STEP 2: Class Configuration
###############################################################################

def render_step2():
    cfg  = ss.configuration
    ppd  = cfg['periods_per_day']
    wdays = cfg['working_days']
    required = ppd * wdays

    st.markdown('<div class="step-header">👨‍🏫 Step 2: Configure Each Class</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-sub">For each class, set the class teacher and add all subjects with their teachers, period counts, and constraints.</div>', unsafe_allow_html=True)

    # ── Save / Load assignments ─────────────────────────────────────────────
    with st.expander("💾 Save / Load Assignments", expanded=False):
        col_s, col_l = st.columns(2)
        with col_s:
            a_name = st.text_input("Assignment name", value="Assignments_{}".format(
                datetime.now().strftime("%Y%m%d_%H%M%S")), key="s2_save_name")
            if st.button("💾 Download Assignments JSON", key="s2_dl_cfg"):
                data = {
                    "assignments": {
                        cn: {
                            "teacher": ss.class_config_data[cn]['teacher_var'].get(),
                            "teacher_period": ss.class_config_data[cn]['teacher_period_var'].get(),
                            "subjects": ss.class_config_data[cn]['subjects'],
                        }
                        for cn in _all_classes() if cn in ss.class_config_data
                    },
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "step": 2,
                }
                fname = "{}.json".format(a_name.strip() or "assignments")
                st.download_button("⬇️ Click to Save", data=json.dumps(data, indent=2),
                                   file_name=fname, mime="application/json", key="s2_dl_btn2")
        with col_l:
            a_up = st.file_uploader("Upload assignments JSON", type=["json"], key="s2_upload_asgn")
            if a_up:
                try:
                    d = json.load(a_up)
                    if 'assignments' not in d:
                        st.error("This does not appear to be a Step 2 assignments file.")
                    else:
                        for cn, cd_saved in d['assignments'].items():
                            if cn in ss.class_config_data:
                                ss.class_config_data[cn]['teacher_var'].set(cd_saved.get('teacher', ''))
                                ss.class_config_data[cn]['teacher_period_var'].set(cd_saved.get('teacher_period', 1))
                                ss.class_config_data[cn]['subjects'] = cd_saved.get('subjects', [])
                        ss._s2_cfg_upload_status = "✓ Assignments loaded!"
                        st.success("✓ Assignments loaded!")
                        st.rerun()
                except Exception as e:
                    st.error("Error: {}".format(e))

    # ── Nav buttons ──────────────────────────────────────────────────────────
    c_nav1, c_nav2 = st.columns([1, 2])
    with c_nav1:
        if st.button("← Back to Step 1", key="s2_back"):
            ss.step = 1; st.rerun()
    with c_nav2:
        if st.button("✓ Validate & Complete →", type="primary", key="s2_validate"):
            report = app._validate_and_complete()
            ss.validation_report = report
            ss.show_validation_report = True
            st.rerun()

    if ss.show_validation_report and ss.validation_report:
        render_validation_report(ss.validation_report)
        return

    # ── Class tabs ──────────────────────────────────────────────────────────
    classes = _all_classes()
    if not classes:
        st.warning("No classes configured. Go back to Step 1 and set class sections.")
        return

    tabs = st.tabs(classes)
    for tab, cn in zip(tabs, classes):
        with tab:
            render_class_tab(cn, ppd, wdays, required, cfg)


def render_class_tab(cn, ppd, wdays, required, cfg):
    _ensure_class_data(cn)
    cd = ss.class_config_data[cn]
    teacher_names = sorted(cfg.get('teacher_names', []))
    day_names = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][:wdays]

    total_periods = _period_count(cn)
    counter_color = "period-counter-ok" if total_periods == required else "period-counter-err"
    st.markdown(
        '<span class="{}">{}/{} periods assigned</span>'.format(counter_color, total_periods, required),
        unsafe_allow_html=True)

    # ── Class Teacher row ────────────────────────────────────────────────────
    st.markdown("**Class Teacher**")
    ct_col1, ct_col2 = st.columns([3, 1])
    with ct_col1:
        ct_options = [""] + teacher_names
        ct_current = cd['teacher_var'].get()
        ct_idx = ct_options.index(ct_current) if ct_current in ct_options else 0
        ct_sel = st.selectbox("Class Teacher", options=ct_options, index=ct_idx,
                              key="ct_sel_{}".format(cn))
        cd['teacher_var'].set(ct_sel)
    with ct_col2:
        ct_per = st.number_input("CT Period", min_value=1, max_value=ppd,
                                  value=cd['teacher_period_var'].get(), key="ct_per_{}".format(cn))
        cd['teacher_period_var'].set(ct_per)

    st.markdown("---")
    left_col, right_col = st.columns([2, 3])

    # ── Left: Add/Edit Subject Form ──────────────────────────────────────────
    with left_col:
        editing_idx = ss.get('step2_editing_{}_{}'.format(cn, 'idx'), None)
        is_editing = editing_idx is not None

        st.markdown("**{} Subject**".format("✏️ Edit" if is_editing else "➕ Add"))

        # Pre-fill form if editing
        subj_data = {}
        if is_editing and 0 <= editing_idx < len(cd['subjects']):
            subj_data = cd['subjects'][editing_idx]

        form_name = st.text_input("Subject Name", value=subj_data.get('name',''),
                                   key="fn_{}_{}".format(cn, editing_idx))
        s_opts = [""] + teacher_names
        s_cur = subj_data.get('teacher','')
        s_idx = s_opts.index(s_cur) if s_cur in s_opts else 0
        form_teacher = st.selectbox("Teacher", options=s_opts, index=s_idx,
                                     key="ft_{}_{}".format(cn, editing_idx))
        form_periods = st.number_input("Periods/Week", min_value=1, max_value=ppd*wdays,
                                        value=subj_data.get('periods',1),
                                        key="fp_{}_{}".format(cn, editing_idx))

        st.markdown("**Period Preference** (leave unchecked = any period)")
        p_pref_cols = st.columns(min(ppd, 8))
        p_pref_vals = {}
        existing_pref = set(subj_data.get('periods_pref', []))
        for pi in range(ppd):
            with p_pref_cols[pi % len(p_pref_cols)]:
                p_pref_vals[pi+1] = st.checkbox(str(pi+1), value=(pi+1) in existing_pref,
                                                  key="pp_{}_{}_{}_{}".format(cn, pi, editing_idx, 'v'))

        st.markdown("**Day Preference** (leave unchecked = any day)")
        d_pref_cols = st.columns(len(day_names))
        d_pref_vals = {}
        existing_dpref = set(subj_data.get('days_pref', []))
        for di, day in enumerate(day_names):
            with d_pref_cols[di]:
                d_pref_vals[day] = st.checkbox(day, value=day in existing_dpref,
                                                key="dp_{}_{}_{}".format(cn, day, editing_idx))

        c_rad1, c_rad2 = st.columns(2)
        with c_rad1:
            form_consec = st.radio("Consecutive", ["No", "Yes"],
                                    index=1 if subj_data.get('consecutive','No')=='Yes' else 0,
                                    key="fc_{}_{}".format(cn, editing_idx), horizontal=True)
        with c_rad2:
            form_parallel = st.checkbox("Parallel to another subject?",
                                         value=subj_data.get('parallel', False),
                                         key="fpar_{}_{}".format(cn, editing_idx))

        par_subj = ""; par_teach = ""
        if form_parallel:
            par_subj = st.text_input("Parallel Subject Name",
                                      value=subj_data.get('parallel_subject',''),
                                      key="fps_{}_{}".format(cn, editing_idx))
            pt_opts = [""] + teacher_names
            pt_cur  = subj_data.get('parallel_teacher','')
            pt_idx  = pt_opts.index(pt_cur) if pt_cur in pt_opts else 0
            par_teach = st.selectbox("Parallel Teacher", options=pt_opts, index=pt_idx,
                                      key="fpt_{}_{}".format(cn, editing_idx))

        err_key = "form_err_{}_{}".format(cn, editing_idx)

        btn_col_a, btn_col_b = st.columns(2)
        with btn_col_a:
            btn_label = "✏️ Update Subject" if is_editing else "➕ Add Subject"
            if st.button(btn_label, key="fadd_{}_{}".format(cn, editing_idx)):
                err = None
                if not form_name.strip():
                    err = "⚠ Subject name is required!"
                elif not form_teacher:
                    err = "⚠ Teacher is required!"
                else:
                    entry = {
                        'name':            form_name.strip(),
                        'teacher':         form_teacher,
                        'periods':         form_periods,
                        'periods_pref':    [p for p, v in p_pref_vals.items() if v],
                        'days_pref':       [d for d, v in d_pref_vals.items() if v],
                        'consecutive':     form_consec,
                        'parallel':        form_parallel,
                        'parallel_subject': par_subj.strip() if form_parallel else '',
                        'parallel_teacher': par_teach if form_parallel else '',
                    }
                    if is_editing:
                        cd['subjects'][editing_idx] = entry
                        st.session_state['step2_editing_{}_idx'.format(cn)] = None
                    else:
                        cd['subjects'].append(entry)
                    st.rerun()
                if err:
                    st.error(err)

        with btn_col_b:
            if is_editing and st.button("✕ Cancel Edit", key="fcancel_{}_{}".format(cn, editing_idx)):
                st.session_state['step2_editing_{}_idx'.format(cn)] = None
                st.rerun()
            elif not is_editing and st.button("✕ Clear", key="fclear_{}".format(cn)):
                st.rerun()

    # ── Right: Subject List ───────────────────────────────────────────────────
    with right_col:
        st.markdown("**📚 Subjects Added**")
        subjects = cd.get('subjects', [])
        if not subjects:
            st.info("No subjects added yet.")
        else:
            for i, s in enumerate(subjects):
                with st.container():
                    pref_str = ""
                    if s.get('periods_pref'):
                        pref_str += " | Periods: {}".format(s['periods_pref'])
                    if s.get('days_pref'):
                        pref_str += " | Days: {}".format(s['days_pref'])
                    if s.get('consecutive','No') == 'Yes':
                        pref_str += " | Consec✓"
                    par_str = ""
                    if s.get('parallel'):
                        par_str = " | ∥ {}/{}".format(
                            s.get('parallel_subject','?'), s.get('parallel_teacher','?'))

                    label = "**{}**  ({} periods)  —  {}{}{}".format(
                        s['name'], s['periods'], s['teacher'], pref_str, par_str)
                    srow1, srow2, srow3 = st.columns([5, 1, 1])
                    with srow1:
                        st.markdown(label, unsafe_allow_html=False)
                    with srow2:
                        if st.button("✏️", key="edit_{}_{}_{}".format(cn, i, 's'), help="Edit"):
                            st.session_state['step2_editing_{}_idx'.format(cn)] = i
                            st.rerun()
                    with srow3:
                        if st.button("🗑", key="del_{}_{}_{}".format(cn, i, 's'), help="Delete"):
                            cd['subjects'].pop(i)
                            if ss.get('step2_editing_{}_idx'.format(cn)) == i:
                                st.session_state['step2_editing_{}_idx'.format(cn)] = None
                            st.rerun()


def render_validation_report(report):
    period_ok     = report['period_ok']
    period_errors = report['period_errors']
    hard_conflicts = report['hard_conflicts']
    within_class   = report['within_class_conflicts']
    required = report['required']

    any_error = bool(period_errors or hard_conflicts or within_class)

    if any_error:
        st.error("❌ VALIDATION FAILED — Fix the errors shown below, then validate again.")
    else:
        st.success("✅ All checks passed!")

    col_a, col_b, col_c, col_d = st.columns(4)
    with col_a: st.metric("Period Errors", len(period_errors))
    with col_b: st.metric("Teacher Conflicts", len(hard_conflicts))
    with col_c: st.metric("Within-Class Conflicts", len(within_class))
    with col_d: st.metric("Classes OK", len(period_ok))

    with st.expander("① Period Count Details", expanded=bool(period_errors)):
        st.write("Required per class: **{} periods/week**".format(required))
        if period_errors:
            for cn, em in sorted(period_errors):
                st.error("❌  Class {}  {}".format(cn, em))
        if period_ok:
            st.markdown("**Classes with correct period count:**")
            for cn, pm in sorted(period_ok):
                st.write("✓ {}  {}".format(cn, pm))

    if hard_conflicts:
        with st.expander("② Teacher Conflicts ({})".format(len(hard_conflicts)), expanded=True):
            for i, c in enumerate(hard_conflicts, 1):
                st.error("[{}] Teacher: {}".format(i, c['teacher']))
                st.write("Assignment A: {}".format(c['slot_a']))
                st.write("Assignment B: {}".format(c['slot_b']))
                st.write("⚠ {}".format(c['reason']))
                st.write("FIX: Change the period/day preference for one of the two assignments.")
                st.markdown("---")

    if within_class:
        with st.expander("③ Within-Class Slot Conflicts ({})".format(len(within_class)), expanded=True):
            for i, c in enumerate(within_class, 1):
                st.error("[{}] Class: {}  Day: {}".format(i, c['class'], c['day']))
                st.write("Item A: {}".format(c['item_a']))
                if c.get('item_b'): st.write("Item B: {}".format(c['item_b']))
                st.write("⚠ {}".format(c['reason']))
                st.write("FIX: Adjust period or day preference so subjects don't compete for the same slot.")
                st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Back to Editing", key="back_from_validation"):
            ss.show_validation_report = False
            st.rerun()
    with col2:
        if not any_error:
            if st.button("▶ Proceed to Step 3 →", type="primary", key="goto_s3"):
                ss.show_validation_report = False
                ss.step3_data = {}
                ss.step3_unavailability = {}
                ss._step3_selected_teacher = None
                ss.step = 3
                st.rerun()


###############################################################################
#  STEP 3: Teacher Manager
###############################################################################

def render_step3():
    cfg  = ss.configuration
    ppd  = cfg['periods_per_day']
    wdays = cfg['working_days']
    total_week  = ppd * wdays
    max_allowed = (ppd - 2) * wdays

    st.markdown('<div class="step-header">📋 Step 3: Manage Class Combines</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="step-sub">Overload threshold: {} periods/week  ({} days × ({}-2) periods). '
        'All teachers shown — overloaded ones highlighted in red.</div>'.format(max_allowed, wdays, ppd),
        unsafe_allow_html=True)

    # Compute workload
    teacher_workload = app._compute_teacher_workload()
    st.session_state['_step3_teacher_wl'] = teacher_workload
    overloaded_set = {t for t, info in teacher_workload.items() if info['total'] > max_allowed}
    st.session_state['_step3_overloaded'] = overloaded_set
    st.session_state['_step3_max_allowed'] = max_allowed
    st.session_state['_step3_total_week']  = total_week

    # ── Save / Load ──────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Step 3 Configuration", expanded=False):
        col_s3, col_l3 = st.columns(2)
        with col_s3:
            s3_name = st.text_input("Config name",
                value="Step3_{}".format(datetime.now().strftime("%Y%m%d_%H%M%S")),
                key="s3_save_name")
            if st.button("💾 Download Step 3 Config", key="s3_dl_btn"):
                data = {
                    'step3_data': {
                        t: {'skipped': v['skipped'], 'combines': v['combines']}
                        for t, v in (ss.step3_data or {}).items()
                    },
                    'step3_unavailability': {
                        t: {'days': list(v['days']), 'periods': list(v['periods'])}
                        for t, v in (ss.step3_unavailability or {}).items()
                    },
                    'saved_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'step': 3,
                }
                fname = "{}.json".format(s3_name.strip() or "step3")
                st.download_button("⬇️ Click to Save", data=json.dumps(data, indent=2),
                                   file_name=fname, mime="application/json", key="s3_dl_btn2")
        with col_l3:
            s3_up = st.file_uploader("Upload Step 3 JSON", type=["json"], key="s3_upload")
            if s3_up:
                try:
                    d = json.load(s3_up)
                    if d.get('step') != 3:
                        st.error("This does not appear to be a Step 3 config.")
                    else:
                        ss.step3_data = d.get('step3_data', {})
                        ss.step3_unavailability = d.get('step3_unavailability', {})
                        st.success("✓ Step 3 config loaded!")
                        st.rerun()
                except Exception as e:
                    st.error("Error: {}".format(e))

    # ── Navigation ───────────────────────────────────────────────────────────
    nav1, nav2, nav3 = st.columns(3)
    with nav1:
        if st.button("◄ Back to Step 2", key="s3_back"):
            ss.step = 2; st.rerun()
    with nav2:
        if st.button("✓ Validate Step 3", key="s3_validate"):
            _validate_step3_ui(overloaded_set, max_allowed)
    with nav3:
        if st.button("▶ Proceed to Step 4 →", type="primary", key="s3_proceed"):
            ss.step = 4; ss.s4_substep = 'main'; ss._gen_stage = 0
            ss._gen = None; ss._timetable = None
            ss._last_allocation = {}; ss._last_all_rows = []
            ss._relaxed_consec_keys = set(); ss._relaxed_main_keys = set()
            st.rerun()

    st.markdown("---")

    # ── Two-column layout ────────────────────────────────────────────────────
    left_col, right_col = st.columns([1, 2])

    with left_col:
        st.markdown("### All Teachers")
        if not teacher_workload:
            st.warning("No teacher assignments found. Make sure teachers are assigned to subjects in Step 2.")
        else:
            overloaded_teachers = sorted([t for t in teacher_workload if t in overloaded_set])
            normal_teachers     = sorted([t for t in teacher_workload if t not in overloaded_set])

            if overloaded_teachers:
                st.markdown("**⚠ Overloaded (>{}/wk)**".format(max_allowed))
                for teacher in overloaded_teachers:
                    render_teacher_card(teacher, teacher_workload, overloaded_set, max_allowed, is_overloaded=True)
            if normal_teachers:
                st.markdown("**✓ Within limit (≤{}/wk)**".format(max_allowed))
                for teacher in normal_teachers:
                    render_teacher_card(teacher, teacher_workload, overloaded_set, max_allowed, is_overloaded=False)

    with right_col:
        sel = ss._step3_selected_teacher
        if sel and sel in teacher_workload:
            render_teacher_detail(sel, teacher_workload, overloaded_set, max_allowed)
        else:
            st.markdown("### Assignments & Combines")
            st.info("👈 Select a teacher from the left to view assignments and create combines.")

    # ── Unavailability ────────────────────────────────────────────────────────
    st.markdown("---")
    render_unavailability_ui(cfg, ppd, wdays)


def render_teacher_card(teacher, teacher_workload, overloaded_set, max_allowed, is_overloaded):
    info      = teacher_workload[teacher]
    effective = app._effective_total(teacher)
    s3d       = (ss.step3_data or {}).get(teacher, {})
    skipped   = s3d.get('skipped', False)
    n_comb    = len(s3d.get('combines', []))

    if skipped:
        cls = "card-skipped"; badge = "SKIPPED"; badge_cls = "badge-skipped"
    elif is_overloaded and effective > max_allowed:
        cls = "card-overloaded"; badge = "OVERLOADED"; badge_cls = "badge-overloaded"
    elif is_overloaded and effective <= max_allowed:
        cls = "card-resolved"; badge = "{} combine{}".format(n_comb, 's' if n_comb>1 else ''); badge_cls = "badge-ok"
    else:
        cls = "card-ok"; badge = "OK" + (" | {} combine{}".format(n_comb,'s' if n_comb>1 else '') if n_comb else ""); badge_cls = "badge-ok"

    stat_text = "Assigned: {}".format(info['total'])
    if effective != info['total']:
        stat_text += " → Effective: {}".format(effective)
    stat_text += " / Max: {}".format(max_allowed)

    st.markdown(
        '<div class="{}">'
        '<strong>{}</strong> <span class="{}">{}</span><br>'
        '<small>{}</small></div>'.format(cls, teacher, badge_cls, badge, stat_text),
        unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Edit/Combine", key="s3_edit_{}".format(teacher)):
            ss._step3_selected_teacher = teacher
            st.rerun()
    with c2:
        skip_lbl = "Un-skip" if skipped else "Skip"
        if st.button(skip_lbl, key="s3_skip_{}".format(teacher)):
            if ss.step3_data is None: ss.step3_data = {}
            d = ss.step3_data.setdefault(teacher, {'skipped': False, 'combines': []})
            d['skipped'] = not d['skipped']
            st.rerun()


def render_teacher_detail(teacher, teacher_workload, overloaded_set, max_allowed):
    info      = teacher_workload[teacher]
    entries   = info['entries']
    s3d       = (ss.step3_data or {}).setdefault(teacher, {'skipped': False, 'combines': []})
    if ss.step3_data is None: ss.step3_data = {}
    if teacher not in ss.step3_data:
        ss.step3_data[teacher] = {'skipped': False, 'combines': []}
    s3d = ss.step3_data[teacher]
    effective = app._effective_total(teacher)
    is_over   = teacher in overloaded_set

    hdr_color = "#c0392b" if (is_over and effective > max_allowed) else "#1a7a1a"
    st.markdown(
        '<div style="background:{};color:white;padding:8px 12px;border-radius:6px;margin-bottom:8px">'
        '<strong>{}</strong> &nbsp;|&nbsp; Assigned: {}  Effective: {}  Max: {}</div>'.format(
            hdr_color, teacher, info['total'], effective, max_allowed),
        unsafe_allow_html=True)

    combined_indices = set()
    for cb in s3d['combines']:
        for idx in cb.get('entry_indices', []):
            combined_indices.add(idx)

    left_d, right_d = st.columns([1,1])

    with left_d:
        st.markdown("**Assignments** — check to combine")
        check_states = {}
        for ei, entry in enumerate(entries):
            in_cb = ei in combined_indices
            disabled_note = " ✅ in combine" if in_cb else ""
            label = "[{}] {} ({} periods){}".format(
                entry['class'], entry['subject'], entry['periods'], disabled_note)

            ct_info  = app._get_class_ct_info(entry['class'], teacher, entry['subject'])
            sub_info = "CT: {}  |  CT Subjects: {}".format(
                ct_info['ct'] or '—',
                ', '.join(ct_info['ct_subjects']) if ct_info['ct_subjects'] else '—')
            par_warn = ""
            if ct_info['is_parallel_with_ct']:
                par_warn = "⚠ '{}' (teacher) ∥ '{}' (CT) — parallel conflict!".format(
                    entry['subject'], ct_info['parallel_ct_subject'])

            if in_cb:
                st.markdown("☑ ~~{}~~  \n*{}*".format(label, sub_info), unsafe_allow_html=False)
                if par_warn: st.warning(par_warn)
            else:
                val = st.checkbox(label, value=False,
                                   key="cb_entry_{}_{}".format(teacher, ei))
                check_states[ei] = val
                st.caption(sub_info)
                if par_warn: st.warning(par_warn)

        if st.button("✓ Combine Checked Entries", key="do_combine_{}".format(teacher)):
            idxs = [ei for ei, v in check_states.items() if v]
            if len(idxs) < 2:
                st.error("Check at least 2 assignments to combine.")
            else:
                periods_list = [entries[i]['periods'] for i in idxs]
                if len(set(periods_list)) > 1:
                    st.error("All entries to combine must have the same period count. Selected: {}".format(periods_list))
                else:
                    already = [i for i in idxs if i in combined_indices]
                    if already:
                        st.error("Some selected entries are already part of a combine.")
                    else:
                        blocked = []
                        for i in idxs:
                            e = entries[i]
                            info2 = app._get_class_ct_info(e['class'], teacher, e['subject'])
                            if info2['is_parallel_with_ct']:
                                blocked.append(e['class'])
                        if blocked:
                            st.error("Classes {} cannot be combined due to parallel-CT conflict. "
                                     "Remove them from selection.".format(', '.join(blocked)))
                        else:
                            s3d['combines'].append({
                                'entry_indices': idxs,
                                'periods_each':  periods_list[0],
                                'classes':       [entries[i]['class'] for i in idxs],
                                'subjects':      [entries[i]['subject'] for i in idxs],
                            })
                            st.rerun()

    with right_d:
        st.markdown("**Combines**")
        if not s3d['combines']:
            st.info("No combines yet.\nCheck assignments on the left and click Combine.")
        else:
            for ci, cb in enumerate(s3d['combines']):
                st.markdown(
                    '<div class="combine-card">'
                    '<strong>Combine {}: {}</strong><br>'.format(ci+1, "  +  ".join(cb['classes'])) +
                    ''.join("<small>• {}</small><br>".format(
                        entries[idx]['label'] if idx < len(entries) else '?')
                        for idx in cb.get('entry_indices', [])) +
                    '<em style="color:#2980b9">Saves {} periods/week for {}</em>'.format(
                        (len(cb['entry_indices'])-1)*cb['periods_each'], teacher) +
                    '</div>', unsafe_allow_html=True)
                if st.button("✕ Remove Combine {}".format(ci+1), key="rm_cb_{}_{}".format(teacher, ci)):
                    s3d['combines'].pop(ci)
                    st.rerun()


def _validate_step3_ui(overloaded_set, max_allowed):
    issues = []; resolved = []
    for teacher in sorted(overloaded_set):
        s3d = (ss.step3_data or {}).get(teacher, {})
        skipped = s3d.get('skipped', False)
        eff     = app._effective_total(teacher)
        if skipped:
            resolved.append("{}: SKIPPED".format(teacher))
        elif eff <= max_allowed:
            resolved.append("{}: Resolved ({} ≤ {})".format(teacher, eff, max_allowed))
        else:
            issues.append("{}: still overloaded ({}/{}) — over by {}".format(
                teacher, eff, max_allowed, eff-max_allowed))

    if not overloaded_set:
        st.success("✅ No overloaded teachers — all clear! You may proceed to Step 4.")
        return

    if issues:
        st.error("❌ {} unresolved overload(s):".format(len(issues)))
        for i in issues:
            st.write("• " + i)
        if resolved:
            st.success("Resolved ({}):\n{}".format(len(resolved), '\n'.join(resolved)))
    else:
        st.success("✅ All overloads resolved or skipped!\n" + '\n'.join(resolved))


def render_unavailability_ui(cfg, ppd, wdays):
    st.markdown("### 🚫 Teacher Unavailability")
    st.caption("Block specific days & periods for a teacher. The scheduler will not assign them during these slots.")

    day_names = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][:wdays]
    teachers  = sorted(cfg.get('teacher_names', []))
    unavail   = ss.step3_unavailability or {}

    col_form, col_list = st.columns([1, 2])

    with col_form:
        st.markdown("**Add/Edit Unavailability**")

        edit_t = ss.get('s3_unavail_edit_teacher', None)
        t_idx  = teachers.index(edit_t) if edit_t and edit_t in teachers else 0

        sel_teacher = st.selectbox("Teacher", options=[""] + teachers,
                                    index=t_idx + 1 if edit_t else 0,
                                    key="unavail_teacher_sel")

        st.markdown("*Unavailable Days:*")
        day_cols = st.columns(len(day_names))
        sel_days = []
        existing_days = set(unavail.get(sel_teacher, {}).get('days', []) if sel_teacher else [])
        for di, day in enumerate(day_names):
            with day_cols[di]:
                if st.checkbox(day, value=day in existing_days, key="ud_{}_{}".format(sel_teacher, day)):
                    sel_days.append(day)

        st.markdown("*Unavailable Periods:*")
        per_cols = st.columns(min(ppd, 8))
        sel_pers = []
        existing_pers = set(unavail.get(sel_teacher, {}).get('periods', []) if sel_teacher else [])
        for pi in range(ppd):
            with per_cols[pi % len(per_cols)]:
                if st.checkbox("P{}".format(pi+1), value=(pi+1) in existing_pers,
                                key="up_{}_{}".format(sel_teacher, pi)):
                    sel_pers.append(pi+1)

        if st.button("💾 Save Unavailability", key="unavail_save"):
            if not sel_teacher:
                st.error("⚠ Select a teacher first.")
            elif sel_teacher not in teachers:
                st.error("⚠ '{}' is not in the teacher list.".format(sel_teacher))
            elif not sel_days or not sel_pers:
                st.error("⚠ Select at least one day and one period.")
            else:
                ok, msg = app._check_unavailability_feasible(sel_teacher, sel_days, sel_pers)
                if not ok:
                    st.error("❌ " + msg)
                else:
                    if ss.step3_unavailability is None:
                        ss.step3_unavailability = {}
                    ss.step3_unavailability[sel_teacher] = {'days': sel_days, 'periods': sel_pers}
                    ss.s3_unavail_edit_teacher = None
                    st.success("✓ Saved unavailability for {}.".format(sel_teacher))
                    st.rerun()

    with col_list:
        st.markdown("**Current Unavailability Rules**")
        if not unavail:
            st.info("No unavailability rules set.")
        else:
            for teacher, inf in sorted(unavail.items()):
                st.markdown(
                    '<div class="unavail-card">'
                    '<strong>{}</strong><br>'
                    'Days: {}   |   Periods: {}'
                    '</div>'.format(
                        teacher,
                        ', '.join(inf.get('days', [])),
                        ', '.join('P{}'.format(p) for p in sorted(inf.get('periods', [])))),
                    unsafe_allow_html=True)
                ec1, ec2 = st.columns(2)
                with ec1:
                    if st.button("✏ Edit", key="unavail_edit_{}".format(teacher)):
                        ss.s3_unavail_edit_teacher = teacher
                        st.rerun()
                with ec2:
                    if st.button("✕ Remove", key="unavail_rm_{}".format(teacher)):
                        del ss.step3_unavailability[teacher]
                        st.rerun()


###############################################################################
#  STEP 4: Generation
###############################################################################

def render_step4():
    cfg   = ss.configuration
    ppd   = cfg['periods_per_day']
    wdays = cfg['working_days']

    substep = ss.s4_substep  # 'main', 'ta', 'final'

    if substep == 'ta':
        render_task_analysis()
        return
    if substep == 'final':
        render_final_timetable()
        return

    # ── Main Step 4 page ─────────────────────────────────────────────────────
    st.markdown('<div class="step-header">📅 Step 4: Generate Timetable</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="step-sub">{} days/week · {} periods/day · {} slots/week per class</div>'.format(
            wdays, ppd, wdays*ppd),
        unsafe_allow_html=True)

    if st.button("◄ Back to Step 3", key="s4_back"):
        ss.step = 3; st.rerun()

    # ── Stage guide ──────────────────────────────────────────────────────────
    with st.expander("📖 Stage Guide", expanded=False):
        st.markdown("""
| Stage | What it does |
|-------|-------------|
| **Stage 1** | Places Class Teacher (★) slots at the fixed period every day, then fills preference-constrained subjects (period/day pinned). |
| **Task Analysis** | Shows all parallel/consecutive groups. Run "Allocate Periods" to reserve their shared slots. All groups must succeed before Stage 2. |
| **Stage 2 / 3** | Fills the remaining free periods using greedy placement + constraint relaxation. |
| **Force Fill** | CSP solver — guarantees 100% placement by relaxing soft constraints as a last resort. |
        """)

    # ── Generation controls ───────────────────────────────────────────────────
    gen_stage = ss._gen_stage

    st.markdown("### Generation Control")
    stage_labels = {0: "Not started", 1: "Stage 1 complete", 2: "Task Analysis done", 3: "Full timetable generated"}
    st.info("**Current stage:** {}".format(stage_labels.get(gen_stage, "?")))

    col_g1, col_g2 = st.columns(2)
    with col_g1:
        if st.button("▶ Stage 1: Fill CT Slots & Fixed Periods", type="primary", key="run_stage1"):
            with st.spinner("Running Stage 1…"):
                app._relaxed_consec_keys = set()
                app._relaxed_main_keys   = set()
                app._init_gen_state()
                app._run_stage1_phases()
                ss._gen_stage = 1
                ss._timetable = app._gen_snapshot_tt()
            st.success("✓ Stage 1 complete!")
            issues = ss.get('_stage1_issues', [])
            if issues:
                st.warning("Stage 1 issues:\n" + "\n".join(issues[:5]))
            st.rerun()

    with col_g2:
        s1_done = gen_stage >= 1
        if st.button("📋 Task Analysis →", disabled=not s1_done, key="goto_ta"):
            ss.s4_substep = 'ta'; st.rerun()

    if gen_stage >= 3 and ss._timetable:
        st.markdown("---")
        if st.button("📊 View Final Timetable", type="primary", key="view_final"):
            ss.s4_substep = 'final'; st.rerun()

    # ── Stage 1 preview (if done) ────────────────────────────────────────────
    if gen_stage >= 1 and ss._timetable:
        tt = ss._timetable
        unplaced = tt.get('unplaced', 0)
        placed   = sum(t['periods'] - t['remaining'] for t in tt['tasks'])
        total    = sum(t['periods'] for t in tt['tasks'])
        st.markdown("---")
        mc1, mc2, mc3 = st.columns(3)
        with mc1: st.metric("Periods Placed", placed)
        with mc2: st.metric("Unplaced", unplaced)
        with mc3: st.metric("Total Periods", total)

        if unplaced > 0 and gen_stage >= 3:
            st.warning("⚠ {} period(s) still unplaced. Consider using Force Fill.".format(unplaced))


###############################################################################
#  TASK ANALYSIS PAGE
###############################################################################

def render_task_analysis():
    st.markdown('<div class="step-header">📋 Task Analysis</div>', unsafe_allow_html=True)
    st.caption("Review all parallel period groups before Stage 2 begins.")

    nav1, nav2, nav3, nav4 = st.columns(4)
    with nav1:
        if st.button("◄ Back to Stage 1", key="ta_back"):
            ss.s4_substep = 'main'; st.rerun()
    with nav2:
        if st.button("🗓 Allocate Periods", type="primary", key="ta_allocate"):
            with st.spinner("Allocating periods…"):
                slots, allocation, rows = app._run_task_analysis_allocation()
                ss._last_group_slots = slots
                ss._last_allocation  = allocation
                ss._last_all_rows    = rows
                ss._ta_alloc_done    = True
            st.rerun()
    with nav3:
        if ss._relaxed_consec_keys:
            if st.button("🔒 Reset Relaxations", key="ta_reset_relax"):
                ss._relaxed_consec_keys = set()
                ss._last_allocation = {}
                ss._ta_alloc_done = False
                st.rerun()
    with nav4:
        all_ok = bool(ss._last_allocation) and all(
            ar.get('ok', False) for ar in ss._last_allocation.values())
        if st.button("▶ Proceed to Stage 2 →", disabled=not all_ok, type="primary" if all_ok else "secondary",
                     key="ta_proceed"):
            with st.spinner("Running Stage 2/3…"):
                app._run_stage2_phases()
                ss._gen_stage  = 3
                ss._timetable  = app._gen_snapshot_tt()
            ss.s4_substep = 'final'; st.rerun()

    # ── Show task analysis table ──────────────────────────────────────────────
    s3   = ss.step3_data or {}
    cfg  = ss.configuration
    all_cls = _all_classes()
    group_allocation = ss._last_allocation or {}
    DAYS_l = (ss._gen or {}).get('DAYS', []) if ss._gen else []

    def _find_parallel(cn, sn):
        for s in ss.class_config_data.get(cn, {}).get('subjects', []):
            if s['name'] == sn and s.get('parallel'):
                return (s.get('parallel_subject','').strip() or '?',
                        s.get('parallel_teacher','').strip() or '—')
            if (s.get('parallel') and s.get('parallel_subject','').strip() == sn and s['name'] != sn):
                return (s['name'], s.get('teacher','').strip() or '—')
        return ('—','—')

    all_rows_ta = []
    group_no = 0; covered = set()
    for teacher, s3d in sorted(s3.items()):
        for cb in s3d.get('combines', []):
            classes = cb.get('classes', []); subjects = cb.get('subjects', [])
            if not classes: continue
            group_no += 1
            for j, cn in enumerate(classes):
                tsub = subjects[j] if j < len(subjects) else (subjects[0] if subjects else '?')
                ps, pt_ = _find_parallel(cn, tsub)
                all_rows_ta.append({'group': group_no, 'class': cn, 'subject': tsub,
                                    'teacher': teacher, 'par_subj': ps, 'par_teacher': pt_, 'section': 'A'})
                covered.add((cn, tsub))
                if ps not in ('—','?'): covered.add((cn, ps))

    seen_pairs = set()
    for cn in all_cls:
        for s in ss.class_config_data.get(cn, {}).get('subjects', []):
            if not s.get('parallel'): continue
            sn = s['name']; st_ = s.get('teacher','').strip()
            ps = s.get('parallel_subject','').strip(); pt_ = s.get('parallel_teacher','').strip()
            if not ps: continue
            if (cn, sn) in covered or (cn, ps) in covered: continue
            pk = frozenset([(cn, sn), (cn, ps)])
            if pk in seen_pairs: continue
            seen_pairs.add(pk); group_no += 1
            all_rows_ta.append({'group': group_no, 'class': cn, 'subject': sn, 'teacher': st_,
                                'par_subj': ps or '?', 'par_teacher': pt_ or '—', 'section': 'B'})

    consec_covered = set(covered)
    for r in all_rows_ta:
        if r.get('section') == 'B':
            consec_covered.add((r['class'], r['subject']))
            if r.get('par_subj') not in ('—','?',''): consec_covered.add((r['class'], r['par_subj']))
    seen_consec = set()
    for cn in all_cls:
        for s in ss.class_config_data.get(cn, {}).get('subjects', []):
            if s.get('consecutive','No') != 'Yes': continue
            sn = s['name']; st_ = s.get('teacher','').strip()
            key = (cn, sn)
            if key in seen_consec: continue
            seen_consec.add(key); group_no += 1
            all_rows_ta.append({'group': group_no, 'class': cn, 'subject': sn, 'teacher': st_,
                                'par_subj': '—', 'par_teacher': '—', 'section': 'C',
                                'periods': s.get('periods','')})

    if not all_rows_ta:
        st.info("No parallel/consecutive groups found.")
        return

    # Display table
    def _alloc_text(gn):
        ar = group_allocation.get(gn)
        if not group_allocation: return "—"
        if ar is None: return "—"
        if ar['ok'] and not ar.get('slots') and ar.get('s1_placed', 0) > 0:
            return "✓ Stage 1 ({})".format(ar['s1_placed'])
        placed_s = []
        for d, p in ar.get('slots', []):
            dn = DAYS_l[d] if d < len(DAYS_l) else "D{}".format(d+1)
            placed_s.append("{} P{}".format(dn, p+1))
        s1_note = " +{}S1".format(ar['s1_placed']) if ar.get('s1_placed',0) > 0 else ""
        if ar['ok']:
            return " · ".join(placed_s) + s1_note
        if placed_s:
            return " · ".join(placed_s) + s1_note + " ⚠partial"
        return "FAILED: " + ar.get('reason','?')[:60]

    SEC_CLR = {'A': '#eaf4fb', 'B': '#f5eef8', 'C': '#eafaf1'}
    sec_headers = {'A': '🔵 Section A — Combined Groups', 'B': '🟣 Section B — Standalone Parallel Pairs', 'C': '🟢 Section C — Consecutive Periods'}

    prev_sec = None
    for row in all_rows_ta:
        sec = row['section']
        if sec != prev_sec:
            st.markdown("#### {}".format(sec_headers.get(sec, sec)))
            prev_sec = sec

        gn    = row['group']
        ar    = group_allocation.get(gn)
        alloc = _alloc_text(gn)
        ok_str = "✅" if (ar and ar.get('ok')) else ("⏳" if not group_allocation else "❌")

        par_str = ""
        if row.get('par_subj','—') not in ('—','?',''):
            par_str = "  ∥  {} / {}".format(row['par_subj'], row.get('par_teacher','—'))

        consec_str = ""
        if sec == 'C' and (row['class'], row['subject']) in (ss._relaxed_consec_keys or set()):
            consec_str = "🔓 relaxed"
        elif sec == 'C':
            consec_str = "🔒 strict"

        bg = SEC_CLR.get(sec, '#fff')
        st.markdown(
            '<div style="background:{};border:1px solid #ccc;border-radius:4px;padding:6px 10px;margin:2px 0;font-size:0.85rem">'
            '<strong>Group {}</strong> · Class {} · {}/{}{}'
            '<span style="float:right">{} {}</span>'
            '{}'
            '</div>'.format(
                bg, gn, row['class'], row['subject'], row['teacher'], par_str,
                ok_str, alloc,
                '<br><small style="color:#888">{}</small>'.format(ar.get('reason','')[:80]) if ar and not ar.get('ok') else ''
            ), unsafe_allow_html=True)

        # Relax button for C section failures
        if sec == 'C' and ar and not ar.get('ok') and (row['class'], row['subject']) not in (ss._relaxed_consec_keys or set()):
            if st.button("🔓 Relax Consecutive for {}:{} in {}".format(row['class'], row['subject'], row['class']),
                         key="relax_{}_{}".format(row['class'], row['subject'])):
                if ss._relaxed_consec_keys is None: ss._relaxed_consec_keys = set()
                ss._relaxed_consec_keys.add((row['class'], row['subject']))
                ss._last_allocation = {}
                ss._ta_alloc_done = False
                st.rerun()

    # Check if all failed groups need showing
    if group_allocation:
        failed = {gn: ar for gn, ar in group_allocation.items() if not ar.get('ok')}
        if failed:
            st.error("⚠ {} group(s) not fully allocated. Fix issues above and re-run Allocate Periods.".format(len(failed)))
        else:
            st.success("✅ All groups allocated successfully! You can now Proceed to Stage 2.")


###############################################################################
#  FINAL TIMETABLE PAGE
###############################################################################

def render_final_timetable():
    tt = ss._timetable
    if not tt:
        st.error("No timetable available. Please run generation first.")
        if st.button("◄ Back", key="ft_back_empty"):
            ss.s4_substep = 'main'; st.rerun()
        return

    unplaced = tt.get('unplaced', 0)
    total    = sum(t['periods'] for t in tt['tasks'])
    placed   = total - unplaced

    # Header
    badge_col = "#1a7a1a" if unplaced == 0 else "#c0392b"
    badge_txt = "✅ Complete — {} periods placed".format(total) if unplaced == 0 else \
                "⚠ {} period(s) unplaced".format(unplaced)
    st.markdown(
        '<div style="background:#2c3e50;color:white;padding:10px 16px;border-radius:8px;margin-bottom:8px">'
        '<strong style="font-size:1.3rem">📅 Generated Timetable</strong>'
        '&nbsp;<span style="background:{};padding:3px 10px;border-radius:4px;font-size:0.9rem">{}</span>'
        '</div>'.format(badge_col, badge_txt), unsafe_allow_html=True)

    # Action row
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("◄ Back to Task Analysis", key="ft_back"):
            ss.s4_substep = 'ta'; st.rerun()
    with c2:
        if st.button("🔧 Force Fill", key="ft_ff"):
            with st.spinner("Force Fill running — this may take a moment…"):
                relaxed = app._force_fill_backtrack()
                ss._timetable = app._gen_snapshot_tt()
            remaining = sum(t['remaining'] for t in app._gen['tasks'])
            if remaining == 0:
                msg = "✅ All periods placed!"
                if relaxed: msg += "\n\nConstraints relaxed:\n" + relaxed
                st.success(msg)
            else:
                st.warning("⚠ {} period(s) still unplaced.".format(remaining))
            st.rerun()

    # Excel download section
    st.markdown("### 📥 Download Excel")
    dl_cols = st.columns(5)
    excel_modes = [
        ("class",     "📚 Classwise Timetable"),
        ("teacher",   "👩‍🏫 Teacherwise Timetable"),
        ("ct_list",   "📋 Class Teacher List"),
        ("workload",  "⚖️ Teacher Workload"),
        ("one_sheet", "📄 One-Sheet Teacherwise"),
    ]
    for (mode, label), col in zip(excel_modes, dl_cols):
        with col:
            try:
                xls_bytes = app._write_excel_bytes(mode)
                fname = "{}_{}.xlsx".format(mode, datetime.now().strftime("%Y%m%d_%H%M%S"))
                st.download_button(label, data=xls_bytes, file_name=fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_{}".format(mode))
            except Exception as e:
                st.error("Export error: {}".format(e))

    st.markdown("---")

    # Timetable tabs
    tab_cls, tab_tch, tab_sum = st.tabs(["📚 Class Timetables", "👩‍🏫 Teacher Timetables", "✅ Summary"])

    with tab_cls:
        render_class_timetable(tt)

    with tab_tch:
        render_teacher_timetable(tt)

    with tab_sum:
        render_summary(tt)


def render_class_timetable(tt):
    grid  = tt['grid']
    days  = tt['days']
    ppd   = tt['ppd']
    half1 = tt['half1']
    classes = tt['all_classes']

    sel_class = st.selectbox("Select Class", options=classes, key="cls_tt_sel")
    if not sel_class: return

    rows_html = []
    # Header row
    header_cells = '<th class="tt-hdr tt-day">Day</th>'
    for p in range(ppd):
        lbl = "①" if p < half1 else "②"
        header_cells += '<th class="tt-hdr">P{} {}</th>'.format(p+1, lbl)
    rows_html.append("<tr>{}</tr>".format(header_cells))

    for d, dname in enumerate(days):
        cells = '<td class="tt-day">{}</td>'.format(dname)
        for p in range(ppd):
            e = grid.get(sel_class, [[]])[d][p] if d < len(grid.get(sel_class,[])) else None
            if e is None:
                cells += '<td class="tt-free">FREE</td>'
            else:
                etype = e.get('type','normal')
                if etype == 'combined_parallel':
                    l1, l2 = app._get_combined_par_display(sel_class, e)
                    txt = "{}<br><small>{}</small>".format(l1.replace('\n','<br>'), l2.replace('\n','<br>'))
                    cells += '<td class="tt-cpar">{}</td>'.format(txt)
                elif etype == 'parallel':
                    txt = "{}/{}<br><small>{}/{}</small>".format(
                        e['subject'], e.get('par_subj',''), e['teacher'], e.get('par_teach',''))
                    cells += '<td class="tt-par">{}</td>'.format(txt)
                elif etype == 'combined':
                    cc = e.get('combined_classes',[])
                    mark = " ★" if e.get('is_ct') else ""
                    txt = "{}{}[{}]<br><small>{}</small>".format(e['subject'],mark,'+'.join(cc),e['teacher'])
                    cells += '<td class="tt-comb">{}</td>'.format(txt)
                else:
                    mark = " ★" if e.get('is_ct') else ""
                    txt = "{}{}<br><small>{}</small>".format(e['subject'], mark, e['teacher'])
                    td_cls = "tt-ct" if e.get('is_ct') else "tt-normal"
                    cells += '<td class="{}">{}</td>'.format(td_cls, txt)
        rows_html.append("<tr>{}</tr>".format(cells))

    cfg = ss.class_config_data.get(sel_class, {})
    ct_name = cfg.get('teacher_var', Var('')).get()
    ct_per  = cfg.get('teacher_period_var', Var(1)).get()
    caption = "Class: {}   |   Class Teacher: {}   |   CT Period: {}".format(
        sel_class, ct_name or '—', ct_per)

    html = '<p style="font-weight:700;font-size:0.9rem">{}</p>'.format(caption)
    html += '<div style="overflow-x:auto"><table class="tt-table">{}</table></div>'.format(
        "".join(rows_html))
    html += '<p style="font-size:0.75rem;margin-top:6px">★=Class Teacher &nbsp; 🔵=Combined &nbsp; 🟠=Parallel &nbsp; 🔴=Combined+Parallel</p>'
    st.markdown(html, unsafe_allow_html=True)


def render_teacher_timetable(tt):
    grid  = tt['grid']
    days  = tt['days']
    ppd   = tt['ppd']
    half1 = tt['half1']
    classes = tt['all_classes']

    # Build teacher grid
    teacher_grid = {}
    for cn in classes:
        for d in range(len(days)):
            for p in range(ppd):
                e = grid.get(cn, [[]])[d][p] if d < len(grid.get(cn,[])) else None
                if not e: continue
                cc    = e.get('combined_classes', [])
                etype = e.get('type','normal')
                is_cp = bool(cc) and etype == 'combined_parallel'
                is_c  = bool(cc) and etype == 'combined'

                def _tg_add(tname, tcls, tsubj, tct):
                    if not tname: return
                    teacher_grid.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                    teacher_grid[tname][d][p] = {'subject': tsubj, 'class': tcls, 'is_ct': tct, 'type': etype}

                if is_cp:
                    combined_teacher = ''; combined_subj = ''
                    s3 = ss.step3_data or {}
                    for _t, s3d in s3.items():
                        for cb in s3d.get('combines', []):
                            if set(cb.get('classes',[])) == set(cc):
                                combined_teacher = _t
                                combined_subj = cb.get('subjects',[''])[0] if cb.get('subjects') else ''
                                break
                        if combined_teacher: break
                    if not combined_teacher:
                        combined_teacher = e.get('teacher',''); combined_subj = e.get('subject','')
                    if combined_teacher and (not cc or cn == cc[0]):
                        _tg_add(combined_teacher, '+'.join(cc), combined_subj, False)
                    class_teacher2 = e.get('par_teach',''); class_subj2 = e.get('par_subj','')
                    if combined_subj and cn in ss.class_config_data:
                        for _s in ss.class_config_data[cn].get('subjects', []):
                            sname = _s.get('name','').strip(); pname = (_s.get('parallel_subject') or '').strip()
                            if sname == combined_subj:
                                class_subj2 = pname; class_teacher2 = (_s.get('parallel_teacher') or '').strip(); break
                            elif pname == combined_subj:
                                class_subj2 = sname; class_teacher2 = _s.get('teacher','').strip(); break
                    if class_teacher2:
                        _tg_add(class_teacher2, cn, class_subj2, e.get('is_ct',False))
                elif is_c:
                    t = e.get('teacher')
                    if t and (not cc or cn == cc[0]):
                        _tg_add(t, '+'.join(cc) if cc else cn, e['subject'], e.get('is_ct',False))
                else:
                    t = e.get('teacher')
                    if t: _tg_add(t, cn, e['subject'], e.get('is_ct',False))
                    pt = e.get('par_teach')
                    if pt and pt not in ('—','?',''):
                        _tg_add(pt, cn, e.get('par_subj',''), False)

    teacher_list = sorted(teacher_grid.keys())
    if not teacher_list:
        st.info("No teacher data available.")
        return

    sel_teacher = st.selectbox("Select Teacher", options=teacher_list, key="tch_tt_sel")
    if not sel_teacher: return

    tg = teacher_grid.get(sel_teacher, [[None]*ppd for _ in range(len(days))])

    rows_html = []
    header_cells = '<th class="tt-hdr tt-day">{}</th>'.format(sel_teacher[:16])
    for p in range(ppd):
        lbl = "①" if p < half1 else "②"
        header_cells += '<th class="tt-hdr">P{} {}</th>'.format(p+1, lbl)
    rows_html.append("<tr>{}</tr>".format(header_cells))

    for d, dname in enumerate(days):
        cells = '<td class="tt-day">{}</td>'.format(dname)
        for p in range(ppd):
            e = tg[d][p] if tg and d < len(tg) else None
            if e is None:
                cells += '<td class="tt-free">FREE</td>'
            else:
                bg_cls = "tt-ct" if e.get('is_ct') else "tt-normal"
                txt = "{}<br><small>{}</small>".format(e.get('class',''), e.get('subject',''))
                cells += '<td class="{}">{}</td>'.format(bg_cls, txt)
        rows_html.append("<tr>{}</tr>".format(cells))

    html = '<div style="overflow-x:auto"><table class="tt-table">{}</table></div>'.format(
        "".join(rows_html))
    st.markdown(html, unsafe_allow_html=True)


def render_summary(tt):
    tasks   = tt['tasks']
    days    = tt['days']
    ppd     = tt['ppd']
    half1   = tt['half1']
    grid    = tt['grid']
    classes = tt['all_classes']

    st.markdown("### Unplaced Periods")
    any_unplaced = False
    for t in tasks:
        if t['remaining'] > 0:
            st.error("❌  Class {}  '{}'  teacher {}  — {} period(s) unplaced".format(
                '+'.join(t['cn_list']), t['subject'], t['teacher'], t['remaining']))
            any_unplaced = True
    if not any_unplaced:
        st.success("✓ All periods placed.")

    st.markdown("### Teacher Free-Period Distribution")
    all_teachers = sorted(set(
        t for cn in classes
        for d in range(len(days)) for p in range(ppd)
        for e in [grid.get(cn,[[]])[d][p] if d < len(grid.get(cn,[])) else None]
        if e and e.get('teacher')
        for t in [e['teacher']]))

    for teacher in all_teachers:
        busy = {}
        for cn in classes:
            for d in range(len(days)):
                for p in range(ppd):
                    e = grid.get(cn,[[]])[d][p] if d < len(grid.get(cn,[])) else None
                    if e and (e.get('teacher') == teacher or e.get('par_teach') == teacher):
                        busy.setdefault(d, set()).add(p)
        lines = []
        for d in range(len(days)):
            bd = busy.get(d, set())
            fh1 = half1 - len([x for x in bd if x < half1])
            fh2 = (ppd - half1) - len([x for x in bd if x >= half1])
            if fh1 + fh2 == ppd: continue
            ok = "✓" if (fh1 >= 1 and fh2 >= 1) else "⚠"
            lines.append("{} {}  H1={} H2={}".format(ok, days[d], fh1, fh2))
        if lines:
            with st.expander("{}".format(teacher)):
                for line in lines:
                    if line.startswith("⚠"):
                        st.warning(line)
                    else:
                        st.write(line)

    st.markdown("### Legend")
    st.write("★ = Class Teacher period  |  Green = CT period  |  Blue = Combined  |  Orange = Parallel  |  Pink = Combined+Parallel")


###############################################################################
#  MAIN ROUTER
###############################################################################

def main():
    # Sidebar navigation
    with st.sidebar:
        st.title("🗓 Timetable Generator")
        st.caption("V4.0")
        st.markdown("---")
        step = ss.step
        steps = ["Step 1: Configuration", "Step 2: Classes", "Step 3: Combines", "Step 4: Generate"]
        for i, s in enumerate(steps, 1):
            if i == step:
                st.markdown("**▶ {}**".format(s))
            else:
                st.markdown("{}".format(s))
        st.markdown("---")
        st.caption("Tip: All configs download as JSON files to your computer. Upload them to restore your work.")

    if step == 1:
        render_step1()
    elif step == 2:
        if not ss.configuration:
            st.error("Please complete Step 1 first.")
            if st.button("Go to Step 1"):
                ss.step = 1; st.rerun()
        else:
            render_step2()
    elif step == 3:
        if not ss.configuration:
            st.error("Please complete Step 1 first.")
        else:
            render_step3()
    elif step == 4:
        if not ss.configuration:
            st.error("Please complete Step 1 first.")
        else:
            render_step4()
    else:
        render_step1()


if __name__ == "__main__":
    main()
