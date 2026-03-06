"""
streamlit_app.py — Timetable Generator (Streamlit) v4.1

BUGS FIXED vs v1:
  BUG-1 (engine): 11 tkinter StringVar .get() remnants → replaced with plain dict access
  BUG-2 (upload hang): file-uploader fires on every rerun → guarded by SHA-1 hash
         dedup; widget keys deleted after JSON load so widgets reinitialise correctly
  BUG-3 (manual entry stuck): number_input widget keys now ARE the session_state keys
         (ni_ppd etc.) — no shadow s1_ppd copies that get silently overwritten
  LOGGING: Python logging to console + in-memory buffer (sidebar Debug Log)
           Format: [LEVEL] filename:function:lineno — message
"""

import io
import json
import logging
import traceback
import hashlib
from datetime import datetime

import streamlit as st
from engine import TimetableEngine


# ─────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────────────────────────────────────
class _MemHandler(logging.Handler):
    """Keeps last N log lines in memory for the sidebar debug console."""
    MAX = 500
    def __init__(self):
        super().__init__()
        self.lines = []
    def emit(self, record):
        self.lines.append(self.format(record))
        if len(self.lines) > self.MAX:
            self.lines = self.lines[-self.MAX:]

_LOG_FMT = "[%(levelname)s] %(filename)s:%(funcName)s:%(lineno)d — %(message)s"
_mem_handler = _MemHandler()
_mem_handler.setFormatter(logging.Formatter(_LOG_FMT))

logging.basicConfig(level=logging.DEBUG, format=_LOG_FMT,
                    handlers=[logging.StreamHandler(), _mem_handler], force=True)
log = logging.getLogger("timetable")
log.info("streamlit_app module loaded")


# ─────────────────────────────────────────────────────────────────────────────
#  Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Timetable Generator", page_icon="📅",
                   layout="wide", initial_sidebar_state="collapsed")


# ─────────────────────────────────────────────────────────────────────────────
#  Session-state bootstrap
# ─────────────────────────────────────────────────────────────────────────────
def _init_state():
    defaults = {
        "page":              "step1",
        "engine":            TimetableEngine(),
        # Step-1 — widget keys ARE the canonical values (no shadow copies)
        "ni_ppd":            7,
        "ni_wdays":          6,
        "ni_fhalf":          4,
        "ni_shalf":          3,
        "s1_teachers":       [],
        "s1_teacher_fname":  "",
        "s1_sections":       {cls: 4 for cls in range(6, 13)},
        # Upload dedup: hash of the last bytes processed per uploader key
        "_upload_hash":      {},
        # Staged config upload: bytes held until user clicks Load Config
        "_s1_pending_raw":   None,
        "_s1_pending_hash":  None,
        "_s1_pending_name":  None,
        # Step-4
        "s4_stage":          0,
        "s4_s1_status":      None,
        "s4_s3_status":      None,
        # Task-analysis
        "ta_allocation":     None,
        "ta_group_slots":    None,
        "ta_all_rows":       None,
        "ta2_allocation":    None,
        # Relaxed keys
        "relaxed_consec":    set(),
        "relaxed_main":      set(),
        # Validation caches
        "s2_validation_result": None,
        "s3_validation_result": None,
        # Pending notifications (cleared each render)
        "_notify":           [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    log.debug("_init_state: session state ready")

_init_state()
eng: TimetableEngine = st.session_state.engine


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────
DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

def _nav(page: str):
    log.info("_nav: → %s", page)
    st.session_state.page = page
    st.rerun()

def _notify(msg: str, kind: str = "info"):
    log.debug("_notify [%s]: %s", kind, msg)
    st.session_state["_notify"].append((kind, msg))

def _show_notifications():
    for kind, msg in st.session_state.get("_notify", []):
        if kind == "success": st.success(msg)
        elif kind == "error":   st.error(msg)
        elif kind == "warning": st.warning(msg)
        else:                   st.info(msg)
    st.session_state["_notify"] = []

def _file_hash(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()

def _already_processed(key: str, data: bytes) -> bool:
    """Return True if we already handled this exact file upload (same bytes)."""
    h = _file_hash(data)
    if st.session_state["_upload_hash"].get(key) == h:
        log.debug("_already_processed: key=%s hash=%s already done", key, h[:8])
        return True
    st.session_state["_upload_hash"][key] = h
    log.debug("_already_processed: key=%s hash=%s NEW", key, h[:8])
    return False

def _all_classes():
    cfg = eng.configuration
    return [
        f"{cls}{chr(65+si)}"
        for cls in range(6, 13)
        for si in range(cfg.get("classes", {}).get(cls, 0))
    ]

def _json_download(data: dict, label: str, filename: str):
    st.download_button(label=label, data=json.dumps(data, indent=2),
                       file_name=filename, mime="application/json",
                       use_container_width=True)

def _excel_download(mode: str, label: str):
    log.info("_excel_download: mode=%s", mode)
    try:
        xbytes = eng.get_excel_bytes(mode)
        fname  = f"{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(label=label, data=xbytes, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        log.info("_excel_download: ready %s (%d bytes)", fname, len(xbytes))
    except Exception as ex:
        log.error("_excel_download: %s\n%s", ex, traceback.format_exc())
        st.error(f"Export error: {ex}")

PAGE_LABELS = {
    "step1":           "Step 1 — Basic Config",
    "step2":           "Step 2 — Class Assignments",
    "step3":           "Step 3 — Teacher Settings",
    "step4":           "Step 4 — Generate",
    "task_analysis":   "Task Analysis",
    "task_analysis2":  "Task Analysis 2",
    "stage2_page":     "Stage 2",
    "final_timetable": "Final Timetable",
}

def _header(title: str, sub: str = ""):
    st.markdown(f"## {title}")
    if sub: st.caption(sub)
    st.divider()


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 1
# ═════════════════════════════════════════════════════════════════════════════
def page_step1():
    log.info("page_step1: render")
    _header("📋 Step 1: Basic Configuration",
            "Set periods, working days, upload teachers and define class sections.")
    _show_notifications()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Configuration", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Save current config**")
            if st.button("⬇ Prepare Download", key="s1_dl_btn", use_container_width=True):
                log.info("page_step1: preparing download")
                if not st.session_state.s1_teachers:
                    _notify("Upload a teacher file first.", "warning"); st.rerun()
                data = {
                    "periods_per_day":     st.session_state.get("ni_ppd", 7),
                    "working_days":        st.session_state.get("ni_wdays", 6),
                    "periods_first_half":  st.session_state.get("ni_fhalf", 4),
                    "periods_second_half": st.session_state.get("ni_shalf", 3),
                    "teacher_file_path":   st.session_state.s1_teacher_fname,
                    "teacher_names":       st.session_state.s1_teachers,
                    "classes":             {str(k): v for k, v in
                                            st.session_state.s1_sections.items()},
                    "saved_at":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.session_state["_s1_dl_data"] = data
            if "_s1_dl_data" in st.session_state:
                _json_download(st.session_state["_s1_dl_data"], "📥 Click to Download",
                               f"Config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            st.markdown("**Load saved config**")
            up = st.file_uploader("Upload Config JSON", type=["json"], key="s1_load_json")
            if up is not None:
                # Stage 1: read bytes once and store — do NOT process yet.
                # up.read() returns empty on the second rerun of the same file,
                # so we only read when we actually have bytes.
                raw = up.read()
                if len(raw) > 0:
                    h = _file_hash(raw)
                    if st.session_state.get("_s1_pending_hash") != h:
                        # New file — cache it and wait for button click
                        st.session_state["_s1_pending_raw"]  = raw
                        st.session_state["_s1_pending_hash"] = h
                        st.session_state["_s1_pending_name"] = up.name
                        log.debug("page_step1: cached new file '%s' (%d bytes)", up.name, len(raw))

            pending = st.session_state.get("_s1_pending_raw")
            if pending:
                pname = st.session_state.get("_s1_pending_name", "config.json")
                st.info(f"📄 Ready to load: **{pname}** ({len(pending):,} bytes)")
                if st.button("📂 Load Config", key="s1_load_btn", type="primary",
                             use_container_width=True):
                    log.info("page_step1: Load Config button clicked for '%s'", pname)
                    _load_step1_config(pending)
            else:
                st.caption("Upload a JSON config file, then click Load Config.")

    # ── Periods & Days ────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**1. Periods & Working Days**")
        c1, c2 = st.columns(2)
        with c1:
            # FIX: widget key == session_state key → ONE source of truth
            ppd   = st.number_input("Periods per day",   1, 20,
                                    st.session_state.get("ni_ppd",   7), key="ni_ppd")
            wdays = st.number_input("Working days/week", 1, 7,
                                    st.session_state.get("ni_wdays", 6), key="ni_wdays")
        with c2:
            fhalf = st.number_input("Periods — first half",  1, 20,
                                    st.session_state.get("ni_fhalf", 4), key="ni_fhalf")
            shalf = st.number_input("Periods — second half", 1, 20,
                                    st.session_state.get("ni_shalf", 3), key="ni_shalf")
        if fhalf + shalf == ppd:
            st.success(f"✓ Valid: {fhalf} + {shalf} = {ppd}")
        else:
            st.error(f"✗ {fhalf} + {shalf} = {fhalf+shalf}, need {ppd}")

    # ── Teachers ──────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**2. Teachers — Upload Excel File**")
        st.caption("Column A only: one name per row. "
                   "Header 'Teacher Name' is automatically skipped.")
        col_up, col_info = st.columns([3, 1])
        with col_up:
            tf = st.file_uploader("Teacher Excel (.xlsx/.xls)", type=["xlsx","xls"],
                                   key="s1_teacher_up")
            if tf is not None:
                raw_tf = tf.read()
                log.debug("page_step1: teacher upload size=%d", len(raw_tf))
                if len(raw_tf) == 0:
                    log.warning("page_step1: teacher 0 bytes, skipping")
                elif not _already_processed("s1_teacher_up", raw_tf):
                    _load_teacher_bytes(raw_tf, tf.name)
        with col_info:
            if st.session_state.s1_teachers:
                st.success(f"✓ {len(st.session_state.s1_teachers)} teachers")
                if st.button("👁 Preview", key="s1_preview"):
                    st.session_state["_s1_show_t"] = not st.session_state.get("_s1_show_t", False)
            else:
                st.warning("No teachers loaded")
        if st.session_state.get("_s1_show_t") and st.session_state.s1_teachers:
            st.dataframe({"Teacher Name": st.session_state.s1_teachers},
                         use_container_width=True, height=200)

    # ── Classes ───────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("**3. Classes 6–12 — Number of Sections**")
        cols = st.columns(7)
        for idx, cls in enumerate(range(6, 13)):
            with cols[idx]:
                cur = st.session_state.s1_sections.get(cls, 4)
                val = st.number_input(f"Class {cls}", 1, 10, cur, key=f"ni_cls_{cls}")
                st.session_state.s1_sections[cls] = val

    # ── Navigation ────────────────────────────────────────────────────────────
    cb, cc = st.columns([1, 3])
    with cc:
        if st.button("✓ Continue to Step 2 →", type="primary", use_container_width=True):
            _step1_save_and_continue()
    with cb:
        if st.button("⟲ Reset", use_container_width=True):
            log.info("page_step1: reset")
            for k in ["ni_ppd","ni_wdays","ni_fhalf","ni_shalf",
                      "s1_teachers","s1_teacher_fname","_s1_dl_data","_s1_show_t",
                      "_s1_pending_raw","_s1_pending_hash","_s1_pending_name"]:
                st.session_state.pop(k, None)
            st.session_state.s1_sections  = {cls: 4 for cls in range(6, 13)}
            st.session_state["_upload_hash"] = {}
            st.rerun()


def _load_step1_config(raw: bytes):
    log.info("_load_step1_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
        ppd   = int(d.get("periods_per_day",    7))
        wdays = int(d.get("working_days",        6))
        fhalf = int(d.get("periods_first_half",  4))
        shalf = int(d.get("periods_second_half", 3))
        teachers   = d.get("teacher_names", [])
        classes_raw = d.get("classes", {})
        log.info("_load_step1_config: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
                 ppd, wdays, fhalf, shalf, len(teachers))

        # ── Clear ALL widget keys that will be set from JSON ───────────────────
        # Streamlit ignores value= if the widget key is already stored.
        # Deleting forces re-initialisation from value= on the next render.
        # IMPORTANT: class-section keys (ni_cls_6..12) were missing before —
        # that caused section counts to silently revert to stale widget values.
        keys_to_clear = (
            ["ni_ppd", "ni_wdays", "ni_fhalf", "ni_shalf"]
            + [f"ni_cls_{cls}" for cls in range(6, 13)]
        )
        for wk in keys_to_clear:
            st.session_state.pop(wk, None)
            log.debug("_load_step1_config: cleared widget key %s", wk)

        sections = {int(k): v for k, v in classes_raw.items()}

        st.session_state["ni_ppd"]           = ppd
        st.session_state["ni_wdays"]         = wdays
        st.session_state["ni_fhalf"]         = fhalf
        st.session_state["ni_shalf"]         = shalf
        st.session_state["s1_teachers"]      = teachers
        st.session_state["s1_teacher_fname"] = d.get("teacher_file_path", "")
        st.session_state["s1_sections"]      = sections
        # Pre-populate class-section widget keys so they show the loaded values
        for cls, nsec in sections.items():
            st.session_state[f"ni_cls_{cls}"] = nsec
            log.debug("_load_step1_config: ni_cls_%d = %d", cls, nsec)

        # Clear the pending buffer so the Load button disappears after success
        for k in ("_s1_pending_raw", "_s1_pending_hash", "_s1_pending_name"):
            st.session_state.pop(k, None)

        _notify(f"✓ Config loaded — {len(teachers)} teachers, "
                f"{ppd} periods/day, {wdays} days/week, "
                f"{sum(sections.values())} total sections.", "success")
        log.info("_load_step1_config: applied OK — sections=%s", sections)
        st.rerun()
    except json.JSONDecodeError as ex:
        log.error("_load_step1_config: JSON error: %s", ex)
        _notify(f"Invalid JSON: {ex}", "error")
    except Exception as ex:
        log.error("_load_step1_config: %s\n%s", ex, traceback.format_exc())
        _notify(f"Failed to load: {ex}", "error")


def _load_teacher_bytes(raw: bytes, fname: str):
    log.info("_load_teacher_bytes: '%s' %d bytes", fname, len(raw))
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        max_col = max(
            (ci for row in ws.iter_rows() for ci, c in enumerate(row, 1) if c.value is not None),
            default=0)
        if max_col > 1:
            log.warning("_load_teacher_bytes: %d columns found, expected 1", max_col)
            _notify(f"❌ Data in {max_col} columns — use Column A only.", "error"); return
        names = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            v = row[0].value
            if v:
                n = str(v).strip()
                if n and n.lower() != "teacher name":
                    names.append(n)
        if not names:
            log.warning("_load_teacher_bytes: no names found")
            _notify("❌ No teacher names found.", "error"); return
        seen, dups = set(), []
        for n in names:
            if n in seen: dups.append(n)
            seen.add(n)
        if dups:
            log.warning("_load_teacher_bytes: duplicates: %s", dups)
            _notify("❌ Duplicates: " + ", ".join(set(dups)), "error"); return
        names.sort()
        st.session_state["s1_teachers"]      = names
        st.session_state["s1_teacher_fname"] = fname
        _notify(f"✓ {len(names)} teachers loaded (A→Z).", "success")
        log.info("_load_teacher_bytes: %d teachers loaded", len(names))
        st.rerun()
    except Exception as ex:
        log.error("_load_teacher_bytes: %s\n%s", ex, traceback.format_exc())
        _notify(f"Error reading file: {ex}", "error")


def _step1_save_and_continue():
    log.info("_step1_save_and_continue: called")
    # FIX: read directly from widget-key state (ni_ppd etc.) — these are the
    # ACTUAL current values shown in the widgets, not stale shadow copies.
    ppd   = int(st.session_state.get("ni_ppd",   7))
    wdays = int(st.session_state.get("ni_wdays", 6))
    fhalf = int(st.session_state.get("ni_fhalf", 4))
    shalf = int(st.session_state.get("ni_shalf", 3))
    teachers = st.session_state.s1_teachers
    log.debug("_step1_save_and_continue: ppd=%d wdays=%d fhalf=%d shalf=%d teachers=%d",
              ppd, wdays, fhalf, shalf, len(teachers))

    if ppd <= 0 or wdays <= 0:
        log.warning("_step1_save_and_continue: invalid ppd/wdays")
        _notify("Periods and days must be ≥ 1.", "error"); return
    if fhalf + shalf != ppd:
        log.warning("_step1_save_and_continue: halves %d+%d≠%d", fhalf, shalf, ppd)
        _notify(f"Halves mismatch: {fhalf}+{shalf}={fhalf+shalf}, need {ppd}.", "error"); return
    if not teachers:
        log.warning("_step1_save_and_continue: no teachers")
        _notify("Upload teacher file first.", "error"); return

    sections = dict(st.session_state.s1_sections)
    eng.configuration = {
        "periods_per_day":    ppd,
        "working_days":       wdays,
        "periods_first_half": fhalf,
        "periods_second_half":shalf,
        "teacher_file":       st.session_state.s1_teacher_fname,
        "teacher_names":      teachers,
        "classes":            sections,
    }
    log.info("_step1_save_and_continue: configuration set, %d total sections",
             sum(sections.values()))

    for cls in range(6, 13):
        for si in range(sections.get(cls, 0)):
            cn = f"{cls}{chr(65+si)}"
            if cn not in eng.class_config_data:
                eng.class_config_data[cn] = {
                    "subjects": [], "teacher": "",
                    "teacher_period": 1, "editing_index": None,
                }
    log.info("_step1_save_and_continue: %d classes ready → step2",
             len(eng.class_config_data))
    _nav("step2")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 2
# ═════════════════════════════════════════════════════════════════════════════
def page_step2():
    log.info("page_step2: render (%d classes)", len(_all_classes()))
    _header("👨‍🏫 Step 2: Configure Each Class",
            "Set class teacher and add subjects with periods, preferences and constraints.")
    _show_notifications()

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    teachers  = sorted(cfg["teacher_names"])
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    # ── Save / Load ───────────────────────────────────────────────────────────
    with st.expander("💾 Save / Load Assignments", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s2_dl_btn", use_container_width=True):
                log.info("page_step2: preparing assignments download")
                payload = {
                    cn: {"teacher": cd.get("teacher",""),
                         "teacher_period": cd.get("teacher_period",1),
                         "subjects": cd.get("subjects",[])}
                    for cn, cd in eng.class_config_data.items()
                }
                st.session_state["_s2_dl_data"] = {
                    "assignments": payload,
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            if "_s2_dl_data" in st.session_state:
                _json_download(st.session_state["_s2_dl_data"], "📥 Click to Download",
                               f"Assignments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Assignments JSON", type=["json"], key="s2_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step2: upload size=%d", len(raw))
                if len(raw) == 0:
                    log.warning("page_step2: 0 bytes, skipping")
                elif not _already_processed("s2_load_json", raw):
                    _load_step2_assignments(raw)

    # ── Navigation ────────────────────────────────────────────────────────────
    cb, cc = st.columns([1, 3])
    with cb:
        if st.button("← Back to Step 1"): _nav("step1")
    with cc:
        if st.button("✓ Validate & Continue to Step 3 →", type="primary"):
            _step2_validate_and_continue()

    vr = st.session_state.get("s2_validation_result")
    if vr is not None:
        _display_s2_validation(vr)
        st.divider()

    if not all_cn:
        st.warning("No classes found — go back to Step 1.")
        return

    tabs = st.tabs(all_cn)
    for tab, cn in zip(tabs, all_cn):
        with tab:
            _class_config_tab(cn, teachers, ppd, wdays, required, day_names)


def _load_step2_assignments(raw: bytes):
    log.info("_load_step2_assignments: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
        n = 0
        for cn, saved in d.get("assignments", {}).items():
            if cn in eng.class_config_data:
                eng.class_config_data[cn].update({
                    "teacher":        saved.get("teacher",""),
                    "teacher_period": saved.get("teacher_period",1),
                    "subjects":       saved.get("subjects",[]),
                    "editing_index":  None,
                })
                n += 1
        _notify(f"✓ Assignments loaded for {n} classes.", "success")
        log.info("_load_step2_assignments: %d classes loaded", n)
        st.rerun()
    except Exception as ex:
        log.error("_load_step2_assignments: %s\n%s", ex, traceback.format_exc())
        _notify(f"Failed: {ex}", "error")


def _class_config_tab(cn, teachers, ppd, wdays, required, day_names):
    log.debug("_class_config_tab: %s", cn)
    cd = eng.class_config_data.setdefault(cn, {
        "subjects": [], "teacher": "", "teacher_period": 1, "editing_index": None})

    total = sum(s.get("periods",0) for s in cd.get("subjects",[]))
    icon  = "🟢" if total==required else ("🔴" if total>required else "🟡")
    diff  = required - total
    note  = "✓ exact" if total==required else (f"need {diff} more" if diff>0 else f"over by {-diff}")
    st.caption(f"{icon}  Assigned: **{total}** / {required}  ({note})")

    with st.container(border=True):
        st.markdown(f"**Class Teacher — {cn}**")
        c1, c2 = st.columns([3,1])
        with c1:
            opts    = [""] + teachers
            cur_ct  = cd.get("teacher","")
            idx_ct  = opts.index(cur_ct) if cur_ct in opts else 0
            sel_ct  = st.selectbox("Class Teacher", opts, index=idx_ct, key=f"ct_{cn}")
            cd["teacher"] = sel_ct
        with c2:
            cur_per = int(cd.get("teacher_period",1))
            sel_per = st.number_input("CT Period", 1, ppd, cur_per, key=f"ctp_{cn}")
            cd["teacher_period"] = int(sel_per)

    subjects = cd.get("subjects",[])
    if subjects:
        with st.container(border=True):
            st.markdown("**Subjects**")
            hdr = st.columns([0.4,2.2,2.2,0.9,1.0,1.0,0.7,0.7])
            for h, t in zip(hdr, ["#","Subject","Teacher","Periods","Consec","Parallel","✏","🗑"]):
                h.markdown(f"**{t}**")
            for i, s in enumerate(subjects):
                row = st.columns([0.4,2.2,2.2,0.9,1.0,1.0,0.7,0.7])
                row[0].write(str(i+1))
                row[1].write(s.get("name","—"))
                row[2].write(s.get("teacher","—"))
                row[3].write(str(s.get("periods","")))
                row[4].write("Yes" if s.get("consecutive")=="Yes" else "—")
                row[5].write("✓" if s.get("parallel") else "—")
                if row[6].button("✏", key=f"edit_{cn}_{i}"):
                    log.debug("_class_config_tab: %s edit %d", cn, i)
                    cd["editing_index"] = i; st.rerun()
                if row[7].button("🗑", key=f"del_{cn}_{i}"):
                    log.info("_class_config_tab: %s delete %d '%s'", cn, i, s.get("name",""))
                    subjects.pop(i)
                    ei = cd.get("editing_index")
                    if ei == i: cd["editing_index"] = None
                    elif ei and ei > i: cd["editing_index"] = ei - 1
                    st.rerun()

    editing_idx = cd.get("editing_index")
    prefill     = (subjects[editing_idx]
                   if editing_idx is not None and editing_idx < len(subjects) else {})
    form_title  = f"✏ Edit Subject #{editing_idx+1}" if editing_idx is not None else "➕ Add Subject"

    with st.container(border=True):
        st.markdown(f"**{form_title}**")
        c1,c2,c3 = st.columns(3)
        with c1:
            name = st.text_input("Subject Name", value=prefill.get("name",""), key=f"sf_name_{cn}")
        with c2:
            t_val = prefill.get("teacher","")
            opts  = [""] + teachers
            sel_t = st.selectbox("Teacher", opts,
                                  index=opts.index(t_val) if t_val in opts else 0,
                                  key=f"sf_teach_{cn}")
        with c3:
            pers = st.number_input("Periods/week", 1, ppd*wdays,
                                   int(prefill.get("periods",1)), key=f"sf_per_{cn}")
        c4,c5 = st.columns(2)
        with c4:
            consec = st.selectbox("Consecutive?", ["No","Yes"],
                                   index=1 if prefill.get("consecutive")=="Yes" else 0,
                                   key=f"sf_cons_{cn}")
        with c5:
            p_prefs = st.multiselect("Period prefs (optional)", list(range(1,ppd+1)),
                                      default=[int(p) for p in prefill.get("periods_pref",[])],
                                      key=f"sf_pref_{cn}")
        d_prefs = st.multiselect("Day prefs (optional)", day_names,
                                  default=[d for d in prefill.get("days_pref",[]) if d in day_names],
                                  key=f"sf_day_{cn}")
        par = st.checkbox("Parallel teaching?", value=bool(prefill.get("parallel")),
                           key=f"sf_par_{cn}")
        par_subj = par_teach = ""
        if par:
            cp1,cp2 = st.columns(2)
            with cp1:
                par_subj = st.text_input("Parallel subject",
                                          value=prefill.get("parallel_subject",""),
                                          key=f"sf_psub_{cn}")
            with cp2:
                pt_val = prefill.get("parallel_teacher","")
                opts2  = [""] + teachers
                par_teach = st.selectbox("Parallel teacher", opts2,
                                          index=opts2.index(pt_val) if pt_val in opts2 else 0,
                                          key=f"sf_pteach_{cn}")

        btn1, btn2 = st.columns(2)
        with btn1:
            lbl = "✓ Update" if editing_idx is not None else "✓ Add Subject"
            if st.button(lbl, key=f"sf_save_{cn}", type="primary"):
                log.info("_class_config_tab: %s save '%s'", cn, name)
                if not name.strip():
                    _notify("Subject name required.", "error")
                elif not sel_t:
                    _notify("Select a teacher.", "error")
                else:
                    entry = {
                        "name":             name.strip(), "teacher": sel_t,
                        "periods":          int(pers),    "consecutive": consec,
                        "periods_pref":     p_prefs,      "days_pref": d_prefs,
                        "parallel":         par,          "parallel_subject": par_subj.strip(),
                        "parallel_teacher": par_teach,
                    }
                    if editing_idx is not None:
                        subjects[editing_idx] = entry
                        cd["editing_index"]   = None
                    else:
                        subjects.append(entry)
                    _notify(f"✓ '{name.strip()}' saved.", "success")
                    st.rerun()
        with btn2:
            if editing_idx is not None:
                if st.button("✕ Cancel", key=f"sf_cancel_{cn}"):
                    cd["editing_index"] = None; st.rerun()


def _step2_validate_and_continue():
    log.info("_step2_validate_and_continue: running")
    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    required  = ppd * wdays
    all_cn    = _all_classes()
    day_names = DAY_NAMES[:wdays]
    errors, ok_list, hc, teacher_slots = [], [], [], {}

    def _add(t, desc):
        if t: teacher_slots.setdefault(t,[]).append(desc)

    for cn in all_cn:
        cd    = eng.class_config_data.get(cn,{})
        subjs = cd.get("subjects",[])
        if not subjs:
            errors.append((cn,"NO SUBJECTS added"))
            log.warning("_step2_validate: %s no subjects", cn); continue
        total = sum(s.get("periods",0) for s in subjs)
        diff  = total - required
        if total == required:
            ok_list.append((cn, f"{total}/{required} ({len(subjs)} subjects)"))
        else:
            errors.append((cn, f"Mismatch: {total} assigned, need {required} "
                               f"({'+'if diff>0 else ''}{diff})"))
            log.warning("_step2_validate: %s %d≠%d", cn, total, required)

        ct = cd.get("teacher","").strip()
        ct_per = cd.get("teacher_period",1)
        if ct:
            _add(ct, {"class":cn,"is_ct":True,"fixed_period":ct_per,
                      "day_set":set(day_names),"period_prefs":[ct_per]})
        for s in subjs:
            t = s.get("teacher","").strip()
            if t:
                _add(t, {"class":cn,"is_ct":False,"fixed_period":None,
                         "day_set":set(s.get("days_pref",[]) or day_names),
                         "period_prefs":list(s.get("periods_pref",[])),
                         "subj_name":s["name"]})
            if s.get("parallel") and s.get("parallel_teacher"):
                _add(s["parallel_teacher"].strip(),
                     {"class":cn,"is_ct":False,"fixed_period":None,
                      "day_set":set(s.get("days_pref",[]) or day_names),
                      "period_prefs":list(s.get("periods_pref",[]))})

    for teacher, slots in teacher_slots.items():
        for i in range(len(slots)):
            for j in range(i+1, len(slots)):
                a, b = slots[i], slots[j]
                if a["class"]==b["class"]: continue
                if not (a["day_set"] & b["day_set"]): continue
                if a["is_ct"] and b["period_prefs"] and a["fixed_period"] in b["period_prefs"]:
                    hc.append({"teacher":teacher,
                               "reason":(f"CT period {a['fixed_period']} in {a['class']} "
                                         f"conflicts with pref for '{b.get('subj_name','')}' "
                                         f"in {b['class']}")})
                    log.warning("_step2_validate: hard conflict %s: %s",
                                teacher, hc[-1]["reason"])

    vr = {"ok": not errors and not hc,
          "period_errors": errors, "period_ok": ok_list, "hard_conflicts": hc}
    st.session_state["s2_validation_result"] = vr
    log.info("_step2_validate: errors=%d hc=%d ok=%s", len(errors), len(hc), vr["ok"])

    if vr["ok"]:
        _notify("✅ Validation passed.", "success")
        _nav("step3")
    else:
        _notify(f"❌ {len(errors)} period error(s), {len(hc)} conflict(s) — see below.", "error")
        st.rerun()


def _display_s2_validation(vr):
    if vr.get("period_errors"):
        with st.expander(f"❌ {len(vr['period_errors'])} Period Error(s)", expanded=True):
            for cn, msg in vr["period_errors"]: st.error(f"**{cn}**: {msg}")
    if vr.get("hard_conflicts"):
        with st.expander(f"⚠ {len(vr['hard_conflicts'])} Conflict(s)", expanded=True):
            for c in vr["hard_conflicts"]: st.warning(f"**{c['teacher']}**: {c['reason']}")
    if vr.get("period_ok"):
        with st.expander(f"✓ {len(vr['period_ok'])} Class(es) OK", expanded=False):
            for cn, msg in vr["period_ok"]: st.success(f"**{cn}**: {msg}")


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 3
# ═════════════════════════════════════════════════════════════════════════════
def page_step3():
    log.info("page_step3: render")
    _header("⚙ Step 3: Teacher Settings",
            "Review workload, define combined classes and mark teacher unavailability.")
    _show_notifications()

    cfg       = eng.configuration
    ppd       = cfg["periods_per_day"]
    wdays     = cfg["working_days"]
    teachers  = sorted(cfg.get("teacher_names",[]))
    day_names = DAY_NAMES[:wdays]
    all_cn    = _all_classes()

    if not hasattr(eng,"step3_data"):           eng.step3_data = {}
    if not hasattr(eng,"step3_unavailability"): eng.step3_unavailability = {}

    with st.expander("💾 Save / Load Step 3 Config", expanded=False):
        c1,c2 = st.columns(2)
        with c1:
            if st.button("⬇ Prepare Download", key="s3_dl_btn", use_container_width=True):
                log.info("page_step3: preparing download")
                st.session_state["_s3_dl_data"] = {
                    "step3_data": {t:{"skipped":v.get("skipped",False),"combines":v.get("combines",[])}
                                   for t,v in eng.step3_data.items()},
                    "step3_unavailability": {t:{"days":list(v["days"]),"periods":list(v["periods"])}
                                             for t,v in eng.step3_unavailability.items()},
                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "step":3,
                }
            if "_s3_dl_data" in st.session_state:
                _json_download(st.session_state["_s3_dl_data"], "📥 Click to Download",
                               f"Step3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with c2:
            up = st.file_uploader("Upload Step 3 JSON", type=["json"], key="s3_load_json")
            if up is not None:
                raw = up.read()
                log.debug("page_step3: upload size=%d", len(raw))
                if len(raw)==0:
                    log.warning("page_step3: 0 bytes, skipping")
                elif not _already_processed("s3_load_json", raw):
                    _load_step3_config(raw)

    cb, cv, cc = st.columns(3)
    with cb:
        if st.button("← Back to Step 2"): _nav("step2")
    with cv:
        if st.button("🔍 Validate Step 3"):
            log.info("page_step3: validate")
            try:
                eng._compute_teacher_workload()
                vr = eng.validate_step3()
                st.session_state["s3_validation_result"] = vr
                log.info("page_step3: validate result can_proceed=%s issues=%d",
                         vr["can_proceed"], len(vr["issues"]))
                if vr["can_proceed"]: _notify("✅ All clear.", "success")
                else: _notify(f"❌ {len(vr['issues'])} overload(s) unresolved.", "error")
            except Exception as ex:
                log.error("page_step3 validate: %s\n%s", ex, traceback.format_exc())
                _notify(f"Validation error: {ex}", "error")
            st.rerun()
    with cc:
        if st.button("➤ Proceed to Step 4 →", type="primary"):
            log.info("page_step3: proceed to step4")
            try:
                eng._compute_teacher_workload()
                vr = eng.validate_step3()
                if vr["can_proceed"]:
                    _nav("step4")
                else:
                    log.warning("page_step3: blocked, %d overloads", len(vr["issues"]))
                    _notify("Fix or skip overloaded teachers first.", "error"); st.rerun()
            except Exception as ex:
                log.error("page_step3 proceed: %s\n%s", ex, traceback.format_exc())
                _notify(f"Error: {ex}", "error"); st.rerun()

    vr = st.session_state.get("s3_validation_result")
    if vr:
        if vr["can_proceed"]: st.success("✅ All overloads resolved or skipped.")
        if vr["issues"]:
            with st.expander("❌ Still Overloaded", expanded=True):
                for ln in vr["issues"]: st.error(ln)
        if vr["resolved"]:
            with st.expander("✓ Resolved / Skipped"):
                for ln in vr["resolved"]: st.success(ln)

    st.divider()
    tab_wl, tab_cb, tab_un = st.tabs(
        ["📊 Teacher Workload","🔗 Combine Classes","🚫 Unavailability"])
    with tab_wl: _render_workload(teachers, wdays, ppd)
    with tab_cb: _render_combine_tab(teachers, all_cn)
    with tab_un: _render_unavailability_tab(teachers, day_names, ppd)


def _load_step3_config(raw: bytes):
    log.info("_load_step3_config: %d bytes", len(raw))
    try:
        d = json.loads(raw.decode("utf-8"))
        eng.step3_data = d.get("step3_data",{})
        eng.step3_unavailability = {
            t:{"days":v.get("days",[]),"periods":v.get("periods",[])}
            for t,v in d.get("step3_unavailability",{}).items()
        }
        _notify("✓ Step 3 config loaded.", "success")
        log.info("_load_step3_config: %d teachers, %d unavail",
                 len(eng.step3_data), len(eng.step3_unavailability))
        st.rerun()
    except Exception as ex:
        log.error("_load_step3_config: %s\n%s", ex, traceback.format_exc())
        _notify(f"Failed: {ex}", "error")


def _render_workload(teachers, wdays, ppd):
    log.debug("_render_workload: %d teachers", len(teachers))
    max_all = wdays * ppd
    all_cn  = _all_classes()
    for teacher in teachers:
        total, details = 0, []
        for cn in all_cn:
            cd = eng.class_config_data.get(cn,{})
            for s in cd.get("subjects",[]):
                if (s.get("teacher","").strip()==teacher or
                        s.get("parallel_teacher","").strip()==teacher):
                    total += s.get("periods",0)
                    details.append(f"{cn}: {s['name']} ×{s['periods']}")
        s3d  = eng.step3_data.get(teacher,{})
        skip = s3d.get("skipped",False)
        over = total > max_all
        icon = "🔴" if (over and not skip) else "✅"
        with st.expander(f"{icon} **{teacher}** — {total} periods"+(f" (SKIP)" if skip else ""),
                         expanded=False):
            if over and not skip:
                st.warning(f"Overloaded by {total-max_all} (max {max_all})")
            for d in details: st.write(f"  • {d}")
            new_skip = st.checkbox("Skip / Accept overload", value=skip, key=f"skip_{teacher}")
            if new_skip != skip:
                log.info("_render_workload: %s skip→%s", teacher, new_skip)
                eng.step3_data.setdefault(teacher,{})["skipped"] = new_skip
                st.rerun()


def _render_combine_tab(teachers, all_cn):
    log.debug("_render_combine_tab")
    st.markdown("**Combine classes** — one teacher, different classes, same period slot.")
    with st.container(border=True):
        st.markdown("**Add Combine**")
        c1,c2 = st.columns(2)
        with c1: sel_t       = st.selectbox("Teacher", [""]+teachers, key="cb_teacher")
        with c2: sel_classes = st.multiselect("Classes", all_cn, key="cb_classes")

        sub_map = {}
        if sel_t and sel_classes:
            st.markdown("*Select subject per class:*")
            cols = st.columns(min(len(sel_classes),4))
            for idx, cn in enumerate(sel_classes):
                with cols[idx%4]:
                    opts = [s["name"] for s in eng.class_config_data.get(cn,{}).get("subjects",[])]
                    sub_map[cn] = st.selectbox(f"{cn}", [""]+opts, key=f"cb_sub_{cn}")

        if st.button("➕ Add Combine", key="cb_add", type="primary"):
            if not sel_t: _notify("Select teacher.","warning")
            elif len(sel_classes)<2: _notify("Select ≥ 2 classes.","warning")
            elif not all(sub_map.get(cn) for cn in sel_classes): _notify("Select subject for each class.","warning")
            else:
                eng.step3_data.setdefault(sel_t,{"skipped":False,"combines":[]})
                eng.step3_data[sel_t]["combines"].append(
                    {"classes":sel_classes,"subjects":[sub_map[cn] for cn in sel_classes]})
                log.info("_render_combine_tab: combine added for %s: %s", sel_t, sel_classes)
                _notify(f"✓ Combine added for {sel_t}.","success"); st.rerun()

    st.markdown("**Existing Combines**")
    any_cb = False
    for teacher, s3d in sorted(eng.step3_data.items()):
        cbs = s3d.get("combines",[])
        if not cbs: continue
        any_cb = True
        st.markdown(f"**{teacher}**")
        for ci, cb in enumerate(cbs):
            c1,c2 = st.columns([5,1])
            with c1: st.write(f"  📌 {' + '.join(cb.get('classes',[]))}  ·  {', '.join(cb.get('subjects',[]))}")
            with c2:
                if st.button("🗑", key=f"del_cb_{teacher}_{ci}"):
                    log.info("_render_combine_tab: del %d for %s", ci, teacher)
                    cbs.pop(ci); st.rerun()
    if not any_cb: st.info("No combines defined.")


def _render_unavailability_tab(teachers, day_names, ppd):
    log.debug("_render_unavailability_tab")
    unavail = eng.step3_unavailability
    with st.container(border=True):
        st.markdown("**Add / Update Unavailability**")
        sel_t = st.selectbox("Teacher", [""]+teachers, key="un_teacher")
        c1,c2 = st.columns(2)
        with c1: sel_days    = st.multiselect("Unavailable Days",    day_names,           key="un_days")
        with c2: sel_periods = st.multiselect("Unavailable Periods", list(range(1,ppd+1)), key="un_periods")
        cs,ck,cl = st.columns(3)
        with cs:
            if st.button("✓ Save", key="un_save"):
                if not sel_t: _notify("Select teacher.","warning")
                elif not sel_days or not sel_periods: _notify("Select days and periods.","warning")
                else:
                    ok, msg = eng._check_unavailability_feasible(sel_t, sel_days, sel_periods)
                    if not ok:
                        log.warning("_render_unavailability_tab: feasibility fail %s: %s", sel_t, msg)
                        _notify(f"❌ Feasibility failed: {msg}","error")
                    else:
                        unavail[sel_t] = {"days":sel_days,"periods":sel_periods}
                        log.info("_render_unavailability_tab: saved %s days=%s", sel_t, sel_days)
                        _notify(f"✓ Saved for {sel_t}.","success"); st.rerun()
        with ck:
            if st.button("🔍 Check Only", key="un_check"):
                if sel_t and sel_days and sel_periods:
                    ok,msg = eng._check_unavailability_feasible(sel_t,sel_days,sel_periods)
                    log.info("_render_unavailability_tab: check %s → ok=%s", sel_t, ok)
                    if ok: _notify(f"✓ Feasible: {msg}","success")
                    else:  _notify(f"❌ {msg}","error")
                    st.rerun()
        with cl:
            if st.button("✕ Clear", key="un_clear"): st.rerun()
    _show_notifications()
    st.markdown("**Current Unavailability**")
    if not unavail:
        st.info("No unavailability set.")
    else:
        for teacher, info in sorted(unavail.items()):
            ok,short = eng._check_unavailability_feasible(teacher,info.get("days",[]),info.get("periods",[]))
            days_str = ", ".join(info.get("days",[]))
            pers_str = ", ".join(f"P{p}" for p in sorted(info.get("periods",[])))
            with st.expander(f"{'🟢' if ok else '🔴'} **{teacher}** — {days_str}  |  {pers_str}", expanded=False):
                st.write(f"Days: {days_str}  ·  Periods: {pers_str}")
                if not ok: st.error(short)
                if st.button("🗑 Remove", key=f"del_un_{teacher}"):
                    log.info("_render_unavailability_tab: remove %s", teacher)
                    del unavail[teacher]; st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  STEP 4
# ═════════════════════════════════════════════════════════════════════════════
def page_step4():
    log.info("page_step4: render stage=%d", st.session_state.get("s4_stage",0))
    cfg   = eng.configuration
    ppd   = cfg["periods_per_day"]
    wdays = cfg["working_days"]
    _header("📅 Step 4: Generate Timetable",
            f"{wdays} days/week · {ppd} periods/day · {wdays*ppd} slots/week per class")
    _show_notifications()
    if st.button("← Back to Step 3"): _nav("step3")
    st.divider()

    stage   = st.session_state.get("s4_stage",0)
    s1_stat = st.session_state.get("s4_s1_status")

    with st.container(border=True):
        st.markdown("**Stage 1 — HC1/HC2: Place Class-Teacher & Fixed/Preference Periods**")
        if stage == 0:
            if st.button("▶ Run Stage 1", type="primary", key="s4_run_s1"):
                log.info("page_step4: Run Stage 1")
                with st.spinner("Running Stage 1…"):
                    try:
                        result = eng.run_stage1()
                        st.session_state.update({"s4_s1_status":result,"s4_stage":1})
                        eng._relaxed_consec_keys = set()
                        eng._relaxed_main_keys   = set()
                        log.info("page_step4: Stage 1 done has_issues=%s", result.get("has_issues"))
                        _notify("Stage 1 complete.","success")
                    except Exception as ex:
                        log.error("page_step4 stage1: %s\n%s", ex, traceback.format_exc())
                        _notify(f"Stage 1 error: {ex}","error")
                st.rerun()
        else:
            if s1_stat:
                bg   = s1_stat.get("stage_bg","#1a7a1a")
                stxt = s1_stat.get("stage_txt","Stage 1 done")
                st.markdown(f"<div style='background:{bg};color:white;padding:8px 16px;"
                            f"border-radius:4px;font-weight:bold'>{stxt}</div>",
                            unsafe_allow_html=True)
                st.info(s1_stat.get("status",""))
            c1,c2 = st.columns(2)
            with c1:
                if st.button("📋 Task Analysis →", type="primary", key="s4_ta_btn"):
                    log.info("page_step4: → task_analysis")
                    eng._relaxed_consec_keys = set(st.session_state.get("relaxed_consec",set()))
                    eng._relaxed_main_keys   = set(st.session_state.get("relaxed_main",set()))
                    _nav("task_analysis")
            with c2:
                if st.button("↺ Re-run Stage 1", key="s4_rerun_s1"):
                    log.info("page_step4: re-run Stage 1")
                    for k in ("s4_stage","s4_s1_status","ta_allocation","ta2_allocation","s4_s3_status"):
                        st.session_state[k] = 0 if k=="s4_stage" else None
                    st.rerun()

    if stage >= 1 and eng._timetable:
        st.divider()
        st.markdown("**Stage 1 Preview**")
        _render_timetable_tabs(eng._timetable, key_prefix="s4")
        st.divider()
        st.markdown("**Export Stage 1 snapshot**")
        c1,c2 = st.columns(2)
        with c1: _excel_download("class","📥 Class Timetables")
        with c2: _excel_download("teacher","📥 Teacher Timetables")


# ═════════════════════════════════════════════════════════════════════════════
#  TASK ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════
def page_task_analysis():
    log.info("page_task_analysis: render")
    _header("📋 Task Analysis","Review groups before Stage 2. Allocate period slots.")
    _show_notifications()
    nav1,nav2 = st.columns(2)
    with nav1:
        if st.button("← Back to Stage 1"): _nav("step4")
    with nav2:
        if st.button("🗓 Allocate Periods", type="primary", key="ta_alloc_btn"):
            log.info("page_task_analysis: allocating")
            with st.spinner("Allocating…"):
                try:
                    slots, allocation, rows = eng._run_task_analysis_allocation()
                    st.session_state.update({"ta_allocation":allocation,"ta_group_slots":slots,"ta_all_rows":rows})
                    eng._last_allocation  = allocation
                    eng._last_group_slots = slots
                    eng._last_all_rows    = rows
                    ok_n = sum(1 for ar in allocation.values() if ar.get("ok"))
                    log.info("page_task_analysis: done %d ok / %d total", ok_n, len(allocation))
                    _notify(f"✓ {ok_n} OK, {len(allocation)-ok_n} failed.",
                            "success" if ok_n==len(allocation) else "warning")
                except Exception as ex:
                    log.error("page_task_analysis: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Allocation error: {ex}","error")
            st.rerun()

    allocation  = st.session_state.get("ta_allocation")
    group_slots = st.session_state.get("ta_group_slots")
    all_rows    = st.session_state.get("ta_all_rows")

    if allocation:
        all_ok = all(ar.get("ok",False) for ar in allocation.values())
        if all_ok:
            if st.button("▶ Proceed to Stage 2 →", type="primary", key="ta_proceed"):
                log.info("page_task_analysis: → task_analysis2")
                _nav("task_analysis2")
        else:
            st.error("Some groups failed — relax constraints or fix Step 2.")

    _show_notifications()
    if all_rows is None:
        st.info("Click **Allocate Periods** to compute slot assignments.")
        try:
            _, _, rows = eng._run_task_analysis_allocation()
            _render_ta_table(rows, None, None)
        except Exception as ex:
            log.warning("page_task_analysis: preview error: %s", ex)
    else:
        _render_ta_table(all_rows, group_slots, allocation)


def _render_ta_table(all_rows, group_slots, allocation):
    log.debug("_render_ta_table: %d rows", len(all_rows) if all_rows else 0)
    if not all_rows:
        st.info("No parallel/combined/consecutive groups found."); return
    DAYS_A = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    sections = {"A":"Combined Groups","B":"Standalone Parallel Pairs","C":"Consecutive Groups"}
    for sec, sec_title in sections.items():
        sec_rows = [r for r in all_rows if r.get("section")==sec]
        if not sec_rows: continue
        st.markdown(f"#### {sec_title}")
        groups = {}
        for r in sec_rows: groups.setdefault(r["group"],[]).append(r)
        for gn, grows in sorted(groups.items()):
            first = grows[0]
            cls_s = ", ".join(r["class"] for r in grows)
            slot_info = ""
            if group_slots and gn in group_slots:
                gs = group_slots[gn]
                slot_info = f"  ·  **{gs['slots']} slot(s)**" if gs.get("ok") else f"  ·  ⚠ {gs.get('reason','?')}"
            alloc_info = ""; alloc_ok = False
            if allocation and gn in allocation:
                ar = allocation[gn]; alloc_ok = ar.get("ok",False)
                if alloc_ok:
                    placed = ar.get("placed", ar.get("slots",[]))
                    ps = "  ·  ".join(f"{DAYS_A[d]} P{p+1}" for d,p in sorted(placed)) if placed else "placed"
                    alloc_info = f"✅ {ps}"
                else:
                    alloc_info = f"❌ {ar.get('remaining','?')} unplaced. {ar.get('reason','')}"
            icon = "🟢" if (allocation and alloc_ok) else ("🔴" if allocation else "⚪")
            with st.expander(f"{icon} **Group {gn}** — {cls_s} · {first['subject']} / {first['teacher']}{slot_info}", expanded=False):
                for r in grows:
                    par = (f"  ‖  Parallel: {r['par_subj']} / {r['par_teacher']}"
                           if r.get("par_subj") not in ("—","?","",None) else "")
                    st.write(f"  📌 **{r['class']}** — {r['subject']} / {r['teacher']}{par}")
                if alloc_info:
                    if "✅" in alloc_info: st.success(alloc_info)
                    else:                  st.error(alloc_info)
                if sec=="C":
                    rk = (first["class"], first["subject"])
                    rel = st.session_state.get("relaxed_consec",set())
                    is_rel = rk in rel
                    if st.button("🔓 Un-relax" if is_rel else "🔒 Relax to Filler", key=f"relax_c_{gn}"):
                        for r in grows:
                            k = (r["class"],r["subject"])
                            if is_rel: rel.discard(k); eng._relaxed_consec_keys.discard(k)
                            else:      rel.add(k);     eng._relaxed_consec_keys.add(k)
                        st.session_state["relaxed_consec"] = rel
                        log.info("_render_ta_table: group %s relax=%s", gn, not is_rel)
                        st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  TASK ANALYSIS 2
# ═════════════════════════════════════════════════════════════════════════════
def page_task_analysis2():
    log.info("page_task_analysis2: render")
    _header("📋 Task Analysis — Stage 2","Allocate remaining slots before Stage 3.")
    _show_notifications()
    eng._relaxed_consec_keys = set(st.session_state.get("relaxed_consec",set()))
    eng._relaxed_main_keys   = set(st.session_state.get("relaxed_main",set()))
    nav1,nav2 = st.columns(2)
    with nav1:
        if st.button("← Back to Task Analysis"): _nav("task_analysis")
    with nav2:
        if st.button("🗓 Allocate Slots", type="primary", key="ta2_alloc_btn"):
            log.info("page_task_analysis2: allocating")
            with st.spinner("Allocating Stage 2 slots…"):
                try:
                    result = eng._run_ta2_allocation()
                    st.session_state["ta2_allocation"] = result
                    eng._last_ta2_allocation = result
                    ok_n = sum(1 for ar in result.values() if isinstance(ar,dict) and ar.get("remaining",1)==0)
                    log.info("page_task_analysis2: done %d ok", ok_n)
                    _notify(f"✓ Stage 2 done: {ok_n} groups OK.","success")
                except Exception as ex:
                    log.error("page_task_analysis2: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Error: {ex}","error")
            st.rerun()

    ta2 = st.session_state.get("ta2_allocation")
    if ta2:
        fails = [k for k,ar in ta2.items() if isinstance(ar,dict) and ar.get("remaining",1)>0]
        if not fails:
            st.success("✅ All groups allocated.")
            if st.button("📅 Proceed to Stage 3 →", type="primary", key="ta2_proceed"):
                log.info("page_task_analysis2: → stage2_page")
                _nav("stage2_page")
        else:
            st.warning(f"⚠ {len(fails)} group(s) have unplaced periods.")
            if st.button("📅 Proceed to Stage 3 anyway →", key="ta2_proceed_any"):
                log.info("page_task_analysis2: proceed with %d failures", len(fails))
                _nav("stage2_page")
        _render_ta2_table(ta2)
    else:
        st.info("Click **Allocate Slots** to compute allocations.")


def _render_ta2_table(result):
    log.debug("_render_ta2_table: %d groups", len(result) if result else 0)
    if not result: return
    DAYS_A = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    def slot_str(placed):
        return "  ·  ".join(f"{DAYS_A[d]} P{p+1}" for d,p in sorted(placed)) if placed else "—"
    for gn, ar in sorted(result.items()):
        if not isinstance(ar,dict): continue
        placed    = ar.get("slots", ar.get("placed",[]))
        remaining = ar.get("remaining",0)
        ok        = remaining == 0
        with st.expander(f"{'✅' if ok else '❌'} Group {gn}", expanded=not ok):
            c1,c2 = st.columns(2)
            with c1:
                st.metric("Total",          ar.get("total","?"))
                st.metric("Stage 1 placed", ar.get("s1_placed","?"))
            with c2:
                st.metric("Stage 2 placed", ar.get("new_placed","?"))
                st.metric("Remaining",      remaining)
            if placed: st.success(f"Slots: {slot_str(placed)}")
            if not ok:
                st.error(f"Reason: {ar.get('reason','unknown')}")
                rel = st.session_state.get("relaxed_main",set()); k_str = str(gn)
                if st.button("🔓 Un-relax" if k_str in rel else "🔒 Relax to Filler", key=f"relax_m_{gn}"):
                    rel.discard(k_str) if k_str in rel else rel.add(k_str)
                    st.session_state["relaxed_main"] = rel
                    log.info("_render_ta2_table: group %s main-relax=%s", gn, k_str in rel)
                    st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
#  STAGE 2 PAGE
# ═════════════════════════════════════════════════════════════════════════════
def _render_force_fill_summary(ff_result):
    """Render the Force Fill result summary — mirrors the tkinter scrollable dialog."""
    remaining   = ff_result.get("remaining", 0)
    relaxed     = ff_result.get("relaxed")
    overloaded  = ff_result.get("overloaded", [])
    blocked     = ff_result.get("blocked_only", [])
    wdays       = ff_result.get("wdays", "?")
    ppd         = ff_result.get("ppd", "?")
    total_slots = ff_result.get("total_slots", "?")

    if remaining == 0:
        st.success("✅ All periods placed — timetable is 100% complete!")
        if relaxed:
            with st.expander("ℹ️ Constraints relaxed during Force Fill", expanded=True):
                st.code(relaxed, language="")
        return

    st.error(f"⚠  {remaining} period(s) could not be placed.")

    if overloaded:
        st.markdown("#### ❌ Overloaded Teachers")
        st.caption(f"Total grid capacity: {wdays} days × {ppd} periods = {total_slots} slots per teacher")
        for tname, assigned, cap, excess, unp in overloaded:
            with st.container(border=True):
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.markdown(f"**❌ {tname}**")
                with col2:
                    st.markdown(
                        f"Assigned: **{assigned}** &nbsp;|&nbsp; "
                        f"Capacity: **{cap}** &nbsp;|&nbsp; "
                        f"Excess: **+{excess}** ← fix in Step 2 &nbsp;|&nbsp; "
                        f"Unplaced: **{unp}**"
                    )

    if blocked:
        st.markdown("#### ⚠ Blocked Teachers *(within capacity, but slots clash)*")
        for tname, assigned, cap, unp in blocked:
            with st.container(border=True):
                col1, col2 = st.columns([2, 3])
                with col1:
                    st.markdown(f"**⚠ {tname}**")
                with col2:
                    st.markdown(
                        f"Assigned: **{assigned} / {cap}** &nbsp;|&nbsp; Unplaced: **{unp}**"
                    )

    if not overloaded and not blocked:
        st.info("Could not identify a specific cause. Please review teacher workloads in Step 2.")

    if relaxed:
        with st.expander("ℹ️ Constraints relaxed during Force Fill", expanded=False):
            st.code(relaxed, language="")


def page_stage2():
    log.info("page_stage2: render")
    _header("⚙ Stage 2 — Fill Remaining Periods",
            "Stage 3 of engine: consecutive pairs, daily subjects, fillers and repair.")
    _show_notifications()
    if st.button("← Back to Task Analysis 2"): _nav("task_analysis2")

    s3_status  = st.session_state.get("s4_s3_status")
    ff_result  = st.session_state.get("s4_ff_result")   # Force Fill result

    # ── Phase 1: Run Stage 3 ──────────────────────────────────────────────────
    if s3_status is None:
        st.info("Click **Run Stage 3** to fill all remaining empty slots.")
        if st.button("▶ Run Stage 3", type="primary", key="s2pg_run"):
            log.info("page_stage2: Run Stage 3")
            with st.spinner("Running Stage 3 — filling remaining periods…"):
                try:
                    result = eng.run_stage3()
                    st.session_state.update({"s4_s3_status": result, "s4_stage": 3,
                                             "s4_ff_result": None})
                    log.info("page_stage2: Stage 3 done ok=%s unplaced=%s",
                             result.get("ok"), result.get("unplaced"))
                    _notify(result.get("msg", "Stage 3 complete."),
                            "success" if result.get("ok") else "warning")
                except Exception as ex:
                    log.error("page_stage2 stage3: %s\n%s", ex, traceback.format_exc())
                    _notify(f"Stage 3 error: {ex}", "error")
            st.rerun()
        return

    # ── Phase 2: Stage 3 done — show status + action buttons ─────────────────
    if s3_status.get("ok"):
        st.success(s3_status.get("msg", "✅ Complete!"))
    else:
        st.warning(s3_status.get("msg", ""))

    # Show Force Fill option only when there are still unplaced periods
    unplaced_after_s3 = s3_status.get("unplaced", 0)

    col_view, col_rerun, col_ff = st.columns([2, 1, 2])
    with col_view:
        if st.button("📊 View Final Timetable →", type="primary", key="s2pg_view"):
            log.info("page_stage2: → final_timetable")
            _nav("final_timetable")
    with col_rerun:
        if st.button("↺ Re-run Stage 3", key="s2pg_rerun"):
            log.info("page_stage2: re-run")
            st.session_state["s4_s3_status"] = None
            st.session_state["s4_ff_result"] = None
            st.rerun()
    with col_ff:
        if unplaced_after_s3 > 0:
            if st.button("🔧 Force Fill", type="secondary", key="s2pg_ff",
                         help="Min-Conflicts CSP solver — up to 1500 iterations, "
                              "stops as soon as all periods are placed."):
                log.info("page_stage2: running Force Fill")
                progress_placeholder = st.empty()
                progress_msgs = []

                def _progress(msg):
                    progress_msgs.append(msg)
                    if msg:
                        progress_placeholder.info(f"⏳ {msg}")

                with st.spinner("🔧 Force Fill running…"):
                    try:
                        ff = eng.run_force_fill(progress_cb=_progress)
                        st.session_state["s4_ff_result"] = ff
                        # Update s3_status so unplaced count is refreshed
                        st.session_state["s4_s3_status"] = {
                            **s3_status,
                            "unplaced": ff["remaining"],
                            "ok":       ff["ok"],
                            "msg":      ff.get("msg", ""),
                        }
                        log.info("page_stage2: Force Fill done ok=%s remaining=%s",
                                 ff["ok"], ff["remaining"])
                        _notify(
                            "✅ Force Fill complete — all periods placed!" if ff["ok"]
                            else f"⚠ Force Fill done — {ff['remaining']} period(s) still unplaced.",
                            "success" if ff["ok"] else "warning",
                        )
                    except Exception as ex:
                        log.error("page_stage2 force_fill: %s\n%s", ex, traceback.format_exc())
                        _notify(f"Force Fill error: {ex}", "error")
                progress_placeholder.empty()
                st.rerun()

    # ── Phase 3: Show Force Fill summary if available ─────────────────────────
    if ff_result is not None:
        st.divider()
        st.markdown("### 🔧 Force Fill Summary")
        _render_force_fill_summary(ff_result)


# ═════════════════════════════════════════════════════════════════════════════
#  FINAL TIMETABLE
# ═════════════════════════════════════════════════════════════════════════════
def page_final_timetable():
    log.info("page_final_timetable: render")
    _header("📊 Final Timetable","View and export the complete timetable.")
    _show_notifications()
    if st.button("← Back to Stage 2"): _nav("stage2_page")
    tt = eng._timetable
    if not tt:
        st.error("No timetable generated."); log.warning("page_final_timetable: no timetable"); return
    with st.container(border=True):
        st.markdown("**📥 Export to Excel**")
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1: _excel_download("class",    "📥 Classwise")
        with c2: _excel_download("teacher",  "📥 Teacherwise")
        with c3: _excel_download("ct_list",  "📥 CT List")
        with c4: _excel_download("workload", "📥 Workload")
        with c5: _excel_download("one_sheet","📥 One-Sheet")
    st.divider()
    tc, tt2, ts = st.tabs(["🏫 Classwise","👨‍🏫 Teacherwise","📋 Summary"])
    with tc:  _render_class_view(tt)
    with tt2: _render_teacher_view(tt)
    with ts:  _render_summary_view(tt)


# ─────────────────────────────────────────────────────────────────────────────
#  Timetable renderers
# ─────────────────────────────────────────────────────────────────────────────
def _render_timetable_tabs(tt, key_prefix="tt"):
    tc, tt2 = st.tabs(["🏫 Class View","👨‍🏫 Teacher View"])
    with tc:  _render_class_view(tt,   key_prefix=key_prefix+"_c")
    with tt2: _render_teacher_view(tt, key_prefix=key_prefix+"_t")


def _render_class_view(tt, key_prefix="cls"):
    import pandas as pd
    log.debug("_render_class_view: key=%s", key_prefix)
    all_classes = tt["all_classes"]; days = tt["days"]; ppd = tt["ppd"]
    half1 = tt["half1"];             grid = tt["grid"]
    sel_cn = st.selectbox("Select Class", all_classes, key=f"{key_prefix}_sel")
    if not sel_cn: return
    header = ["Day"] + [f"P{p+1}{'①' if p<half1 else '②'}" for p in range(ppd)]
    rows = []
    for d, dname in enumerate(days):
        row = [dname]
        for p in range(ppd):
            g = grid.get(sel_cn,[])
            cell = g[d][p] if d<len(g) and g else None
            if cell:
                row.append(f"{'★' if cell.get('is_ct') else ''}{cell.get('subject','')} / {cell.get('teacher','')}")
            else:
                row.append("—")
        rows.append(row)
    st.dataframe(pd.DataFrame(rows, columns=header), use_container_width=True, hide_index=True)


def _render_teacher_view(tt, key_prefix="tch"):
    import pandas as pd
    log.debug("_render_teacher_view: key=%s", key_prefix)
    all_classes = tt["all_classes"]; days = tt["days"]; ppd = tt["ppd"]; grid = tt["grid"]
    tg = {}
    for cn in all_classes:
        g = grid.get(cn,[])
        for d in range(len(days)):
            if d >= len(g): continue
            for p in range(ppd):
                cell = g[d][p]
                if not cell: continue
                for tname, sname in [(cell.get("teacher",""), cell.get("subject","")),
                                     (cell.get("par_teach",""), cell.get("par_subj",""))]:
                    if not tname or tname in ("—","?"): continue
                    tg.setdefault(tname, [[None]*ppd for _ in range(len(days))])
                    tg[tname][d][p] = {"class":cn,"subject":sname,"is_ct":cell.get("is_ct",False)}
    tlist = sorted(tg.keys())
    if not tlist: st.info("No teacher data."); return
    sel_t = st.selectbox("Select Teacher", tlist, key=f"{key_prefix}_sel")
    if not sel_t: return
    trows = tg.get(sel_t,[])
    header = ["Day"] + [f"P{p+1}" for p in range(ppd)]
    rows = []
    for d, dname in enumerate(days):
        row = [dname]
        for p in range(ppd):
            cell = trows[d][p] if d<len(trows) else None
            row.append(f"{'★' if cell and cell.get('is_ct') else ''}"
                       f"{cell['class']} / {cell['subject']}" if cell else "FREE")
        rows.append(row)
    st.dataframe(pd.DataFrame(rows, columns=header), use_container_width=True, hide_index=True)


def _render_summary_view(tt):
    log.debug("_render_summary_view")
    tasks=tt.get("tasks",[]); days=tt["days"]; ppd=tt["ppd"]
    half1=tt["half1"]; grid=tt["grid"]; all_classes=tt["all_classes"]
    unplaced = [t for t in tasks if t.get("remaining",0)>0]
    if unplaced:
        st.error(f"**{len(unplaced)} task(s) with unplaced periods:**")
        for t in unplaced:
            st.write(f"  ❌ {'+'.join(t.get('cn_list',[]))} | {t['subject']} | "
                     f"{t['teacher']} — {t['remaining']} unplaced")
    else:
        st.success("✅ All periods placed.")
    st.markdown("**Teacher Free-Period Distribution**")
    all_teachers = sorted({cell.get("teacher","")
                            for cn in all_classes
                            for d_row in grid.get(cn,[])
                            for cell in d_row if cell and cell.get("teacher")})
    for teacher in all_teachers:
        busy = {}
        for cn in all_classes:
            for d, d_row in enumerate(grid.get(cn,[])):
                for p, cell in enumerate(d_row):
                    if cell and (cell.get("teacher")==teacher or cell.get("par_teach")==teacher):
                        busy.setdefault(d,set()).add(p)
        lines = []
        for d in range(len(days)):
            bd  = busy.get(d,set())
            fh1 = half1 - len([x for x in bd if x<half1])
            fh2 = (ppd-half1) - len([x for x in bd if x>=half1])
            if fh1+fh2 == ppd: continue
            lines.append((days[d], fh1, fh2, fh1>=1 and fh2>=1))
        if lines:
            with st.expander(f"**{teacher}**", expanded=False):
                for dname,fh1,fh2,ok in lines:
                    st.write(f"  {'✓' if ok else '⚠'} {dname} — free H1:{fh1}  H2:{fh2}")
    st.divider()
    st.caption("★=CT  ⊕=Combined  ∥=Parallel  ⊕∥=Combined+Parallel")


# ═════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🗓 Timetable Generator")
    st.caption("V4.1 — Streamlit Edition")
    st.divider()
    cur = st.session_state.page
    for pid, plabel in PAGE_LABELS.items():
        active = (pid == cur)
        if st.button(("▶ " if active else "   ") + plabel,
                     key=f"nav_{pid}", use_container_width=True, disabled=active):
            _nav(pid)
    st.divider()
    if eng.configuration:
        cfg = eng.configuration
        st.caption(f"**Config:** {cfg.get('working_days','?')}d × {cfg.get('periods_per_day','?')}p")
        st.caption(f"**Teachers:** {len(cfg.get('teacher_names',[]))}")
        st.caption(f"**Classes:** {sum(cfg.get('classes',{}).values())}")
    st.divider()
    with st.expander("🪵 Debug Log", expanded=False):
        st.caption("Shows last 200 log lines. Format: [LEVEL] file:func:line — msg")
        if st.button("Clear", key="clear_log"):
            _mem_handler.lines.clear(); st.rerun()
        log_text = "\n".join(_mem_handler.lines[-200:])
        st.code(log_text if log_text else "(no entries yet)", language="")


# ═════════════════════════════════════════════════════════════════════════════
#  ROUTER
# ═════════════════════════════════════════════════════════════════════════════
PAGES = {
    "step1":           page_step1,
    "step2":           page_step2,
    "step3":           page_step3,
    "step4":           page_step4,
    "task_analysis":   page_task_analysis,
    "task_analysis2":  page_task_analysis2,
    "stage2_page":     page_stage2,
    "final_timetable": page_final_timetable,
}

log.debug("Router: page=%s", st.session_state.page)
PAGES.get(st.session_state.page, page_step1)()

# ─────────────────────────────────────────────────────────────────────────────
#  FOOTER  (bottom-centre on every page)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    .dev-footer {
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
        text-align: center;
        padding: 6px 0 6px 0;
        font-size: 0.78rem;
        color: #888888;
        background: rgba(255,255,255,0.85);
        backdrop-filter: blur(4px);
        border-top: 1px solid #e0e0e0;
        z-index: 9999;
        letter-spacing: 0.01em;
    }
    </style>
    <div class="dev-footer">
        Developed by: <strong>Kanika Tanwar</strong>, TGT CS
    </div>
    """,
    unsafe_allow_html=True,
)
