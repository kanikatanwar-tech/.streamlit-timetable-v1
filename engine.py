"""
engine.py — TimetableEngine

Pure-Python scheduling engine (no tkinter).
All scheduling/generation logic from the original app, adapted for Streamlit:
  • run_stage1()       — synchronous Stage 1 (HC1/HC2 placement)
  • run_stage3()       — synchronous Stage 3 (filler + repair)
  • validate_step3()   — returns dict instead of showing a window
  • get_excel_bytes()  — returns Excel bytes for st.download_button
"""
import random
import copy
from collections import defaultdict
from datetime import datetime


class TimetableEngine:
    """Holds all application state and scheduling logic."""

    def __init__(self):
        self.configuration        = {}
        self.class_config_data    = {}
        self.step3_data           = {}
        self.step3_unavailability = {}
        self._relaxed_consec_keys = set()
        self._relaxed_main_keys   = set()
        self._gen_stage           = 0
        self._progress_log        = []
        self._gen                 = None
        self._timetable           = None
        self._last_allocation     = None
        self._last_all_rows       = None
        self._last_group_slots    = None
        self._last_ta2_allocation = None
        self._stage1_status       = None
        self._stage2_status       = None

    def _check_unavailability_feasible(self, teacher, blocked_days, blocked_periods):
        """Two-part feasibility check for teacher unavailability.

        CHECK 1 — Direct slot conflicts:
          For every subject assigned to this teacher, if the subject has specific
          period preferences AND specific day preferences, check whether ANY of
          those (day, period) pairs fall inside the blocked slots.
          Class-teacher duty is also checked: if the CT period falls on a blocked
          day+period combination that is a conflict.

        CHECK 2 — Total slot availability:
          After removing blocked slots, the remaining available slots per week
          must be >= teacher's effective assigned periods.

        Returns (ok: bool, message: str)
          ok=False if either check fails; message explains what is wrong.
        """
        cfg       = self.configuration
        ppd       = cfg['periods_per_day']
        wdays     = cfg['working_days']
        day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]

        blocked_days_set    = set(blocked_days)
        blocked_periods_set = set(int(p) for p in blocked_periods)

        # ── CHECK 1: Direct assignment conflicts ─────────────────────────
        slot_conflicts = []   # list of human-readable conflict strings

        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd  = self.class_config_data[cn]
                ct  = cd.get('teacher', '').strip()
                ct_per = cd.get('teacher_period', 1)

                # Class-teacher duty: fixed period, every working day
                if ct == teacher and ct_per in blocked_periods_set:
                    # Only a conflict on the blocked days
                    conflict_days = [d for d in day_names if d in blocked_days_set]
                    if conflict_days:
                        slot_conflicts.append(
                            "Class Teacher of {} (Period {}) conflicts on: {}".format(
                                cn, ct_per, ', '.join(conflict_days)))

                # Subject assignments
                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t != teacher:
                        continue
                    s_periods = s.get('periods_pref', [])
                    s_days    = s.get('days_pref', [])

                    if not s_periods and not s_days:
                        # No specific preference — no direct conflict detectable
                        continue

                    # Determine which days are relevant
                    relevant_days = set(s_days) if s_days else set(day_names)
                    conflict_days = relevant_days & blocked_days_set

                    if not conflict_days:
                        continue

                    if s_periods:
                        # Check if any preferred period is in the blocked set
                        bad_periods = set(s_periods) & blocked_periods_set
                        if bad_periods:
                            slot_conflicts.append(
                                "'{}' in {} — Period(s) {} on {} are both "
                                "preferred and blocked".format(
                                    s['name'], cn,
                                    sorted(bad_periods),
                                    ', '.join(sorted(conflict_days))))
                    # If subject has day pref but no period pref: warn only (soft)
                    else:
                        # Soft: some teaching days overlap with blocked days
                        # but no period specified — flag as warning
                        slot_conflicts.append(
                            "'{}' in {} — preferred days {} overlap with "
                            "blocked days (no period preference set, "
                            "scheduler may still place it in a blocked slot)".format(
                                s['name'], cn,
                                ', '.join(sorted(conflict_days))))

        # ── CHECK 2: Total available slots ──────────────────────────────
        total_week    = ppd * wdays
        blocked_total = len(blocked_days_set) * len(blocked_periods_set)
        available     = total_week - blocked_total

        wl = getattr(self, '_step3_teacher_wl', {})
        assigned = self._effective_total(teacher) if teacher in wl else 0

        slot_ok  = available >= assigned
        free     = available - assigned

        # ── Build result message ─────────────────────────────────────────
        parts = []

        if slot_conflicts:
            parts.append(
                "SLOT CONFLICTS ({}):\n{}".format(
                    len(slot_conflicts),
                    "\n".join("  • " + c for c in slot_conflicts)))

        parts.append(
            "CAPACITY: {} assigned, {} available after blocking "
            "({} blocked, {} free).".format(
                assigned, available, blocked_total, free))

        message = "\n".join(parts)

        ok = (not slot_conflicts) and slot_ok
        if not ok:
            if slot_conflicts and not slot_ok:
                message = "Slot conflicts AND capacity problem.\n" + message
            elif slot_conflicts:
                message = "Slot conflicts found (capacity OK).\n" + message
            else:
                message = "Capacity problem.\n" + message

        return (ok, message)


    def _compute_teacher_workload(self):
        """Compute teacher workload.

        NOTE: Class teacher period is already included in the subject's period count.
        For example if teacher A is class teacher of 8A and teaches English (7 periods),
        those 7 periods already include the class teacher period — so we do NOT add
        extra periods for class teacher duty. We only record it as metadata (is_ct=True
        on the subject entry) so the fixed-period constraint is known.
        """
        cfg   = self.configuration
        wdays = cfg['working_days']
        result = {}

        def _add(t, entry):
            if not t: return
            result.setdefault(t, {'total': 0, 'entries': []})
            result[t]['entries'].append(entry)
            result[t]['total'] += entry['periods']

        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                cn = "{}{}".format(cls, chr(65 + si))
                if cn not in self.class_config_data:
                    continue
                cd = self.class_config_data[cn]
                ct = cd.get('teacher', '').strip()
                ct_per = cd.get('teacher_period', 1)

                for s in cd['subjects']:
                    t = s['teacher'].strip()
                    if t:
                        # Mark if this teacher is also the class teacher for this class
                        is_ct_subject = (t == ct)
                        ct_note = "  [incl. CT Period {}]".format(ct_per) if is_ct_subject else ""
                        _add(t, {
                            'class':    cn,
                            'subject':  s['name'],
                            'label':    "'{}' in {}  x{}/wk{}".format(
                                s['name'], cn, s['periods'], ct_note),
                            'periods':  s['periods'],
                            'is_ct':    is_ct_subject,
                            'ct_period': ct_per if is_ct_subject else None,
                        })
                    pt = s['parallel_teacher'].strip() if s['parallel'] else ''
                    if pt:
                        _add(pt, {
                            'class':   cn,
                            'subject': s.get('parallel_subject', '?'),
                            'label':   "Parallel '{}' in {}  x{}/wk".format(
                                s.get('parallel_subject', '?'), cn, s['periods']),
                            'periods': s['periods'],
                            'is_ct':   False,
                            'ct_period': None,
                        })
        return result

    def _effective_total(self, teacher):
        """Total periods after subtracting savings from combines."""
        wl    = self._step3_teacher_wl.get(teacher, {})
        total = wl.get('total', 0)
        for cb in self.step3_data.get(teacher, {}).get('combines', []):
            n   = len(cb.get('entry_indices', []))
            per = cb.get('periods_each', 0)
            if n > 1:
                total -= (n - 1) * per
        return total

    # ── Left panel: ALL teachers ──────────────────────────────────────────
    def prepare_step3_workload(self):
        """Compute workload AND set all step-3 attributes needed by validate_step3.

        Must be called before validate_step3() or _render_workload().
        Sets:
          _step3_teacher_wl   – {teacher: {'total': int, 'entries': [...]}}
          _step3_overloaded   – set of teacher names whose raw total > max_allowed
          _step3_max_allowed  – (ppd - 2) * wdays  (mirrors original formula)
          _step3_total_week   – ppd * wdays

        Returns the workload dict (same as _compute_teacher_workload).
        """
        cfg          = self.configuration
        ppd          = cfg['periods_per_day']
        wdays        = cfg['working_days']
        total_week   = ppd * wdays
        max_allowed  = (ppd - 2) * wdays   # must keep ≥ 2 free periods per day

        wl = self._compute_teacher_workload()

        self._step3_teacher_wl  = wl
        self._step3_total_week  = total_week
        self._step3_max_allowed = max_allowed
        self._step3_overloaded  = {
            t for t, info in wl.items() if info['total'] > max_allowed
        }
        return wl

    def validate_step3(self):
        """Return dict with overload status; no UI.

        Call prepare_step3_workload() first so that _step3_overloaded and
        _step3_max_allowed are properly initialised.
        """
        overloaded = getattr(self, '_step3_overloaded', set())
        max_all    = getattr(self, '_step3_max_allowed', 99999)
        issues, resolved = [], []
        for teacher in sorted(overloaded):
            s3d     = self.step3_data.get(teacher, {})
            skipped = s3d.get('skipped', False)
            eff     = self._effective_total(teacher)
            if skipped:
                resolved.append("{}: SKIPPED by user".format(teacher))
            elif eff <= max_all:
                resolved.append("{}: Resolved  ({} periods  \u2264 {})".format(teacher, eff, max_all))
            else:
                issues.append("{}: still overloaded  ({}/{})  — over by {}".format(
                    teacher, eff, max_all, eff - max_all))
        return {
            'overloaded':   overloaded,
            'issues':       issues,
            'resolved':     resolved,
            'can_proceed':  (not overloaded) or (not issues),
        }

    def get_class_ct_info(self, cn, teacher, teacher_subject):
        """Return class-teacher info and parallel-conflict details for one entry.

        Adapted from the original _get_class_ct_info to use plain dict-based
        class_config_data (no tkinter StringVar).

        Returns a dict:
          ct                  – class teacher name (str)
          ct_subjects         – list[str] of subjects the CT teaches in cn
          is_parallel_with_ct – bool: teacher_subject is parallel to a CT subject
          parallel_ct_subject – str: the CT subject that is parallel ('' if none)
        """
        cd    = self.class_config_data.get(cn, {})
        ct    = cd.get('teacher', '').strip()
        subjs = cd.get('subjects', [])

        ct_subjects = [s['name'] for s in subjs
                       if s.get('teacher', '').strip() == ct]

        is_parallel_with_ct = False
        parallel_ct_subject = ''
        for s in subjs:
            # Primary: teacher teaches teacher_subject, parallel partner is the CT
            if (s.get('teacher', '').strip() == teacher
                    and s['name'] == teacher_subject
                    and s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s.get('parallel_subject', '')
                break
            # Reverse: teacher is the *parallel* teacher; CT teaches the primary
            if (s.get('parallel')
                    and s.get('parallel_teacher', '').strip() == teacher
                    and s.get('parallel_subject', '') == teacher_subject
                    and s.get('teacher', '').strip() == ct):
                is_parallel_with_ct = True
                parallel_ct_subject = s['name']
                break

        return {
            'ct':                   ct,
            'ct_subjects':          ct_subjects,
            'is_parallel_with_ct':  is_parallel_with_ct,
            'parallel_ct_subject':  parallel_ct_subject,
        }


    # =========================================================================
    #  STEP 4 — Timetable Generation Engine
    # =========================================================================


    # =========================================================================
    #  STEP 4 — Timetable Generation  (complete rewrite)
    # =========================================================================

    def run_stage1(self):
        """Run Stage 1 synchronously. Returns status dict."""
        self._progress_log = []
        self._init_gen_state()
        self._run_stage1_phases()
        return getattr(self, '_stage1_status', {})
    # ── Task Analysis page ────────────────────────────────────────────────────
    def _run_task_analysis_allocation(self):
        """
        Orchestrate the full allocation pipeline for the Task Analysis page:
          1. Build all_rows (same logic as _show_task_analysis data phase)
          2. Calculate slots needed per group (_calculate_group_slots)
          3. Allocate slots (_allocate_group_slots)

        Returns (group_slots, group_allocation, all_rows)
        so the caller can pass them directly to _show_task_analysis.
        """
        s3  = getattr(self, 'step3_data', {})
        cfg = self.configuration

        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'].get(cls, 0)):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        # Helper: find parallel partner
        def _find_parallel(cn, subject_name):
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if s['name'] == subject_name and s.get('parallel'):
                    ps = (s.get('parallel_subject') or '').strip()
                    pt = (s.get('parallel_teacher') or '').strip()
                    return (ps or '?', pt or '—')
                if (s.get('parallel')
                        and (s.get('parallel_subject') or '').strip() == subject_name
                        and s['name'] != subject_name):
                    return (s['name'], (s.get('teacher') or '').strip() or '—')
            return ('—', '—')

        # Section A — combined groups
        all_rows = []
        group_no = 0
        covered  = set()

        for teacher, s3d in sorted(s3.items()):
            for cb in s3d.get('combines', []):
                classes  = cb.get('classes', [])
                subjects = cb.get('subjects', [])
                if not classes:
                    continue
                group_no += 1
                for j, cn in enumerate(classes):
                    tsub = (subjects[j] if j < len(subjects)
                            else (subjects[0] if subjects else '?'))
                    par_subj, par_teacher = _find_parallel(cn, tsub)
                    all_rows.append({
                        'group': group_no, 'class': cn,
                        'subject': tsub, 'teacher': teacher,
                        'par_subj': par_subj, 'par_teacher': par_teacher,
                        'section': 'A',
                    })
                    covered.add((cn, tsub))
                    if par_subj not in ('—', '?'):
                        covered.add((cn, par_subj))

        # Section B — standalone parallel pairs
        seen_pairs = set()
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if not s.get('parallel'):
                    continue
                subj_name   = s['name']
                subj_teach  = (s.get('teacher') or '').strip()
                par_subj    = (s.get('parallel_subject') or '').strip()
                par_teacher = (s.get('parallel_teacher') or '').strip()
                if not par_subj:
                    continue
                if (cn, subj_name) in covered or (cn, par_subj) in covered:
                    continue
                pair_key = frozenset([(cn, subj_name), (cn, par_subj)])
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                group_no += 1
                all_rows.append({
                    'group': group_no, 'class': cn,
                    'subject': subj_name, 'teacher': subj_teach,
                    'par_subj': par_subj or '?',
                    'par_teacher': par_teacher or '—',
                    'section': 'B',
                })

        # Section C — consecutive groups
        consec_covered = set(covered)
        for row in all_rows:
            if row['section'] == 'B':
                consec_covered.add((row['class'], row['subject']))
                if row['par_subj'] not in ('—', '?', ''):
                    consec_covered.add((row['class'], row['par_subj']))

        seen_consec = set()
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                if s.get('consecutive', 'No') != 'Yes':
                    continue
                subj_name  = s['name']
                subj_teach = (s.get('teacher') or '').strip()
                periods    = s.get('periods', '')
                key = (cn, subj_name)
                if key in seen_consec:
                    continue
                seen_consec.add(key)
                group_no += 1
                all_rows.append({
                    'group': group_no, 'class': cn,
                    'subject': subj_name, 'teacher': subj_teach,
                    'par_subj': '—', 'par_teacher': '—',
                    'section': 'C', 'periods': periods,
                })

        # Calculate & allocate
        group_slots      = self._calculate_group_slots(all_rows)
        group_allocation = self._allocate_group_slots(all_rows, group_slots)

        return group_slots, group_allocation, all_rows

    def _proceed_to_stage2(self):
        """
        Gate check before entering Stage 2.
        All groups must have been successfully allocated by 'Allocate Periods'.
        If any failed, show a detailed error + suggestion dialog and block.
        If all ok, open the Stage 2 timetable page.
        """
        if not getattr(self, '_last_allocation', None):
            return {'ok': False, 'reason': 'allocation_not_run'}
        failed = {gn: ar for gn, ar in self._last_allocation.items() if not ar.get('ok', False)}
        if failed:
            return {'ok': False, 'reason': 'groups_failed', 'failed': failed,
                    'all_rows': getattr(self, '_last_all_rows', [])}
        return {'ok': True}

    # ── Allocation error dialog ───────────────────────────────────────────────
    def _allocation_suggestion(self, reason, rows, sec):
        """Return a human-readable suggestion string based on the failure reason."""
        reason_l = (reason or '').lower()
        teachers = list(dict.fromkeys(
            t for r in rows
            for t in [r.get('teacher',''), r.get('par_teacher','')]
            if t and t not in ('—','?','')))
        classes  = list(dict.fromkeys(r['class'] for r in rows))

        if 'stage 1 not run' in reason_l:
            return ("Run Stage 1 first (click '▶ Stage 1: Fill CT Slots') "
                    "before allocating periods.")

        if 'not found in engine' in reason_l or 'not in task list' in reason_l:
            return ("The subject name in Step 2 may not exactly match what Step 3 "
                    "recorded. Open Step 2 for class {} and verify the subject "
                    "name spelling matches exactly.".format(classes[0] if classes else '?'))

        if 'teacher' in reason_l and 'busy' in reason_l:
            busy_t = [t for t in teachers if t.lower() in reason_l]
            t_str  = ', '.join(busy_t) if busy_t else ', '.join(teachers)
            return ("Teacher {} is fully occupied at all candidate slots. "
                    "Options:  (a) Reduce total periods for one of their other subjects "
                    "in Step 2,  (b) Remove an unavailability block in Step 3 if one "
                    "was set by mistake,  (c) Reassign this subject to a different "
                    "teacher.".format(t_str))

        if 'occupied' in reason_l or 'class' in reason_l:
            c_str = ', '.join(classes)
            return ("Class {} has no free slots at the required periods. "
                    "Options:  (a) Reduce the period count for another subject "
                    "assigned to this class in Step 2,  (b) Split the combine into "
                    "smaller groups so fewer classes compete for the same "
                    "slots.".format(c_str))

        if sec == 'C':
            return ("No two adjacent free periods found for the consecutive subject. "
                    "Try reducing the period count for another subject in this class, "
                    "or disable the 'Consecutive' flag if back-to-back periods are "
                    "not strictly required.")

        if sec == 'B':
            return ("Both the primary and parallel teachers must be free at the "
                    "same period. Check each teacher's schedule and reduce workload "
                    "or unavailability constraints in Step 3.")

        if sec == 'A':
            return ("All {} classes AND all their teachers must be free at the same "
                    "slot. Reduce the number of classes in this combine group (Step 3) "
                    "or reduce period counts for conflicting subjects "
                    "(Step 2).".format(len(classes)))

        return ("Check that all teachers in this group have free periods available "
                "and that the affected classes have not already been fully "
                "scheduled by Stage 1 / other groups.")

    # ── Task Analysis 2 — allocation engine ─────────────────────────────────
    def _run_ta2_allocation(self):
        """
        Two-phase allocation for Task Analysis 2.

        PHASE 1 — Main Periods  (tasks where periods >= wdays-1, not relaxed):
            Scan periods p=0..ppd-1. At each p, collect all days where the
            class(es) AND teacher(s) are free. If free days >= remaining,
            assign ALL needed slots at that one period. Otherwise fall through.

        PHASE 2 — Filler Periods  (everything else + Phase-1 fall-through):
            Walk d=0..wdays-1, p=0..ppd-1 and place wherever free.

        Returns
        -------
        dict  task_idx -> {
            'phase'      : 'main' | 'filler',
            'placed'     : [(d, p), ...],   # new placements this run
            'remaining'  : int,             # still unplaced after this run
            'fail_reason': str,             # '' if fully placed
        }
        """
        if not hasattr(self, '_gen'):
            return {}

        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()

        results = {}

        # ── helpers ──────────────────────────────────────────────────────────────
        def slot_free(task, d, p):
            for cn in task['cn_list']:
                if grid.get(cn, [[]])[d][p] is not None:
                    return False
            t  = task['teacher']
            pt = task.get('par_teach', '')
            if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                return False
            if pt and pt not in ('', '—', '?'):
                if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                    return False
            return True

        def is_main(task):
            key = (frozenset(task['cn_list']), task['subject'])
            if key in self._relaxed_main_keys:
                return False
            return task['periods'] >= wdays - 1   # use original period count

        # ── PHASE 1 — Main Periods ────────────────────────────────────────────────
        for task in tasks:
            if task['remaining'] <= 0:
                continue
            if not is_main(task):
                continue

            placed      = []
            needed      = task['remaining']
            fail_reason = ''

            for p in range(ppd):
                if task['remaining'] <= 0:
                    break
                avail = [d for d in range(wdays) if slot_free(task, d, p)]
                if len(avail) >= task['remaining']:
                    for d in avail[:task['remaining']]:
                        self._gen_place(task, d, p)
                        placed.append((d, p))
                    break   # fully placed at this period

            if task['remaining'] > 0:
                # Diagnose which period came closest and what was blocking it
                best_days   = 0
                best_period = -1
                all_busy_teachers = set()
                for p in range(ppd):
                    free_days = [d for d in range(wdays) if slot_free(task, d, p)]
                    if len(free_days) > best_days:
                        best_days   = len(free_days)
                        best_period = p + 1
                    for d in range(wdays):
                        t  = task['teacher']
                        pt = task.get('par_teach', '')
                        if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                            all_busy_teachers.add(t)
                        if pt and pt not in ('', '—', '?'):
                            if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                                all_busy_teachers.add(pt)

                busy_t_str = ', '.join(sorted(all_busy_teachers)) or 'teacher(s)'

                if best_days == 0:
                    fail_reason = (
                        'No free slot found at any period across all {} days. '
                        'Teacher(s) {} appear fully booked or all class slots are '
                        'already occupied. Needed {} periods on the SAME period '
                        'each day.'.format(wdays, busy_t_str, needed))
                else:
                    fail_reason = (
                        'Could not find a single period free on all {} required '
                        'days. Best found: Period {} with only {}/{} free days. '
                        'Teacher(s) {} are occupied at conflicting slots. '
                        'Use "Relax to Filler" to allow flexible placement '
                        'across different periods.'.format(
                            needed, best_period, best_days, needed, busy_t_str))

            results[task['idx']] = {
                'phase':       'main',
                'placed':      placed,
                'remaining':   task['remaining'],
                'fail_reason': fail_reason,
            }

        # ── PHASE 2 — Filler Periods (+ relaxed main + main fall-through) ─────────
        for task in tasks:
            if task['remaining'] <= 0:
                if task['idx'] not in results:
                    results[task['idx']] = {
                        'phase':       'main' if is_main(task) else 'filler',
                        'placed':      [],
                        'remaining':   0,
                        'fail_reason': '',
                    }
                continue

            phase  = 'main' if is_main(task) else 'filler'
            placed = list(results.get(task['idx'], {}).get('placed', []))

            for d in range(wdays):
                if task['remaining'] <= 0:
                    break
                for p in range(ppd):
                    if task['remaining'] <= 0:
                        break
                    if slot_free(task, d, p):
                        self._gen_place(task, d, p)
                        placed.append((d, p))

            fail_reason = ''
            if task['remaining'] > 0:
                busy_teachers = set()
                for d in range(wdays):
                    for p in range(ppd):
                        class_ok = all(
                            grid.get(cn, [[]])[d][p] is None
                            for cn in task['cn_list'])
                        if not class_ok:
                            continue
                        t  = task['teacher']
                        pt = task.get('par_teach', '')
                        if not g['t_free'](t, d, p) or g['t_unavail'](t, d, p):
                            busy_teachers.add(t)
                        if pt and pt not in ('', '—', '?'):
                            if not g['t_free'](pt, d, p) or g['t_unavail'](pt, d, p):
                                busy_teachers.add(pt)
                busy_t_str = ', '.join(sorted(busy_teachers)) or 'teacher(s)'
                fail_reason = (
                    '{} slot(s) could not be placed. '
                    'Teacher(s) {} appear fully occupied at all remaining free '
                    'class slots. All {} free grid positions have been '
                    'exhausted.'.format(
                        task['remaining'], busy_t_str,
                        sum(1 for d in range(wdays) for p in range(ppd)
                            if all(grid.get(cn, [[]])[d][p] is None
                                for cn in task['cn_list']))))

            results[task['idx']] = {
                'phase':       phase,
                'placed':      placed,
                'remaining':   task['remaining'],
                'fail_reason': fail_reason,
            }

        return results

    # ── Task Analysis 2 page ─────────────────────────────────────────────────
    def check_ta2_done(self):
        """Return True if stage 2 allocation is complete."""
        return bool(getattr(self, '_last_ta2_allocation', None))

    def run_stage3(self):
        """Run Stage 3 synchronously. Returns status dict."""
        self._progress_log = []
        self._run_stage2_phases()
        return getattr(self, '_stage2_status', {})

    def run_force_fill(self, progress_cb=None):
        """
        Run Force Fill (Min-Conflicts CSP solver) synchronously.
        Stops as soon as all periods are placed (or 1500 iterations max).
        Returns a result dict with:
          - ok:            bool
          - remaining:     int (unplaced period count after force fill)
          - relaxed:       str or None (constraint relaxation notes)
          - overloaded:    list of (teacher, assigned, capacity, excess, unplaced)
          - blocked_only:  list of (teacher, assigned, capacity, unplaced)
          - progress_msgs: list of str
        """
        if self._gen is None:
            return {'ok': False, 'remaining': -1, 'relaxed': None,
                    'overloaded': [], 'blocked_only': [], 'progress_msgs': []}

        progress_msgs = []

        def _cb(msg):
            progress_msgs.append(msg)
            if progress_cb:
                progress_cb(msg)

        relaxed = self._force_fill_backtrack(progress_cb=_cb)

        # Refresh timetable snapshot
        self._timetable = self._gen_snapshot_tt()

        g       = self._gen
        tasks   = g['tasks']
        wdays   = g['wdays']
        ppd     = g['ppd']
        remaining = sum(t['remaining'] for t in tasks)
        total_slots = wdays * ppd

        overloaded   = []
        blocked_only = []

        if remaining > 0:
            teacher_assigned = {}
            teacher_unplaced = {}
            for t in tasks:
                for tname in ([t['teacher']] if t['teacher'] else []):
                    teacher_assigned[tname] = teacher_assigned.get(tname, 0) + t['periods']
                    if t['remaining'] > 0:
                        teacher_unplaced[tname] = teacher_unplaced.get(tname, 0) + t['remaining']
                pt = t.get('par_teach', '')
                if pt and pt not in ('', '—', '?'):
                    teacher_assigned[pt] = teacher_assigned.get(pt, 0) + t['periods']
                    if t['remaining'] > 0:
                        teacher_unplaced[pt] = teacher_unplaced.get(pt, 0) + t['remaining']

            for tname, assigned in sorted(teacher_assigned.items()):
                if assigned > total_slots:
                    excess   = assigned - total_slots
                    unp      = teacher_unplaced.get(tname, 0)
                    overloaded.append((tname, assigned, total_slots, excess, unp))

            for tname, unp in sorted(teacher_unplaced.items()):
                if unp > 0 and tname not in {o[0] for o in overloaded}:
                    assigned = teacher_assigned.get(tname, 0)
                    blocked_only.append((tname, assigned, total_slots, unp))

        # Update stage2 status
        self._stage2_status = {
            'unplaced': remaining,
            'ok':       remaining == 0,
            'msg':      ("✅ Force Fill complete — all periods placed!" if remaining == 0
                         else f"⚠ Force Fill done — {remaining} period(s) still unplaced."),
        }

        return {
            'ok':           remaining == 0,
            'remaining':    remaining,
            'relaxed':      relaxed,
            'overloaded':   overloaded,
            'blocked_only': blocked_only,
            'progress_msgs': progress_msgs,
            'wdays':        wdays,
            'ppd':          ppd,
            'total_slots':  total_slots,
        }

    # ── Task Analysis ────────────────────────────────────────────────────────
    def _calculate_group_slots(self, all_rows):
        """
        For every group in all_rows, determine how many slots that group needs.

        Strategy (per the spec):
          • Take the FIRST row of the group (any class in the group will do,
            as Step 3 guarantees all classes in a combined group share the
            same period count for their combined subject).
          • Look up class_config_data[cn]['subjects'] for that class.
          • Find the entry whose 'name' matches the row's subject.
          • Its 'periods' value is the number of slots required.
          • Parallel subjects share the same slot, so no extra count needed.

        Returns dict:
          { group_no: {'slots': int,  'ok': True} }          — success
          { group_no: {'slots': None, 'ok': False,
                       'reason': '<short reason>'} }          — failure
        """
        # Collect first row per group
        group_first = {}
        for row in all_rows:
            g = row['group']
            if g not in group_first:
                group_first[g] = row

        result = {}
        for g, row in group_first.items():
            cn   = row['class']
            subj = row['subject']

            # ── Guard: class config missing ───────────────────────────────
            if cn not in self.class_config_data:
                result[g] = {'slots': None, 'ok': False,
                             'reason': 'No config for {}'.format(cn)}
                continue

            cd_subjects = self.class_config_data[cn].get('subjects', [])
            if not cd_subjects:
                result[g] = {'slots': None, 'ok': False,
                             'reason': 'No subjects in {}'.format(cn)}
                continue

            # ── Search for the subject by name ────────────────────────────
            periods = None
            for s in cd_subjects:
                if s.get('name', '').strip() == subj:
                    periods = s.get('periods')
                    break

            # ── Fallback: subject may be the PARALLEL side of another entry
            if periods is None:
                for s in cd_subjects:
                    if (s.get('parallel')
                            and s.get('parallel_subject', '').strip() == subj):
                        # Parallel subjects share the same slot as the
                        # primary subject; use the primary's period count.
                        periods = s.get('periods')
                        break

            # ── Evaluate result ───────────────────────────────────────────
            if periods is None:
                result[g] = {'slots': None, 'ok': False,
                             'reason': '"{}" not found in {}'.format(subj, cn)}
            else:
                try:
                    result[g] = {'slots': int(periods), 'ok': True}
                except (ValueError, TypeError):
                    result[g] = {'slots': None, 'ok': False,
                                 'reason': 'Bad period value "{}"'.format(periods)}

        return result

    # ── Slot allocation engine ────────────────────────────────────────────────
    def _allocate_group_slots(self, all_rows, group_slots):
        """
        Allocate timetable slots to every group following RULE1 and RULE2.

        RULE1 – Slot priority order:
            Start from the last period of every day, work backwards.
            i.e. for period p from (ppd-1) down to 0, try all days at that p.

        RULE2 – Only fill what Stage 1 has not already covered:
            remaining = task['remaining']  (already decremented by Stage 1)

        Processing order:   C (consecutive)  →  B (standalone parallel)  →  A (combined)

        Returns
        -------
        dict:  { group_no → alloc_result }

        alloc_result (success):
            {'ok': True, 'total': int, 's1_placed': int,
             'new_placed': int, 'slots': [(d, p), ...]}

        alloc_result (failure / partial):
            {'ok': False, 'total': int, 's1_placed': int,
             'new_placed': int, 'slots': [(d, p), ...],   # partial placements
             'reason': str}
        """
        # ── Guard: Stage 1 must have been run ────────────────────────────────
        if not hasattr(self, '_gen'):
            dummy = {'ok': False, 'total': 0, 's1_placed': 0,
                     'new_placed': 0, 'slots': [],
                     'reason': 'Stage 1 not run yet — generate timetable first'}
            return {row['group']: dummy for row in all_rows}

        g      = self._gen
        grid   = g['grid']
        t_busy = g['t_busy']
        ppd    = g['ppd']
        wdays  = g['wdays']
        DAYS   = g['DAYS']

        # ── Helpers ──────────────────────────────────────────────────────────
        def slot_is_free_for_classes(cn_list, d, p):
            return all(grid.get(cn, [[]])[d][p] is None
                       for cn in cn_list if cn in grid)

        def teacher_free(t, d, p):
            if not t or t in ('—', '?', ''):
                return True
            return ((d, p) not in t_busy.get(t, set())
                    and not g['t_unavail'](t, d, p))

        def all_teachers_free(teachers, d, p):
            return all(teacher_free(t, d, p) for t in teachers)

        def mark_teachers_busy(teachers, d, p):
            for t in teachers:
                if t and t not in ('—', '?', ''):
                    t_busy.setdefault(t, set()).add((d, p))

        def place_slot(task, extra_par_teachers, d, p, class_info_map=None):
            """Place slot then apply per-class cell corrections.

            class_info_map: {cn -> {'type', 'par_subj', 'par_teach'}}
            Built entirely from Task Analysis rows so every class in the group
            gets the correct cell type and parallel-teacher regardless of what
            the engine task stored (which may be incomplete when a combined group
            has mixed parallel/non-parallel classes, e.g. Group 5: 12A no-par,
            12B has CS/Rajender).

            After _gen_place writes a shared cell to all cn in task['cn_list'],
            we overwrite each class's cell individually with correct data.
            """
            self._gen_place(task, d, p)

            # ── Fix each class's cell: type, par_subj, par_teach, teacher ───
            # Always apply — covers every class in combined/parallel groups.
            # primary_teacher fixes 7B/7C cells that still hold Anita (first class)
            # from the shared engine cell created by _gen_place.
            if class_info_map:
                for cn, info in class_info_map.items():
                    if cn in grid and grid[cn][d][p] is not None:
                        patch = {
                            'type':      info['type'],
                            'par_subj':  info['par_subj'],
                            'par_teach': info['par_teach'],
                        }
                        pt = info.get('primary_teacher', '').strip()
                        if pt and pt not in ('—', '?'):
                            patch['teacher'] = pt
                        grid[cn][d][p] = dict(grid[cn][d][p], **patch)

            # ── Mark ALL parallel teachers busy ──────────────────────────────
            # _gen_place marks task['teacher'] + task['par_teach'].
            # Also mark every par_teacher from the rows that is not yet marked.
            all_extra = set(extra_par_teachers)
            if class_info_map:
                for info in class_info_map.values():
                    pt = info.get('par_teach', '') or ''
                    if pt and pt not in ('—', '?', ''):
                        all_extra.add(pt)
            engine_par = (task.get('par_teach') or '').strip()
            for t in all_extra:
                if t and t not in ('—', '?', '') and t != engine_par:
                    t_busy.setdefault(t, set()).add((d, p))

        # ── Build task lookup — index by BOTH primary and parallel identity ─────
        #
        # A task in the engine is always created for the PRIMARY subject (e.g. SKT)
        # with par_subj/par_teach pointing to the parallel subject (e.g. Urdu/Irfan).
        # But a Task Analysis row (Section A/B) may identify a group by the PARALLEL
        # subject (Urdu/Irfan) because that is what is stored in step3_data combines.
        # We must find the task no matter which side the row uses as its identity.
        #
        # task_by_primary  — keyed by (frozenset(cn_list), subject,  teacher)
        # task_by_parallel — keyed by (frozenset(cn_list), par_subj, par_teach)

        task_by_primary  = {}
        task_by_parallel = {}
        for _t in g['tasks']:
            pk = (frozenset(_t['cn_list']), _t['subject'], _t['teacher'])
            task_by_primary[pk] = _t
            ps = (_t.get('par_subj') or '').strip()
            pt = (_t.get('par_teach') or '').strip()
            if ps and pt and ps not in ('—', '?'):
                sk = (frozenset(_t['cn_list']), ps, pt)
                task_by_parallel[sk] = _t

        task_lookup = task_by_primary   # alias used in fallback below

        # ── Organise rows by group ────────────────────────────────────────────
        group_rows    = {}
        group_section = {}
        for row in all_rows:
            gn = row['group']
            if gn not in group_rows:
                group_rows[gn]    = []
                group_section[gn] = row['section']
            group_rows[gn].append(row)

        # ── Process in order: C → B → A ──────────────────────────────────────
        result = {}

        for sec in ('C', 'B', 'A'):
            for gn, rows in sorted(group_rows.items()):
                if group_section[gn] != sec:
                    continue

                gs = group_slots.get(gn)
                if gs is None or not gs['ok']:
                    result[gn] = {
                        'ok': False, 'total': 0, 's1_placed': 0,
                        'new_placed': 0, 'slots': [],
                        'reason': (gs['reason'] if gs else 'Slot count unknown'),
                    }
                    continue

                total_periods  = gs['slots']
                first_row      = rows[0]
                primary_subj   = first_row['subject']
                primary_teach  = first_row['teacher']

                # All classes in this group (in display order, deduplicated)
                all_cn = list(dict.fromkeys(r['class'] for r in rows))

                # ── Find the matching task — 4-pass lookup ────────────────────
                #
                # Pass 1: exact match on (cn_list, subject, teacher)
                # Pass 2: exact match on (cn_list, par_subj, par_teach)  ← key fix
                # Pass 3: loose match on subject/teacher ignoring cn_list size
                # Pass 4: loose match on par_subj/par_teach ignoring cn_list size
                #
                # The row's primary_subj may be the PARALLEL side in the engine
                # (e.g. row says Urdu/Irfan but engine task is SKT/Anita with
                # par_subj=Urdu, par_teach=Irfan).  All passes are checked so we
                # always find the real task regardless of which side is "primary".

                cn_fs = frozenset(all_cn)
                task  = (task_by_primary.get((cn_fs, primary_subj, primary_teach))
                         or task_by_parallel.get((cn_fs, primary_subj, primary_teach)))

                if task is None:
                    # Pass 3 & 4: relax the cn_list requirement (subset match)
                    for t_obj in g['tasks']:
                        cn_overlap = bool(frozenset(t_obj['cn_list']) & cn_fs)
                        if not cn_overlap:
                            continue
                        via_primary  = (t_obj['subject']           == primary_subj
                                        and t_obj['teacher']       == primary_teach)
                        via_parallel = (t_obj.get('par_subj', '')  == primary_subj
                                        and t_obj.get('par_teach', '') == primary_teach)
                        if via_primary or via_parallel:
                            task = t_obj
                            break

                if task is None:
                    # Debug info: show a sample of task subjects to help diagnose
                    sample = [(t['subject'], t['teacher'], t.get('par_subj',''),
                               t.get('par_teach',''), list(t['cn_list']))
                              for t in g['tasks'][:8]]
                    result[gn] = {
                        'ok': False, 'total': total_periods,
                        's1_placed': 0, 'new_placed': 0, 'slots': [],
                        'reason': (
                            'Task not found — "{}"/{} not in engine '
                            '(checked primary + parallel sides). '
                            'Verify subject name matches Step 2 exactly.'.format(
                                primary_subj, primary_teach)),
                    }
                    continue

                # ── Once found, identify which side the row matched ───────────────
                # "found_via_parallel" = the row's primary is the engine's par_subj.
                # This is important for building the all_teachers_needed list correctly:
                # both the task's teacher AND par_teach must be checked as busy.
                found_via_parallel = (
                    task.get('par_subj', '').strip() == primary_subj
                    and task.get('par_teach', '').strip() == primary_teach
                )

                # Stage-1-placed count
                s1_placed = task['periods'] - task['remaining']
                remaining = task['remaining']

                if remaining <= 0:
                    result[gn] = {
                        'ok': True, 'total': total_periods,
                        's1_placed': s1_placed, 'new_placed': 0,
                        'slots': [],
                    }
                    continue

                # ── Collect ALL teachers that must be free at the chosen slot ──
                #
                # We collect teachers from two sources and union them:
                #   Source 1 — the engine task itself: task['teacher'] + task['par_teach']
                #   Source 2 — the Task Analysis rows: each row's teacher + par_teacher
                #              (combined groups have a different par_teacher per class)
                #
                # This ensures that even when the task was found via its parallel side
                # (found_via_parallel=True), we still check both Anita (SKT) and Irfan
                # (Urdu) for availability before claiming the slot.

                task_teachers = []
                if task['teacher']:
                    task_teachers.append(task['teacher'])
                if task.get('par_teach', '') and task['par_teach'] not in ('—','?',''):
                    task_teachers.append(task['par_teach'])

                row_teachers = []
                for row in rows:
                    for fld in ('teacher', 'par_teacher'):
                        t = row.get(fld, '')
                        if t and t not in ('—', '?', ''):
                            row_teachers.append(t)

                all_teachers_needed = list(
                    dict.fromkeys(task_teachers + row_teachers))

                # Extra par teachers: those in all_teachers_needed that the task's
                # _gen_place does NOT already mark busy (it only marks teacher + par_teach)
                engine_marks = set(filter(None, [task['teacher'],
                                                  task.get('par_teach', '')]))
                extra_par = [t for t in all_teachers_needed
                             if t not in engine_marks]

                # ── Per-class info map: type + par_subj + par_teach ──────────
                #
                # Builds {cn → {type, par_subj, par_teach}} entirely from the
                # Task Analysis rows.  This is the AUTHORITATIVE source — the
                # engine task may be incomplete (e.g. Group 5: engine task was
                # built from 12A which has no parallel, so par_teach='' even
                # though 12B has CS/Rajender).  We apply this map to EVERY class
                # unconditionally so each class gets exactly the right cell.
                #
                # Rules for cell type:
                #   Section A + has parallel  → 'combined_parallel'
                #   Section A + no parallel   → 'combined'
                #   Section B                 → 'parallel'
                #   Section C                 → 'normal'
                # ── Per-class info map: type + par_subj + par_teach + primary_teacher ─
                #
                # Section A (combined) rows always show the COMBINE perspective:
                #   row['subject']     = combine subject  (Urdu/Irfan)
                #   row['teacher']     = combine teacher  (Irfan) — same for every row
                #   row['par_subj']    = class primary subject (SKT)
                #   row['par_teacher'] = class primary teacher (Anita/Neha/Mamta)
                #
                # When found_via_parallel=True (Urdu row → engine SKT task):
                #   cell.teacher    = engine's task['teacher'] = Anita (only 7A correct!)
                #   cell.par_teach  = Irfan ✓
                #   cell.subject    = SKT ✓
                # → primary_teacher must come from row['par_teacher'] (Neha for 7B etc.)
                #
                # When found_via_parallel=False (engine task subject matches row):
                #   row['teacher'] maps directly to cell.teacher
                class_info_map = {}
                for row in rows:
                    ps = (row.get('par_subj')    or '').strip()
                    pt = (row.get('par_teacher') or '').strip()
                    has_par = bool(ps and pt and ps not in ('—', '?') and pt not in ('—', '?'))
                    if sec == 'A':
                        cell_type = 'combined_parallel' if has_par else 'combined'
                    elif sec == 'B':
                        cell_type = 'parallel'
                    else:
                        cell_type = 'normal'
                    # Determine the per-class PRIMARY teacher for the cell
                    if found_via_parallel:
                        # row['par_teacher'] = SKT teacher of this class (Anita/Neha/Mamta)
                        primary_teacher = (row.get('par_teacher') or '').strip()
                    else:
                        # row['teacher'] is already the cell's primary teacher
                        primary_teacher = (row.get('teacher') or '').strip()
                    class_info_map[row['class']] = {
                        'type':           cell_type,
                        'par_subj':       ps if has_par else '',
                        'par_teach':      pt if has_par else '',
                        'primary_teacher': primary_teacher,
                    }

                # ── Placement logic ───────────────────────────────────────────
                placed_slots  = []   # (d, p) pairs successfully placed
                last_fail_why = ''

                # Check if this group's consecutive constraint has been relaxed
                _relaxed = getattr(self, '_relaxed_consec_keys', set())
                _group_relaxed = (sec == 'C' and rows and
                                  (rows[0]['class'], rows[0]['subject']) in _relaxed)

                if sec == 'C' and not _group_relaxed:
                    # ── Consecutive: find adjacent pairs (p_start, p_start+1) ──
                    # RULE1 for pairs: start from (ppd-2, ppd-1) and go backwards.
                    for p_start in range(ppd - 2, -1, -1):
                        if len(placed_slots) >= remaining:
                            break
                        p1, p2 = p_start, p_start + 1
                        for d in range(wdays):
                            if len(placed_slots) >= remaining:
                                break
                            # Need both slots free for all classes + all teachers
                            cls_ok = (slot_is_free_for_classes(all_cn, d, p1)
                                      and slot_is_free_for_classes(all_cn, d, p2))
                            tch_ok = (all_teachers_free(all_teachers_needed, d, p1)
                                      and all_teachers_free(all_teachers_needed, d, p2))
                            if cls_ok and tch_ok:
                                # Place both (or only p1 if remaining == 1)
                                if remaining - len(placed_slots) >= 2:
                                    place_slot(task, extra_par, d, p1, class_info_map)
                                    place_slot(task, extra_par, d, p2, class_info_map)
                                    placed_slots.extend([(d, p1), (d, p2)])
                                else:
                                    place_slot(task, extra_par, d, p1, class_info_map)
                                    placed_slots.append((d, p1))
                            else:
                                if not cls_ok:
                                    busy_cn = [cn for cn in all_cn
                                               if cn in grid and (
                                                   grid[cn][d][p1] is not None
                                                   or grid[cn][d][p2] is not None)]
                                    last_fail_why = (
                                        '{} P{}-P{}: class {} occupied'.format(
                                            DAYS[d], p1+1, p2+1,
                                            ', '.join(busy_cn)))
                                else:
                                    busy_t = [t for t in all_teachers_needed
                                              if not teacher_free(t, d, p1)
                                              or not teacher_free(t, d, p2)]
                                    last_fail_why = (
                                        '{} P{}-P{}: teacher {} busy'.format(
                                            DAYS[d], p1+1, p2+1,
                                            ', '.join(busy_t)))

                else:
                    # ── Section B / A (or relaxed-C): standard single-slot placement ──
                    # RULE1: last period first, iterate all days, then earlier periods
                    for p in range(ppd - 1, -1, -1):
                        if len(placed_slots) >= remaining:
                            break
                        for d in range(wdays):
                            if len(placed_slots) >= remaining:
                                break
                            cls_ok = slot_is_free_for_classes(all_cn, d, p)
                            tch_ok = all_teachers_free(all_teachers_needed, d, p)
                            if cls_ok and tch_ok:
                                place_slot(task, extra_par, d, p, class_info_map)
                                placed_slots.append((d, p))
                            else:
                                if not cls_ok:
                                    busy_cn = [cn for cn in all_cn
                                               if cn in grid
                                               and grid[cn][d][p] is not None]
                                    occupant = grid[busy_cn[0]][d][p] if busy_cn else {}
                                    last_fail_why = (
                                        '{} P{}: {} occupied by "{}"'.format(
                                            DAYS[d], p+1,
                                            ', '.join(busy_cn),
                                            occupant.get('subject', '?')))
                                else:
                                    busy_t = [t for t in all_teachers_needed
                                              if not teacher_free(t, d, p)]
                                    last_fail_why = (
                                        '{} P{}: teacher {} busy'.format(
                                            DAYS[d], p+1,
                                            ', '.join(busy_t)))

                # ── Build result ──────────────────────────────────────────────
                new_placed = len(placed_slots)
                if new_placed >= remaining:
                    result[gn] = {
                        'ok': True,
                        'total': total_periods,
                        's1_placed': s1_placed,
                        'new_placed': new_placed,
                        'slots': placed_slots,
                    }
                else:
                    still_short = remaining - new_placed
                    if new_placed == 0:
                        reason = ('No free slots found. '
                                  + (last_fail_why or 'All slots occupied'))
                    else:
                        reason = ('{} slot(s) still unplaced. '
                                  'Last conflict: {}'.format(
                                      still_short,
                                      last_fail_why or 'Unknown'))
                    result[gn] = {
                        'ok': False,
                        'total': total_periods,
                        's1_placed': s1_placed,
                        'new_placed': new_placed,
                        'slots': placed_slots,
                        'reason': reason,
                    }

        return result

    # ─────────────────────────────────────────────────────────────────────────
    def _run_stage2(self):
        """Legacy Stage 2 entry — now redirects to Stage 3 (filler phases)."""
        self._run_stage3()

    # =========================================================================
    #  CORE GENERATION ENGINE  (split into init + stage1 + stage2)
    # =========================================================================

    def _init_gen_state(self):
        """
        Build the grid, task list and all helper closures.
        Called once before Stage 1. Results stored on self._gen.
        """
        cfg   = self.configuration
        ppd   = cfg['periods_per_day']
        wdays = cfg['working_days']
        half1 = cfg['periods_first_half']
        DAYS  = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][:wdays]

        all_classes = []
        for cls in range(6, 13):
            for si in range(cfg['classes'][cls]):
                all_classes.append("{}{}".format(cls, chr(65 + si)))

        grid    = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        task_at = {cn: [[None]*ppd for _ in range(wdays)] for cn in all_classes}
        t_busy  = {}

        def t_free(t, d, p):
            return not t or (d, p) not in t_busy.get(t, set())
        def t_mark(t, d, p):
            if t: t_busy.setdefault(t, set()).add((d, p))
        def t_unmark(t, d, p):
            if t: t_busy.get(t, set()).discard((d, p))

        unavail = getattr(self, 'step3_unavailability', {})
        def t_unavail(t, d, p):
            u = unavail.get(t, {})
            if not u: return False
            return DAYS[d] in u.get('days', []) and (p+1) in u.get('periods', [])

        # ── combine lookup ───────────────────────────────────────────────────────
        # Pass 1: map (cn, combine_subject) → combined_class_list
        #   e.g. Step-3 combine for Irfan/Urdu across 7A+7B+7C
        #   gives: ('7A','Urdu')→['7A','7B','7C'], etc.
        s3 = getattr(self, 'step3_data', {})
        cn_subj_combined = {}
        for _teacher, s3d in s3.items():
            for cb in s3d.get('combines', []):
                classes  = sorted(cb.get('classes', []))
                subjects = cb.get('subjects', [])
                if len(classes) >= 2 and subjects:
                    for cn in classes:
                        cn_subj_combined[(cn, subjects[0])] = classes

        # Pass 2: if a primary subject's par_subj is a combine subject, also
        # map that primary subject to the same combined class list.
        #
        # Example: Step-3 combine subject = 'Urdu' → ('7A','Urdu')=['7A','7B','7C']
        # Class config 7A: primary='SKT', par_subj='Urdu'.
        # After pass 2: ('7A','SKT')=['7A','7B','7C'] is also added.
        # This ensures that when the engine iterates 7A's 'SKT' subject it
        # correctly receives cn_list=['7A','7B','7C'] instead of just ['7A'].
        for cn in all_classes:
            cd_subjects = self.class_config_data.get(cn, {}).get('subjects', [])
            for s in cd_subjects:
                primary_subj = s.get('name', '').strip()
                par_subj     = (s.get('parallel_subject') or '').strip()
                # If the parallel-subject of this entry is already a combine key,
                # map this primary subject to the same combined class list.
                if par_subj and (cn, par_subj) in cn_subj_combined:
                    if (cn, primary_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, primary_subj)] = cn_subj_combined[(cn, par_subj)]
                # Also handle the reverse: if this primary IS the combine subject
                # and it has a par_subj, ensure par_subj is also combined.
                if (cn, primary_subj) in cn_subj_combined and par_subj:
                    if (cn, par_subj) not in cn_subj_combined:
                        cn_subj_combined[(cn, par_subj)] = cn_subj_combined[(cn, primary_subj)]

        # build tasks
        tasks = []
        # Key: (frozenset(cn_list), subject_name)
        # Deduplicates combined groups regardless of which class is processed first
        # or which teacher is recorded — the combine is one task for the whole group.
        seen_combined = set()
        for cn in all_classes:
            if cn not in self.class_config_data:
                continue
            cd     = self.class_config_data[cn]
            ct     = cd.get('teacher', '').strip()
            ct_per = cd.get('teacher_period', 1)

            # Only ONE subject per class is the "CT subject" — the first subject
            # whose teacher matches the class teacher.  The CT period is WHERE
            # that subject is placed (period ct_per, every day it appears).
            # Any other subjects taught by the same teacher in the same class are
            # normal subjects and are scheduled freely by Stage 2.
            ct_subject_assigned = False

            for s in cd['subjects']:
                subj = s['name']
                t    = s['teacher'].strip()
                n    = s['periods']

                cn_list = cn_subj_combined.get((cn, subj), [cn])
                if len(cn_list) > 1:
                    # Use frozenset + subject as dedup key (teacher intentionally
                    # excluded — different classes have different par_teachers but
                    # represent the same combined group)
                    key = (frozenset(cn_list), subj)
                    if key in seen_combined:
                        continue
                    seen_combined.add(key)

                # is_ct: True only for the FIRST subject whose teacher == CT.
                # Additional subjects by the same CT teacher are regular tasks.
                if t == ct and not ct_subject_assigned:
                    is_ct = True
                    ct_subject_assigned = True
                else:
                    is_ct = False

                par    = bool(s.get('parallel', False))
                pt     = s.get('parallel_teacher', '').strip() if par else ''
                ps     = s.get('parallel_subject', '').strip() if par else ''
                consec = (s.get('consecutive', 'No') == 'Yes')
                # If the user has relaxed this group's consecutive constraint,
                # override so the engine also treats it as non-consecutive
                if consec and (cn, subj) in getattr(self, '_relaxed_consec_keys', set()):
                    consec = False
                p_pref = list(s.get('periods_pref', []))
                d_pref = list(s.get('days_pref', []))

                if len(cn_list) > 1 and par:
                    ttype = 'combined_parallel'
                elif len(cn_list) > 1:
                    ttype = 'combined'
                elif par:
                    ttype = 'parallel'
                else:
                    ttype = 'normal'

                if is_ct:
                    priority = 'HC1'
                elif p_pref or d_pref:
                    priority = 'HC2'
                elif consec:
                    priority = 'SC1'
                elif n >= wdays:
                    priority = 'SC2'
                else:
                    priority = 'filler'

                tasks.append({
                    'idx':       len(tasks),
                    'cn_list':   cn_list,
                    'subject':   subj,
                    'teacher':   t,
                    'par_subj':  ps,
                    'par_teach': pt,
                    'periods':   n,
                    'remaining': n,
                    'is_ct':     is_ct,
                    'ct_period': ct_per if is_ct else None,
                    'p_pref':    p_pref,
                    'd_pref':    d_pref,
                    'consec':    consec,
                    'daily':     (n >= wdays),
                    'priority':  priority,
                    'type':      ttype,
                    'rx_sc1':    False,
                    'rx_sc3':    False,
                    'rx_sc2':    False,
                })

        total_atoms = sum(t['periods'] for t in tasks)

        # Store everything on self so stages can share state
        self._gen = {
            'cfg': cfg, 'ppd': ppd, 'wdays': wdays, 'half1': half1,
            'DAYS': DAYS, 'all_classes': all_classes,
            'grid': grid, 'task_at': task_at, 't_busy': t_busy,
            'tasks': tasks, 'total_atoms': total_atoms,
            't_free': t_free, 't_mark': t_mark, 't_unmark': t_unmark,
            't_unavail': t_unavail,
        }

    # ── Shared helpers (use self._gen) ────────────────────────────────────────

    def _gen_can_place(self, task, d, p,
                       ignore_sc1=False, ignore_sc3=False, ignore_sc2=False):
        g      = self._gen
        DAYS   = g['DAYS']; ppd = g['ppd']
        grid   = g['grid']
        t_free = g['t_free']; t_unavail = g['t_unavail']
        t      = task['teacher']; pt = task['par_teach']
        p1     = p + 1  # 1-based

        if task['is_ct'] and p1 != task['ct_period']:
            return False
        if task['p_pref'] and not task['is_ct']:
            if p1 not in task['p_pref']:
                return False
        if task['d_pref']:
            if DAYS[d] not in task['d_pref']:
                return False
        for cn in task['cn_list']:
            if grid[cn][d][p] is not None:
                return False
        if not t_free(t, d, p): return False
        if pt and not t_free(pt, d, p): return False
        if not (ignore_sc3 or task['rx_sc3']):
            if t_unavail(t, d, p): return False
            if pt and t_unavail(pt, d, p): return False
        # FIX BUG 5: consecutive tasks may go at ANY adjacent pair (p, p+1),
        # not locked to only the last two slots.  The partner slot p+1 must
        # also be empty and teacher-free.
        if task['consec'] and not (ignore_sc1 or task['rx_sc1']):
            if p >= ppd - 1:
                return False   # no room for the partner slot
            for cn in task['cn_list']:
                if grid[cn][d][p + 1] is not None:
                    return False
            if not t_free(t, d, p + 1): return False
            if pt and not t_free(pt, d, p + 1): return False
            if not (ignore_sc3 or task['rx_sc3']):
                if t_unavail(t, d, p + 1): return False
                if pt and t_unavail(pt, d, p + 1): return False
        if not task['consec']:
            for cn in task['cn_list']:
                for pp in range(ppd):
                    e = grid[cn][d][pp]
                    if e and e.get('subject') == task['subject']:
                        return False
        return True

    def _gen_count_valid_slots(self, task,
                               ignore_sc1=False, ignore_sc3=False,
                               ignore_sc2=False):
        """Count valid (d,p) placements for *task* right now (MRV helper)."""
        g = self._gen
        return sum(
            1
            for d in range(g['wdays'])
            for p in range(g['ppd'])
            if self._gen_can_place(task, d, p, ignore_sc1, ignore_sc3, ignore_sc2)
        )

    def _gen_make_cell(self, task):
        return {
            'type':      task['type'],
            'subject':   task['subject'],
            'teacher':   task['teacher'],
            'par_subj':  task['par_subj'],
            'par_teach': task['par_teach'],
            'combined_classes': task['cn_list'] if len(task['cn_list'])>1 else [],
            'is_ct':     task['is_ct'],
        }

    def _gen_place(self, task, d, p):
        g = self._gen
        cell = self._gen_make_cell(task)
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = cell
            g['task_at'][cn][d][p] = task['idx']
        g['t_mark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_mark'](task['par_teach'], d, p)
        task['remaining'] -= 1

    def _gen_unplace(self, task, d, p):
        g = self._gen
        for cn in task['cn_list']:
            g['grid'][cn][d][p]    = None
            g['task_at'][cn][d][p] = None
        g['t_unmark'](task['teacher'], d, p)
        if task['par_teach']:
            g['t_unmark'](task['par_teach'], d, p)
        task['remaining'] += 1

    def _gen_prog(self, msg, extra_pct=0):
        g = self._gen
        done = g['total_atoms'] - sum(t['remaining'] for t in g['tasks'])
        pct  = min(97, int(100 * done / max(g['total_atoms'], 1))) + extra_pct
        self._progress_log.append((msg, min(97, pct)))

    def _gen_snapshot_tt(self):
        """Return a tt-dict from current gen state (for display)."""
        g = self._gen
        unplaced = sum(t['remaining'] for t in g['tasks'])
        return {
            'grid':        g['grid'],
            'days':        g['DAYS'],
            'ppd':         g['ppd'],
            'half1':       g['half1'],
            'all_classes': g['all_classes'],
            'tasks':       g['tasks'],
            'unplaced':    unplaced,
        }

    # ── STAGE 1: HC1 (CT fixed periods) + HC2 (preference-constrained) ────────

    def _run_stage1_phases(self):
        """
        Stage 1 — place ALL CT (HC1) and ALL fixed/preference (HC2) periods.

        Logic is intentionally dead-simple:
          • For HC1: the CT period is fixed (same period-index, every working day).
            Just write into the grid cell if it is None.  No teacher conflict check
            is needed — Step 2 already guarantees these slots are conflict-free.
          • For HC2: iterate through the preferred (day, period) combinations in
            the order they were specified and fill the required number of slots.
            Again only check that the cell is still empty.

        Any period that STILL could not be placed (cell was already occupied) is
        reported as an issue with an exact explanation of what was blocking it.
        """
        g     = self._gen
        tasks = g['tasks']
        grid  = g['grid']
        wdays = g['wdays']
        ppd   = g['ppd']
        DAYS  = g['DAYS']

        s1_issues = []   # list of human-readable problem strings (HC1/HC2 only)

        # ══════════════════════════════════════════════════════════════════════
        # PHASE 1 — HC1: Class-teacher subject periods
        #
        # The CT teacher's subject S IS the CT period — they are the same thing.
        # Subject S has n periods configured in Step 2.  Those n periods are placed
        # at ct_per (same period-index) across n different working days.
        # No extra "CT admin" slots are added — the subject count n is the total.
        #
        # Only possible failure: cell already occupied (cannot happen if Step 2
        # validation was completed — reported as an issue if it does occur).
        # ══════════════════════════════════════════════════════════════════════
        self._gen_prog("Stage 1 · Phase 1 — Placing Class Teacher subject periods…")
        for task in tasks:
            if task['priority'] != 'HC1':
                continue
            p_idx = task['ct_period'] - 1          # 0-based period index
            for d in range(wdays):
                if task['remaining'] <= 0:
                    break                           # all n periods placed — done
                blocked_by = None
                for cn in task['cn_list']:
                    existing = grid[cn][d][p_idx]
                    if existing is not None:
                        blocked_by = (cn, existing)
                        break

                if blocked_by is None:
                    # Cell empty → place immediately, no conflict check needed
                    self._gen_place(task, d, p_idx)
                else:
                    cn_blk, cell_blk = blocked_by
                    s1_issues.append(
                        "HC1 — CT subject '{}' (teacher: {}, class: {}) "
                        "could NOT be placed on {} at Period {} — "
                        "cell already occupied by subject '{}' (teacher: {}).  "
                        "Step 2 validation should have prevented this.".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            DAYS[d], p_idx + 1,
                            cell_blk.get('subject', '?'),
                            cell_blk.get('teacher', '?')))

        # ══════════════════════════════════════════════════════════════════════
        # PHASE 2 — HC2: Fixed / preference-constrained subjects
        # Rule: iterate through every (day, period) combination that matches the
        # subject's day and period preferences and fill until remaining == 0.
        # Only block = cell already occupied.
        # ══════════════════════════════════════════════════════════════════════
        self._gen_prog("Stage 1 · Phase 2 — Placing preference-constrained subjects…")

        # Sort most-constrained first (fewest allowed slots → place first)
        hc2_tasks = sorted(
            [t for t in tasks if t['priority'] == 'HC2'],
            key=lambda t: (len(t['p_pref']) or ppd) * (len(t['d_pref']) or wdays))

        for task in hc2_tasks:
            if task['remaining'] <= 0:
                continue

            # Build the ordered list of preferred (day, period) slots
            pref_p = [x - 1 for x in task['p_pref']] if task['p_pref'] else list(range(ppd))
            pref_d = (
                [DAYS.index(x) for x in task['d_pref'] if x in DAYS]
                if task['d_pref'] else list(range(wdays)))
            slots = [(d, p) for d in pref_d for p in pref_p]

            blocked_slots = []   # (day_name, period_1based, blocker_subject, blocker_teacher)

            for d, p in slots:
                if task['remaining'] <= 0:
                    break
                blocked_by = None
                for cn in task['cn_list']:
                    existing = grid[cn][d][p]
                    if existing is not None:
                        blocked_by = (cn, existing)
                        break

                if blocked_by is None:
                    self._gen_place(task, d, p)
                else:
                    cn_blk, cell_blk = blocked_by
                    blocked_slots.append((DAYS[d], p + 1,
                                          cell_blk.get('subject', '?'),
                                          cell_blk.get('teacher', '?')))

            if task['remaining'] > 0:
                # Still unplaced periods — report each blocked slot
                classes_str = ', '.join(task['cn_list'])
                for day_name, per_1b, blk_subj, blk_teach in blocked_slots:
                    s1_issues.append(
                        "HC2 — Subject '{}' (teacher: {}) for class {} could NOT be "
                        "placed on {} at Period {} — slot occupied by '{}' "
                        "(teacher: {}).".format(
                            task['subject'], task['teacher'], classes_str,
                            day_name, per_1b, blk_subj, blk_teach))
                if not blocked_slots:
                    # No preferred slot was even available in the preference list
                    s1_issues.append(
                        "HC2 — Subject '{}' (teacher: {}) for class {} has {} "
                        "period(s) unplaced — no matching preferred slot "
                        "exists in the grid (preferences: days={}, periods={}).".format(
                            task['subject'], task['teacher'],
                            ', '.join(task['cn_list']),
                            task['remaining'],
                            task['d_pref'] or 'Any',
                            task['p_pref'] or 'Any'))

        # ══════════════════════════════════════════════════════════════════════
        # REPORTING
        # ══════════════════════════════════════════════════════════════════════
        hc1_placed = sum(t['periods'] - t['remaining'] for t in tasks if t['priority'] == 'HC1')
        hc2_placed = sum(t['periods'] - t['remaining'] for t in tasks if t['priority'] == 'HC2')
        hc1_fail   = sum(t['remaining']                for t in tasks if t['priority'] == 'HC1')
        hc2_fail   = sum(t['remaining']                for t in tasks if t['priority'] == 'HC2')
        other_rem  = sum(t['remaining'] for t in tasks if t['priority'] not in ('HC1', 'HC2'))

        has_issues = bool(s1_issues)
        stage_bg   = "#c0392b" if has_issues else "#1a7a1a"

        if has_issues:
            stage_txt = ("  ⚠ Stage 1 — {} CT period(s) placed, "
                         "{} preference period(s) placed  |  "
                         "{} issue(s) — see status bar".format(
                            hc1_placed, hc2_placed, len(s1_issues)))
        else:
            stage_txt = ("  ✓ Stage 1 complete — {} CT period(s) placed, "
                         "{} preference period(s) placed — "
                         "no issues".format(hc1_placed, hc2_placed))

        if has_issues:
            issue_lines = "\n".join("  ⚠ {}".format(i) for i in s1_issues)
            status = (
                "⚠ Stage 1 complete with {} issue(s) — these should not occur "
                "if Step 2 was fully validated.\n\n"
                "ISSUES:\n{}\n\n"
                "Click 'Task Analysis →' to review groups, then proceed to Stage 2 "
                "({} more period(s) to place).".format(
                    len(s1_issues), issue_lines, other_rem + hc1_fail + hc2_fail))
        else:
            status = (
                "✅ Stage 1 complete — all CT and fixed/preference periods placed "
                "with zero issues.\n"
                "Click '📋 Task Analysis →' to review parallel groups before Stage 2 "
                "({} period(s) remaining).".format(other_rem))

        self._gen_stage = 1
        tt = self._gen_snapshot_tt()
        self._timetable = tt
        self._stage1_status = {
            'stage_txt': stage_txt,
            'stage_bg':  stage_bg,
            'status':    status,
            'has_issues': has_issues,
        }

    def _run_stage2_phases(self):
        g     = self._gen
        tasks = g["tasks"]
        wdays = g["wdays"]
        ppd   = g["ppd"]
        DAYS  = g["DAYS"]

        # ── Phase 3 — SC1: Consecutive pairs ──────────────────────────────
        # BUG1 FIX: try EVERY adjacent pair (p, p+1), not only last two slots.
        self._gen_prog("Stage 3 · Phase 3 — Consecutive double-periods…")
        sc1_tasks = sorted([t for t in tasks if t["priority"] == "SC1"],
                           key=lambda t: -t["periods"])
        for task in sc1_tasks:
            if task["remaining"] <= 0:
                continue
            day_order = list(range(wdays))
            random.shuffle(day_order)
            for d in day_order:
                if task["remaining"] <= 0:
                    break
                for p_start in range(ppd - 1):
                    if task["remaining"] <= 0:
                        break
                    if self._gen_can_place(task, d, p_start):
                        self._gen_place(task, d, p_start)
                        if (task["remaining"] > 0
                                and self._gen_can_place(task, d, p_start + 1)):
                            self._gen_place(task, d, p_start + 1)
                        break

        # ── Phase 4 — SC2: Daily subjects ─────────────────────────────────
        # BUG2 FIX: try same-period-each-day first; fall back to per-day.
        self._gen_prog("Stage 3 · Phase 4 — Daily subjects…")
        sc2_tasks = sorted([t for t in tasks if t["priority"] == "SC2"
                             and t["remaining"] > 0],
                           key=lambda t: -t["periods"])
        for task in sc2_tasks:
            if task["remaining"] <= 0:
                continue
            placed = False
            for p in range(ppd):
                avail = [d for d in range(wdays)
                         if self._gen_can_place(task, d, p)]
                if len(avail) >= task["remaining"]:
                    for d in avail[:task["remaining"]]:
                        self._gen_place(task, d, p)
                    placed = True
                    break
            if not placed:
                # Fallback: best-effort per-day placement
                for d in range(wdays):
                    if task["remaining"] <= 0:
                        break
                    for p in range(ppd):
                        if self._gen_can_place(task, d, p):
                            self._gen_place(task, d, p)
                            break

        # ── Phase 5 — Fillers ──────────────────────────────────────────────
        # BUG3 FIX: sort by most-remaining-AND-fewest-available (constrained
        # first) without an expensive full-grid MRV scan.
        self._gen_prog("Stage 3 · Phase 5 — Filling remaining slots…")
        remaining = [t for t in tasks if t["remaining"] > 0]
        remaining.sort(key=lambda t: -t["periods"])
        for task in remaining:
            for d in range(wdays):
                if task["remaining"] <= 0:
                    break
                for p in range(ppd):
                    if task["remaining"] <= 0:
                        break
                    if self._gen_can_place(task, d, p):
                        self._gen_place(task, d, p)

        # ── Repair loop ────────────────────────────────────────────────────
        # Direct placement + one-level swap only.  No chain swap (too slow).
        # Relax constraints progressively; stop when no progress possible.
        relax_level = 0
        for rep in range(80):
            remaining_tasks = [t for t in tasks if t["remaining"] > 0]
            if not remaining_tasks:
                break
            self._gen_prog("Stage 3 · Repair {}: {} unplaced, relax={}".format(
                rep + 1,
                sum(t["remaining"] for t in remaining_tasks),
                relax_level))

            ix_sc1 = relax_level >= 1
            ix_sc3 = relax_level >= 2
            ix_sc2 = relax_level >= 3
            if ix_sc1:
                for t in tasks: t["rx_sc1"] = True
            if ix_sc3:
                for t in tasks: t["rx_sc3"] = True
            if ix_sc2:
                for t in tasks: t["rx_sc2"] = True

            progress = False

            for task in sorted(remaining_tasks, key=lambda t: -t["remaining"]):
                if task["remaining"] <= 0:
                    continue
                pt = task["par_teach"]

                # ── Direct placement ──────────────────────────────────────
                for d in range(wdays):
                    if task["remaining"] <= 0: break
                    for p in range(ppd):
                        if task["remaining"] <= 0: break
                        if self._gen_can_place(task, d, p, ix_sc1, ix_sc3, ix_sc2):
                            self._gen_place(task, d, p)
                            progress = True

                if task["remaining"] <= 0:
                    continue

                # ── One-level swap: try to displace one blocker ───────────
                # Stop as soon as we place one slot (don't exhaust all slots).
                swap_done = False
                for d in range(wdays):
                    if task["remaining"] <= 0 or swap_done: break
                    for p in range(ppd):
                        if task["remaining"] <= 0 or swap_done: break

                        # Teacher must be free at (d,p)
                        if not (g["t_free"](task["teacher"], d, p) and
                                (ix_sc3 or not g["t_unavail"](task["teacher"], d, p))):
                            continue
                        if pt and not (g["t_free"](pt, d, p) and
                                       (ix_sc3 or not g["t_unavail"](pt, d, p))):
                            continue

                        # Hard constraint checks
                        if task["is_ct"] and (p + 1) != task["ct_period"]:
                            continue
                        if task["p_pref"] and not task["is_ct"]:
                            if (p + 1) not in task["p_pref"]: continue
                        if task["d_pref"] and DAYS[d] not in task["d_pref"]:
                            continue
                        if not task["consec"]:
                            dup = False
                            for cn in task["cn_list"]:
                                for pp in range(ppd):
                                    e = g["grid"][cn][d][pp]
                                    if e and e.get("subject") == task["subject"]:
                                        dup = True; break
                                if dup: break
                            if dup: continue

                        # Find what is blocking
                        blocking_idx = None
                        for cn in task["cn_list"]:
                            if g["grid"][cn][d][p] is not None:
                                blocking_idx = g["task_at"][cn][d][p]
                                break

                        if blocking_idx is None:
                            self._gen_place(task, d, p)
                            progress = True
                            swap_done = True
                            break

                        if blocking_idx >= len(tasks): continue
                        blocker = tasks[blocking_idx]
                        if blocker["priority"] in ("HC1", "HC2"): continue

                        # Try to move blocker to its first available free slot
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p): continue
                                if not self._gen_can_place(
                                        blocker, d2, p2, ix_sc1, ix_sc3, ix_sc2):
                                    continue
                                self._gen_unplace(blocker, d, p)
                                slot_clear = all(
                                    g["grid"][cn][d][p] is None
                                    for cn in task["cn_list"])
                                t_now = (g["t_free"](task["teacher"], d, p) and
                                         (not pt or g["t_free"](pt, d, p)))
                                if slot_clear and t_now:
                                    self._gen_place(task, d, p)
                                    self._gen_place(blocker, d2, p2)
                                    progress = True
                                    swap_done = True
                                    moved = True
                                    break
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved or swap_done: break
                        if swap_done: break

            if not progress:
                relax_level += 1
                if relax_level > 4:
                    break

        # ── Store result ────────────────────────────────────────────────────
        unplaced = sum(t["remaining"] for t in tasks)
        tt = self._gen_snapshot_tt()
        self._timetable = tt
        self._gen_stage = 3
        self._stage2_status = {
            'unplaced':  unplaced,
            'ok':        unplaced == 0,
            'msg':       ("✅ Complete timetable generated — all periods placed!" if unplaced == 0
                          else "⚠ {} period(s) still unplaced — constraints may be too tight.".format(unplaced)),
            'stage_msg': ("Stage 3 complete ✓ — Full timetable generated!" if unplaced == 0
                          else "Stage 3 done — {} period(s) unplaced".format(unplaced)),
        }

    def _build_timetable(self):
        """Legacy single-shot builder — kept for backward compat. Not used in staged flow."""
        self._init_gen_state()
        self._run_stage1_phases()
        self._run_stage2_phases()
        return self._gen_snapshot_tt()

    # =========================================================================
    #  DISPLAY
    # =========================================================================

    def _force_fill_backtrack(self, progress_cb=None):
        """
        Two-stage guaranteed timetable completion using Min-Conflicts CSP.
        progress_cb(msg) is called periodically to update the UI label.
        """
        import random as _rnd

        def _prog(msg):
            if progress_cb:
                progress_cb(msg)

        g      = self._gen
        tasks  = g['tasks']
        grid   = g['grid']
        wdays  = g['wdays']
        ppd    = g['ppd']

        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()
        if not hasattr(self, '_relaxed_consec_keys'):
            self._relaxed_consec_keys = set()

        relaxed_notes = []
        PRIO_W = {'HC1': 0, 'HC2': 1, 'SC1': 2, 'SC2': 3, 'filler': 4}

        def _prio(t):
            return PRIO_W.get(t['priority'], 4)

        def _unplaced():
            return sum(t['remaining'] for t in tasks)

        def _can(task, d, p, ign_sc1=False, ign_sc3=False):
            return self._gen_can_place(task, d, p,
                                       ignore_sc1=ign_sc1,
                                       ignore_sc3=ign_sc3)

        # ─────────────────────────────────────────────────────────────────────
        # STAGE A: greedy + MRV with progressive constraint relaxation
        # ─────────────────────────────────────────────────────────────────────
        def _greedy_pass(ign_sc1=False, ign_sc3=False):
            remaining_tasks = [t for t in tasks if t['remaining'] > 0]
            remaining_tasks.sort(key=lambda t: sum(
                1 for d in range(wdays) for p in range(ppd)
                if _can(t, d, p, ign_sc1, ign_sc3)))
            for task in remaining_tasks:
                for d in range(wdays):
                    if task['remaining'] <= 0:
                        break
                    for p in range(ppd):
                        if task['remaining'] <= 0:
                            break
                        if _can(task, d, p, ign_sc1, ign_sc3):
                            self._gen_place(task, d, p)

        def _swap_pass(ign_sc1=False, ign_sc3=False):
            for task in sorted(tasks, key=lambda t: -t['remaining']):
                if task['remaining'] <= 0 or _prio(task) == 0:
                    continue
                for d in range(wdays):
                    if task['remaining'] <= 0:
                        break
                    for p in range(ppd):
                        if task['remaining'] <= 0:
                            break
                        tname = task['teacher']
                        pt    = task.get('par_teach', '')
                        t_ok  = g['t_free'](tname, d, p)
                        if not ign_sc3:
                            t_ok = t_ok and not g['t_unavail'](tname, d, p)
                        if pt and pt not in ('', '—', '?'):
                            t_ok = t_ok and g['t_free'](pt, d, p)
                        if not t_ok:
                            continue
                        bidx = None
                        for cn in task['cn_list']:
                            if grid[cn][d][p] is not None:
                                bidx = g['task_at'][cn][d][p]
                                break
                        if bidx is None:
                            if _can(task, d, p, ign_sc1, ign_sc3):
                                self._gen_place(task, d, p)
                            continue
                        blocker = tasks[bidx]
                        if _prio(blocker) <= _prio(task):
                            continue
                        for d2 in range(wdays):
                            moved = False
                            for p2 in range(ppd):
                                if (d2, p2) == (d, p):
                                    continue
                                if not _can(blocker, d2, p2, ign_sc1, ign_sc3):
                                    continue
                                self._gen_unplace(blocker, d, p)
                                clr = all(grid[cn][d][p] is None
                                          for cn in task['cn_list'])
                                tok = (g['t_free'](tname, d, p)
                                       and (not pt or pt in ('','—','?')
                                            or g['t_free'](pt, d, p)))
                                if clr and tok:
                                    self._gen_place(blocker, d2, p2)
                                    if _can(task, d, p, ign_sc1, ign_sc3):
                                        self._gen_place(task, d, p)
                                        moved = True
                                        break
                                    else:
                                        self._gen_unplace(blocker, d2, p2)
                                        self._gen_place(blocker, d, p)
                                else:
                                    self._gen_place(blocker, d, p)
                            if moved:
                                break

        def _run_stage_a(ign_sc1=False, ign_sc3=False):
            for _ in range(4):
                if _unplaced() == 0:
                    return
                _greedy_pass(ign_sc1, ign_sc3)
            for _ in range(4):
                if _unplaced() == 0:
                    return
                _swap_pass(ign_sc1, ign_sc3)
                _greedy_pass(ign_sc1, ign_sc3)

        _prog("Stage A — greedy placement…")
        _run_stage_a()
        if _unplaced() == 0:
            _prog("")
            return None

        _prog("Stage A — relaxing consecutive…")
        consec_items = []
        for t in tasks:
            if t['consec'] and t['remaining'] > 0:
                t['rx_sc1'] = True
                for cn_i in t['cn_list']:
                    self._relaxed_consec_keys.add((cn_i, t['subject']))
                consec_items.append("  • {} — {}".format(
                    '+'.join(t['cn_list']), t['subject']))
        if consec_items:
            relaxed_notes.append(
                "Consecutive constraint relaxed for:\n" + '\n'.join(consec_items))
        _run_stage_a(ign_sc1=True)
        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        _prog("Stage A — relaxing unavailability…")
        unav_set = set()
        for t in tasks:
            if t['remaining'] > 0:
                t['rx_sc3'] = True
                if t['teacher']:
                    unav_set.add(t['teacher'])
                pt = t.get('par_teach', '')
                if pt and pt not in ('', '—', '?'):
                    unav_set.add(pt)
        if unav_set:
            relaxed_notes.append(
                "Teacher unavailability bypassed for:\n"
                + '\n'.join("  • {}".format(x) for x in sorted(unav_set)))
        _run_stage_a(ign_sc1=True, ign_sc3=True)
        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        _prog("Stage A — relaxing preferences…")
        main_items = []
        for t in tasks:
            if t['remaining'] == 0 or t.get('is_ct'):
                continue
            if t['p_pref'] or t['d_pref'] or t.get('daily') or t['priority'] == 'SC2':
                t['p_pref']   = []
                t['d_pref']   = []
                t['daily']    = False
                t['priority'] = 'filler'
                self._relaxed_main_keys.add(
                    (frozenset(t['cn_list']), t['subject']))
                main_items.append("  • {} — {}".format(
                    '+'.join(t['cn_list']), t['subject']))
        if main_items:
            relaxed_notes.append(
                "Period/day preferences relaxed for:\n" + '\n'.join(main_items))
        _run_stage_a(ign_sc1=True, ign_sc3=True)
        if _unplaced() == 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # ─────────────────────────────────────────────────────────────────────
        # STAGE B: Min-Conflicts CSP solver
        # Hard cap: 1500 iterations with anti-cycling random restarts every 150.
        # ─────────────────────────────────────────────────────────────────────
        relaxed_notes.append(
            "Min-Conflicts solver applied: soft constraints overridden "
            "to guarantee complete placement.")

        # Mark everything relaxed
        for t in tasks:
            t['rx_sc1'] = True
            t['rx_sc3'] = True

        # ── B-1: force-complete — stuff remaining into any free class slot ──
        _prog("Stage B — force-completing grid…")
        for task in sorted(tasks, key=lambda t: -t['remaining']):
            if task['remaining'] <= 0:
                continue
            for d in range(wdays):
                if task['remaining'] <= 0:
                    break
                for p in range(ppd):
                    if task['remaining'] <= 0:
                        break
                    cells_free = all(
                        grid[cn][d][p] is None for cn in task['cn_list'])
                    if not cells_free:
                        continue
                    # Don't overwrite HC1 slots
                    hc1 = False
                    for cn in task['cn_list']:
                        bidx = g['task_at'][cn][d][p]
                        if bidx is not None and tasks[bidx]['priority'] == 'HC1':
                            hc1 = True; break
                    if hc1:
                        continue
                    for cn in task['cn_list']:
                        grid[cn][d][p] = self._gen_make_cell(task)
                        g['task_at'][cn][d][p] = task['idx']
                    g['t_mark'](task['teacher'], d, p)
                    pt = task.get('par_teach', '')
                    if pt and pt not in ('', '—', '?'):
                        g['t_mark'](pt, d, p)
                    task['remaining'] -= 1

        # If still unplaced after B-1, the problem is truly infeasible
        if _unplaced() > 0:
            _prog("")
            return '\n\n'.join(relaxed_notes)

        # ── B-2: conflict scorer ─────────────────────────────────────────────
        def _slot_conflicts(tname, pt, d, p, own_idx):
            """Count teacher double-bookings at (d,p) excluding own_idx."""
            score = 0
            for cn2 in g['all_classes']:
                idx2 = g['task_at'][cn2][d][p]
                if idx2 is None or idx2 == own_idx:
                    continue
                other = tasks[idx2]
                if other['teacher'] == tname:
                    score += 1
                if pt and pt not in ('', '—', '?'):
                    if other['teacher'] == pt or other.get('par_teach','') == pt:
                        score += 1
            return score

        def _build_task_slots():
            """task_idx -> list of (d,p) it currently occupies."""
            ts = {t['idx']: [] for t in tasks}
            for cn in g['all_classes']:
                for d in range(wdays):
                    for p in range(ppd):
                        idx = g['task_at'][cn][d][p]
                        if idx is not None:
                            if (d, p) not in ts[idx]:
                                ts[idx].append((d, p))
            return ts

        def _total_conflicts(task_slots):
            total = 0
            for t in tasks:
                if t['priority'] == 'HC1':
                    continue
                pt = t.get('par_teach', '')
                for d, p in task_slots[t['idx']]:
                    total += _slot_conflicts(t['teacher'], pt, d, p, t['idx'])
            return total

        # ── B-3: min-conflicts repair ────────────────────────────────────────
        MAX_ITER      = 1500
        RESTART_EVERY = 150    # random restart if no improvement for this many iters
        best_conflicts = None
        no_improve_count = 0

        for _iter in range(MAX_ITER):
            task_slots = _build_task_slots()
            total_conf = _total_conflicts(task_slots)

            if _iter % 20 == 0:
                _prog("Stage B — conflicts: {}  (iter {}/{})".format(
                    total_conf, _iter, MAX_ITER))

            if total_conf == 0:
                break   # ✅ Done

            # Track improvement for anti-cycling
            if best_conflicts is None or total_conf < best_conflicts:
                best_conflicts   = total_conf
                no_improve_count = 0
            else:
                no_improve_count += 1

            # Random restart if stuck in a plateau
            if no_improve_count >= RESTART_EVERY:
                _prog("Stage B — restart (stuck at {} conflicts)…".format(total_conf))
                # Shuffle all non-HC1 tasks' slots randomly
                non_hc1 = [t for t in tasks if t['priority'] != 'HC1']
                _rnd.shuffle(non_hc1)
                for t in non_hc1:
                    slots = task_slots[t['idx']]
                    for d, p in slots[:]:
                        # Unplace
                        for cn in t['cn_list']:
                            grid[cn][d][p] = None
                            g['task_at'][cn][d][p] = None
                        g['t_unmark'](t['teacher'], d, p)
                        pt2 = t.get('par_teach', '')
                        if pt2 and pt2 not in ('', '—', '?'):
                            g['t_unmark'](pt2, d, p)
                        t['remaining'] += 1
                    # Re-place greedily
                    free_slots = [(d, p) for d in range(wdays)
                                  for p in range(ppd)
                                  if all(grid[cn][d][p] is None
                                         for cn in t['cn_list'])]
                    _rnd.shuffle(free_slots)
                    for d, p in free_slots:
                        if t['remaining'] <= 0:
                            break
                        for cn in t['cn_list']:
                            grid[cn][d][p] = self._gen_make_cell(t)
                            g['task_at'][cn][d][p] = t['idx']
                        g['t_mark'](t['teacher'], d, p)
                        pt2 = t.get('par_teach', '')
                        if pt2 and pt2 not in ('', '—', '?'):
                            g['t_mark'](pt2, d, p)
                        t['remaining'] -= 1
                no_improve_count = 0
                best_conflicts   = None
                continue

            # Pick most-conflicted non-HC1 task
            conflicted = [(t, sum(_slot_conflicts(
                              t['teacher'], t.get('par_teach',''), d, p, t['idx'])
                              for d, p in task_slots[t['idx']]))
                          for t in tasks
                          if t['priority'] != 'HC1'
                          and sum(_slot_conflicts(
                              t['teacher'], t.get('par_teach',''), d, p, t['idx'])
                              for d, p in task_slots[t['idx']]) > 0]
            if not conflicted:
                break

            target, _ = max(conflicted, key=lambda x: x[1])
            t_slots    = task_slots[target['idx']]
            if not t_slots:
                continue

            # Worst slot for this task
            worst_d, worst_p = max(
                t_slots,
                key=lambda dp: _slot_conflicts(
                    target['teacher'], target.get('par_teach',''),
                    dp[0], dp[1], target['idx']))

            # Unplace that slot
            for cn in target['cn_list']:
                grid[cn][worst_d][worst_p] = None
                g['task_at'][cn][worst_d][worst_p] = None
            g['t_unmark'](target['teacher'], worst_d, worst_p)
            pt = target.get('par_teach', '')
            if pt and pt not in ('', '—', '?'):
                g['t_unmark'](pt, worst_d, worst_p)
            target['remaining'] += 1

            # Find free class slot with minimum teacher conflicts
            best_score = None
            best_d, best_p = worst_d, worst_p   # fallback = same slot
            for d in range(wdays):
                for p in range(ppd):
                    cells_free = all(
                        grid[cn][d][p] is None for cn in target['cn_list'])
                    if not cells_free:
                        continue
                    sc = _slot_conflicts(
                        target['teacher'], pt, d, p, target['idx'])
                    if best_score is None or sc < best_score:
                        best_score = sc
                        best_d, best_p = d, p
                    if best_score == 0:
                        break
                if best_score == 0:
                    break

            # Place at best slot
            for cn in target['cn_list']:
                grid[cn][best_d][best_p] = self._gen_make_cell(target)
                g['task_at'][cn][best_d][best_p] = target['idx']
            g['t_mark'](target['teacher'], best_d, best_p)
            if pt and pt not in ('', '—', '?'):
                g['t_mark'](pt, best_d, best_p)
            target['remaining'] -= 1

        _prog("")
        return '\n\n'.join(relaxed_notes) if relaxed_notes else None

    # ── Snapshot / Restore (for undo-on-no-improvement) ──────────────────────

    def _ft_snapshot(self):
        """
        Deep-copy all mutable gen state so that any action can be fully undone.

        Captures:
          - grid cells (dict of CN → list-of-lists of cell dicts)
          - task_at   (dict of CN → list-of-lists of idx or None)
          - t_busy    (dict of teacher → set of (d, p))
          - per-task mutable fields
          - relaxed key sets
        """
        import copy
        g = self._gen

        # Grid: each cell is either None or a small dict — shallow copy of the dict
        # is sufficient because cells are replaced wholesale (never mutated in place).
        grid_snap = {
            cn: [[g['grid'][cn][d][p] for p in range(g['ppd'])]
                 for d in range(g['wdays'])]
            for cn in g['all_classes']
        }
        task_at_snap = {
            cn: [[g['task_at'][cn][d][p] for p in range(g['ppd'])]
                 for d in range(g['wdays'])]
            for cn in g['all_classes']
        }
        t_busy_snap = {t: set(s) for t, s in g['t_busy'].items()}

        tasks_snap = [
            {
                'idx':       task['idx'],
                'remaining': task['remaining'],
                'rx_sc1':    task['rx_sc1'],
                'rx_sc2':    task['rx_sc2'],
                'rx_sc3':    task['rx_sc3'],
                'p_pref':    list(task['p_pref']),
                'd_pref':    list(task['d_pref']),
                'daily':     task['daily'],
                'priority':  task['priority'],
                'consec':    task['consec'],
            }
            for task in g['tasks']
        ]

        return {
            'grid':               grid_snap,
            'task_at':            task_at_snap,
            't_busy':             t_busy_snap,
            'tasks':              tasks_snap,
            'relaxed_consec':     set(self._relaxed_consec_keys),
            'relaxed_main':       set(getattr(self, '_relaxed_main_keys', set())),
        }

    def _ft_restore(self, snap):
        """Restore gen state from a snapshot produced by _ft_snapshot."""
        g = self._gen

        # Restore grid and task_at
        for cn in g['all_classes']:
            for d in range(g['wdays']):
                for p in range(g['ppd']):
                    g['grid'][cn][d][p]    = snap['grid'][cn][d][p]
                    g['task_at'][cn][d][p] = snap['task_at'][cn][d][p]

        # Restore t_busy
        g['t_busy'].clear()
        for t, s in snap['t_busy'].items():
            g['t_busy'][t] = set(s)

        # Restore per-task fields
        task_map = {t['idx']: t for t in g['tasks']}
        for ts in snap['tasks']:
            task = task_map.get(ts['idx'])
            if task is None:
                continue
            task['remaining'] = ts['remaining']
            task['rx_sc1']    = ts['rx_sc1']
            task['rx_sc2']    = ts['rx_sc2']
            task['rx_sc3']    = ts['rx_sc3']
            task['p_pref']    = list(ts['p_pref'])
            task['d_pref']    = list(ts['d_pref'])
            task['daily']     = ts['daily']
            task['priority']  = ts['priority']
            task['consec']    = ts['consec']

        # Restore relaxed-key sets
        self._relaxed_consec_keys = set(snap['relaxed_consec'])
        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()
        self._relaxed_main_keys.clear()
        self._relaxed_main_keys.update(snap['relaxed_main'])

    def _ft_targetable(self):
        """Return tasks that are candidates for re-allocation (filler / consec / parallel)."""
        result = []
        for t in self._gen['tasks']:
            p = t['priority']
            if p in ('filler', 'SC1') or t['type'] in ('parallel', 'combined_parallel'):
                result.append(t)
        return result

    def _ft_teacher_free_slots(self):
        """Return {teacher: count_of_free_usable_slots}."""
        g = self._gen
        wdays = g['wdays']; ppd = g['ppd']
        t_busy = g['t_busy']
        t_unavail = g['t_unavail']
        teachers = set()
        for task in g['tasks']:
            if task['teacher']:
                teachers.add(task['teacher'])
            if task['par_teach']:
                teachers.add(task['par_teach'])
        result = {}
        for t in teachers:
            busy = t_busy.get(t, set())
            free = 0
            for d in range(wdays):
                for p in range(ppd):
                    if (d, p) not in busy and not t_unavail(t, d, p):
                        free += 1
            result[t] = free
        return result

    def _ft_try_place_task(self, task, ignore_sc1=False, ignore_sc3=False):
        """Greedily place all remaining slots of *task*. Returns count newly placed."""
        g = self._gen
        wdays = g['wdays']; ppd = g['ppd']
        placed = 0
        for d in range(wdays):
            if task['remaining'] == 0:
                break
            for p in range(ppd):
                if task['remaining'] == 0:
                    break
                if self._gen_can_place(task, d, p,
                                       ignore_sc1=ignore_sc1,
                                       ignore_sc3=ignore_sc3):
                    self._gen_place(task, d, p)
                    placed += 1
        return placed

    def _ft_unplace_task(self, task):
        """Remove all placed slots of *task* from the grid and return count removed."""
        g = self._gen
        grid = g['grid']; wdays = g['wdays']; ppd = g['ppd']
        removed = 0
        for d in range(wdays):
            for p in range(ppd):
                for cn in task['cn_list']:
                    if g['task_at'][cn][d][p] == task['idx']:
                        self._gen_unplace(task, d, p)
                        removed += 1
                        break   # one unplace per (d,p) slot is enough
        return removed

    # ── Task A: Allocate ──────────────────────────────────────────────────────

    def _task_allocate(self):
        """
        Task A — Smart allocation of unplaced filler/consecutive/parallel periods.

        1. Compute free-slot counts for every teacher.
        2. Allocate tasks whose teacher has fewest free slots first (most constrained).
        3. Second pass to catch anything still unplaced.

        Returns a summary string (or None).
        """
        targets = self._ft_targetable()
        if not targets:
            return None

        from collections import defaultdict
        t_tasks = defaultdict(list)
        for task in targets:
            if task['remaining'] > 0:
                t_tasks[task['teacher']].append(task)

        if not t_tasks:
            return None

        free_counts = self._ft_teacher_free_slots()
        ordered_teachers = sorted(t_tasks.keys(), key=lambda t: free_counts.get(t, 0))

        for teacher in ordered_teachers:
            for task in t_tasks[teacher]:
                if task['remaining'] > 0:
                    self._ft_try_place_task(task)

        # Second pass
        for task in targets:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        return None

    # ── Task S: Shuffle ───────────────────────────────────────────────────────

    def _task_shuffle(self):
        """
        Task S — Unplace all targetable tasks, re-sort by constraint difficulty, re-place.

        Sort order (most constrained first):
          1. Consecutive (SC1)
          2. Combined parallel
          3. Parallel
          4. Daily filler
          5. Standard filler

        Returns a summary string (or None).
        """
        targets = self._ft_targetable()
        if not targets:
            return None

        for task in targets:
            self._ft_unplace_task(task)

        def _sort_key(t):
            if t['consec']:            return (0, -t['periods'])
            if t['type'] == 'combined_parallel': return (1, -t['periods'])
            if t['type'] == 'parallel':          return (2, -t['periods'])
            if t.get('daily'):                   return (3, -t['periods'])
            return (4, -t['periods'])

        free_counts = self._ft_teacher_free_slots()
        targets_sorted = sorted(
            targets,
            key=lambda t: (_sort_key(t), free_counts.get(t['teacher'], 0)))

        for task in targets_sorted:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        # Final sweep
        for task in targets_sorted:
            if task['remaining'] > 0:
                self._ft_try_place_task(task)

        return None

    # ── Task C: Relax Consecutive ─────────────────────────────────────────────

    def _task_relax_consecutive(self):
        """
        Task C — Relax consecutive constraints for unplaced SC1 tasks.

        For each task with consec=True and remaining > 0:
          - Set rx_sc1=True (bypass the consecutive-placement rule)
          - Try placing freely
        Returns a summary string listing which rules were relaxed.
        """
        relaxed = []
        for task in self._gen['tasks']:
            if task['consec'] and task['remaining'] > 0:
                task['rx_sc1'] = True
                placed = self._ft_try_place_task(task, ignore_sc1=True)
                if placed > 0:
                    cn = '+'.join(task['cn_list'])
                    relaxed.append((cn, task['subject']))
                    for cn_i in task['cn_list']:
                        self._relaxed_consec_keys.add((cn_i, task['subject']))

        # Second pass
        for task in self._gen['tasks']:
            if task.get('rx_sc1') and task['remaining'] > 0:
                self._ft_try_place_task(task, ignore_sc1=True)

        if relaxed:
            return ("Consecutive rules relaxed for:\n\n" +
                    "\n".join("• {} → {}".format(cn, subj)
                              for cn, subj in sorted(relaxed)))
        return None

    # ── Task M: Relax Main Periods ────────────────────────────────────────────

    def _task_relax_main_periods(self):
        """
        Task M — Convert unplaced main-period tasks to filler, then run Task A.

        Targets: HC2 tasks (period/day preferences) and SC2 tasks (daily) that
        still have remaining > 0.  CT periods (HC1) are never touched.

        Returns a summary string listing which tasks were converted.
        """
        if not hasattr(self, '_relaxed_main_keys'):
            self._relaxed_main_keys = set()

        relaxed = []
        for task in self._gen['tasks']:
            if task['remaining'] == 0 or task['is_ct']:
                continue
            if not (task['p_pref'] or task['d_pref'] or task.get('daily')):
                continue

            task['p_pref']   = []
            task['d_pref']   = []
            task['daily']    = False
            task['priority'] = 'filler'
            cn = '+'.join(task['cn_list'])
            relaxed.append((cn, task['subject']))
            self._relaxed_main_keys.add((frozenset(task['cn_list']), task['subject']))

        if relaxed:
            self._task_allocate()
            return ("Main periods converted to filler for:\n\n" +
                    "\n".join("• {} → {}".format(cn, subj)
                              for cn, subj in sorted(relaxed)))
        return None

    # ── Task UN: Relax Unavailability ─────────────────────────────────────────

    def _task_relax_unavailability(self):
        """
        Task UN — Override unavailability rules for teachers with unplaced periods,
        then run Task A.

        Returns a summary string listing which teachers' unavailability was bypassed.
        """
        unplaced_teachers = set()
        for task in self._gen['tasks']:
            if task['remaining'] > 0:
                if task['teacher']:
                    unplaced_teachers.add(task['teacher'])
                if task['par_teach']:
                    unplaced_teachers.add(task['par_teach'])

        if not unplaced_teachers:
            return None

        # Mark rx_sc3 on all tasks for those teachers
        for task in self._gen['tasks']:
            if task['remaining'] > 0:
                if (task['teacher'] in unplaced_teachers or
                        task['par_teach'] in unplaced_teachers):
                    task['rx_sc3'] = True

        # Run allocation with unavailability bypassed
        from collections import defaultdict
        targets = self._ft_targetable()
        free_counts = self._ft_teacher_free_slots()
        t_tasks = defaultdict(list)
        for task in targets:
            if task['remaining'] > 0:
                t_tasks[task['teacher']].append(task)

        ordered = sorted(t_tasks.keys(), key=lambda t: free_counts.get(t, 0))
        for teacher in ordered:
            for task in t_tasks[teacher]:
                if task['remaining'] > 0:
                    self._ft_try_place_task(task,
                                            ignore_sc3=task.get('rx_sc3', False))

        # Also try non-targetable tasks with rx_sc3
        for task in self._gen['tasks']:
            if task['remaining'] > 0 and task.get('rx_sc3'):
                self._ft_try_place_task(task, ignore_sc3=True)

        return ("Unavailability rules bypassed for:\n\n" +
                "\n".join("• {}".format(t) for t in sorted(unplaced_teachers)))

    # ── Class timetable view ──────────────────────────────────────────────

    def _get_combined_par_display(self, cn, e):
        """For a combined_parallel cell, return (line1, line2) where:
            line1 = "combined_subject / class_subject"   e.g. "URDU / SKT"
            line2 = "combined_teacher / class_teacher"   e.g. "Irfan / Anita"

        Uses step3_data as the ground truth for which teacher+subject is the
        combined one, then looks up class_config_data[cn] for the per-class entry.
        """
        cc = e.get('combined_classes', [])

        # ── Step 1: find combined teacher + subject from step3_data ──────────
        combined_teacher = ''
        combined_subj    = ''
        s3 = getattr(self, 'step3_data', {})
        for _t, s3d in s3.items():
            for cb in s3d.get('combines', []):
                if set(cb.get('classes', [])) == set(cc):
                    combined_teacher = _t
                    combined_subj    = cb.get('subjects', [''])[0] if cb.get('subjects') else ''
                    break
            if combined_teacher:
                break

        # ── Step 2: look up this class's entry to find the class-specific side ─
        class_subj    = ''
        class_teacher = ''
        if combined_subj and cn in self.class_config_data:
            for _s in self.class_config_data[cn].get('subjects', []):
                sname = _s.get('name', '').strip()
                pname = (_s.get('parallel_subject') or '').strip()
                if sname == combined_subj:
                    # primary = combined, parallel = class-specific
                    class_subj    = pname
                    class_teacher = (_s.get('parallel_teacher') or '').strip()
                    break
                elif pname == combined_subj:
                    # parallel = combined, primary = class-specific
                    class_subj    = sname
                    class_teacher = _s.get('teacher', '').strip()
                    break

        # ── Fallback: use cell data if step3_data lookup failed ──────────────
        if not combined_subj:
            combined_subj    = e.get('subject', '')
            combined_teacher = e.get('teacher', '')
            class_subj       = e.get('par_subj', '')
            class_teacher    = e.get('par_teach', '')

        return (
            "{} / {}".format(combined_subj, class_subj),
            "{} / {}".format(combined_teacher, class_teacher),
        )


    def get_excel_bytes(self, mode):
        """Generate Excel workbook and return raw bytes (for Streamlit download)."""
        import io
        buf = io.BytesIO()
        self._write_excel_buf(buf, mode)
        return buf.getvalue()


    def _write_excel_buf(self, filename, mode):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        from collections import defaultdict

        tt          = self._timetable
        days        = tt['days']
        ppd         = tt['ppd']
        half1       = tt['half1']
        grid        = tt['grid']
        all_classes = tt['all_classes']

        def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
        def _font(bold=False, sz=9, col="000000"):
            return Font(bold=bold, size=sz, color=col.lstrip("#"), name="Arial")
        def _border():
            s = Side(style="thin", color="AAAAAA")
            return Border(left=s, right=s, top=s, bottom=s)
        def _align(h="center", wrap=True):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

        HDR_F  = _fill("#2c3e50"); HDR_N  = _font(True, 10, "FFFFFF")
        DAY_F  = _fill("#34495e"); DAY_N  = _font(True,  9, "FFFFFF")
        SUB_F  = _fill("#d5e8d4")
        COMB_F = _fill("#dae8fc")
        PAR_F  = _fill("#ffe6cc")
        CPAF   = _fill("#f8cecc")
        FREE_F = _fill("#f5f5f5")
        WHT_F  = _fill("#FFFFFF")
        SUM_F  = _fill("#eaf2ff")
        CT_H_F = _fill("#1a5276")
        WRN_F  = _fill("#fdebd0")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── shared: build teacher grid ───────────────────────────────────────
        def _build_tg():
            tg = {}
            for cn in all_classes:
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if not e: continue
                        etype = e.get('type', 'normal')
                        cc    = e.get('combined_classes', [])
                        is_cp = bool(cc) and etype == 'combined_parallel'
                        is_c  = bool(cc) and etype == 'combined'

                        def _add(tname, tcls, tsubj, tct):
                            if not tname: return
                            tg.setdefault(tname,
                                          [[None]*ppd for _ in range(len(days))])
                            tg[tname][d][p] = {
                                'class': tcls, 'subject': tsubj, 'is_ct': tct}

                        if is_cp:
                            if not cc or cn == cc[0]:
                                _add(e.get('teacher'), '+'.join(cc),
                                     e.get('subject',''), False)
                            pt = e.get('par_teach','')
                            if pt and pt not in ('—','?',''):
                                _add(pt, cn, e.get('par_subj',''),
                                     e.get('is_ct', False))
                        elif is_c:
                            if not cc or cn == cc[0]:
                                _add(e.get('teacher'), '+'.join(cc),
                                     e.get('subject',''), e.get('is_ct', False))
                        else:
                            _add(e.get('teacher'), cn,
                                 e.get('subject',''), e.get('is_ct', False))
                            pt = e.get('par_teach','')
                            if pt and pt not in ('—','?',''):
                                _add(pt, cn, e.get('par_subj',''), False)
            return tg

        def _sv(val):
            """Safely extract string from a StringVar or plain string."""
            if hasattr(val, 'get'):
                return val.get()
            return val or ''

        def _ct_map():
            ct = {}
            for cn in all_classes:
                cfg = self.class_config_data.get(cn, {})
                t = cfg.get('teacher', '').strip()
                if t:
                    ct.setdefault(t, []).append(cn)
            return ct

        # ─────────────────────────────────────────────────────────────────────
        # 1. CLASSWISE TIMETABLE — one sheet per class
        # ─────────────────────────────────────────────────────────────────────
        if mode == "class":
            for cn in all_classes:
                ws = wb.create_sheet(cn)
                cfg     = self.class_config_data.get(cn, {})
                ct_name = cfg.get('teacher', '').strip()
                ct_per  = str(cfg.get('teacher_period', ''))
                hdr_txt = "Class: {}   |   Class Teacher: {}{}".format(
                    cn, ct_name or '—',
                    "   |   CT Period: {}".format(ct_per) if ct_per else '')

                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=ppd+1)
                c = ws.cell(1, 1, hdr_txt)
                c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[1].height = 20

                ws.cell(2, 1, "Day")
                ws.cell(2, 1).fill = HDR_F; ws.cell(2, 1).font = HDR_N
                ws.cell(2, 1).alignment = _align(); ws.cell(2, 1).border = _border()
                for p in range(ppd):
                    h = ws.cell(2, p+2, "P{} {}".format(
                        p+1, "①" if p < half1 else "②"))
                    h.fill = HDR_F; h.font = HDR_N
                    h.alignment = _align(); h.border = _border()
                ws.row_dimensions[2].height = 16

                for d, dname in enumerate(days):
                    r = 3 + d
                    ws.row_dimensions[r].height = 48
                    dc = ws.cell(r, 1, dname)
                    dc.fill = DAY_F; dc.font = DAY_N
                    dc.alignment = _align(); dc.border = _border()
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if e is None:
                            txt = "FREE"; fill = FREE_F
                        else:
                            etype = e.get('type', 'normal')
                            if etype == 'combined_parallel':
                                l1, l2 = self._get_combined_par_display(cn, e)
                                txt = "{}\n{}".format(l1, l2); fill = CPAF
                            elif etype == 'parallel':
                                txt = "{} / {}\n{} / {}".format(
                                    e['subject'], e.get('par_subj',''),
                                    e['teacher'],  e.get('par_teach',''))
                                fill = PAR_F
                            elif etype == 'combined':
                                cc = e.get('combined_classes', [])
                                mark = " ★" if e.get('is_ct') else ""
                                txt = "{}{}[{}]\n{}".format(
                                    e['subject'], mark, '+'.join(cc), e['teacher'])
                                fill = COMB_F
                            else:
                                mark = " ★" if e.get('is_ct') else ""
                                txt  = "{}{}\n{}".format(e['subject'], mark, e['teacher'])
                                fill = SUB_F if e.get('is_ct') else WHT_F
                        c = ws.cell(r, p+2, txt)
                        c.fill = fill; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=8)

                # Summary
                sr = 3 + len(days) + 1
                ws.merge_cells(start_row=sr, start_column=1,
                               end_row=sr, end_column=ppd+1)
                c = ws.cell(sr, 1, "Summary — {}".format(cn))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align("left"); c.border = _border()

                smry = defaultdict(int)
                for d in range(len(days)):
                    for p in range(ppd):
                        e = grid.get(cn, [[]])[d][p] \
                            if d < len(grid.get(cn, [])) else None
                        if not e: continue
                        etype = e.get('type', 'normal')
                        if etype == 'combined_parallel':
                            l1, l2 = self._get_combined_par_display(cn, e)
                            for ln in (l1, l2):
                                parts = ln.split('\n')
                                smry[(parts[0].strip(),
                                      parts[1].strip() if len(parts) > 1 else '')] += 1
                        elif etype == 'parallel':
                            smry[(e['subject'], e['teacher'])] += 1
                            smry[(e.get('par_subj',''), e.get('par_teach',''))] += 1
                        else:
                            smry[(e['subject'], e['teacher'])] += 1

                hdr_r = sr + 1
                for col, txt in enumerate(["Subject", "Teacher", "Periods/Week"], 1):
                    c = ws.cell(hdr_r, col, txt)
                    c.fill = HDR_F; c.font = HDR_N
                    c.alignment = _align(); c.border = _border()

                for i, ((subj, teach), cnt) in enumerate(
                        sorted(smry.items())):
                    row = hdr_r + 1 + i
                    for col, val in enumerate([subj, teach, cnt], 1):
                        c = ws.cell(row, col, val)
                        c.fill = SUM_F if i % 2 == 0 else WHT_F
                        c.alignment = _align(); c.border = _border()
                        c.font = _font(sz=9)

                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        # ─────────────────────────────────────────────────────────────────────
        # 2. TEACHERWISE TIMETABLE — one sheet per teacher
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "teacher":
            tg    = _build_tg()
            ct_mp = _ct_map()

            for teacher in sorted(tg.keys()):
                ws    = wb.create_sheet(teacher[:31])
                tdata = tg[teacher]
                ctc   = ct_mp.get(teacher, [])
                hdr_txt = "Teacher: {}   |   Class Teacher of: {}".format(
                    teacher, ', '.join(ctc) if ctc else '—')

                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=ppd+1)
                c = ws.cell(1, 1, hdr_txt)
                c.fill = CT_H_F; c.font = _font(True, 11, "FFFFFF")
                c.alignment = _align(); c.border = _border()
                ws.row_dimensions[1].height = 20

                ws.cell(2, 1, "Day")
                ws.cell(2, 1).fill = HDR_F; ws.cell(2, 1).font = HDR_N
                ws.cell(2, 1).alignment = _align(); ws.cell(2, 1).border = _border()
                for p in range(ppd):
                    h = ws.cell(2, p+2, "P{} {}".format(
                        p+1, "①" if p < half1 else "②"))
                    h.fill = HDR_F; h.font = HDR_N
                    h.alignment = _align(); h.border = _border()
                ws.row_dimensions[2].height = 16

                for d, dname in enumerate(days):
                    r = 3 + d
                    ws.row_dimensions[r].height = 48
                    dc = ws.cell(r, 1, dname)
                    dc.fill = DAY_F; dc.font = DAY_N
                    dc.alignment = _align(); dc.border = _border()
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e is None:
                            txt = "FREE"; fill = FREE_F
                        else:
                            txt  = "{}\n{}".format(e['class'], e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        c = ws.cell(r, p+2, txt)
                        c.fill = fill; c.alignment = _align()
                        c.border = _border(); c.font = _font(sz=8)

                # Summary: class → subject → count
                sr    = 3 + len(days) + 1
                ws.merge_cells(start_row=sr, start_column=1,
                               end_row=sr, end_column=ppd+1)
                c = ws.cell(sr, 1, "Summary — {}".format(teacher))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align("left"); c.border = _border()

                smry  = defaultdict(lambda: defaultdict(int))
                total = 0
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            smry[e['class']][e['subject']] += 1
                            total += 1

                hdr_r = sr + 1
                for col, txt in enumerate(["Class", "Subject", "Periods/Week"], 1):
                    c = ws.cell(hdr_r, col, txt)
                    c.fill = HDR_F; c.font = HDR_N
                    c.alignment = _align(); c.border = _border()

                row = hdr_r + 1
                for cls in sorted(smry.keys()):
                    for subj, cnt in sorted(smry[cls].items()):
                        for col, val in enumerate([cls, subj, cnt], 1):
                            c = ws.cell(row, col, val)
                            c.fill = SUM_F if row % 2 == 0 else WHT_F
                            c.alignment = _align(); c.border = _border()
                            c.font = _font(sz=9)
                        row += 1

                for col, val in enumerate(["", "TOTAL", total], 1):
                    c = ws.cell(row, col, val)
                    c.fill = _fill("#d4e6f1"); c.font = _font(True, 9)
                    c.alignment = _align(); c.border = _border()

                ws.column_dimensions["A"].width = 12
                for p in range(ppd):
                    ws.column_dimensions[get_column_letter(p+2)].width = 20

        # ─────────────────────────────────────────────────────────────────────
        # 3. CLASS TEACHER LIST
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "ct_list":
            ws = wb.create_sheet("Class Teacher List")
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = "Class Teacher List"
            c.fill = HDR_F; c.font = _font(True, 13, "FFFFFF")
            c.alignment = _align(); c.border = _border()
            ws.row_dimensions[1].height = 22

            for col, txt in enumerate(["Class", "Class Teacher", "CT Period"], 1):
                c = ws.cell(2, col, txt)
                c.fill = DAY_F; c.font = DAY_N
                c.alignment = _align(); c.border = _border()

            for i, cn in enumerate(all_classes):
                cfg     = self.class_config_data.get(cn, {})
                ct_name = cfg.get('teacher', '').strip() or '—'
                ct_per  = str(cfg.get('teacher_period', '')) or '—'
                row = 3 + i
                for col, val in enumerate([cn, ct_name, ct_per], 1):
                    c = ws.cell(row, col, val)
                    c.fill = SUM_F if i % 2 == 0 else WHT_F
                    c.alignment = _align(); c.border = _border()
                    c.font = _font(sz=10)

            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 12

        # ─────────────────────────────────────────────────────────────────────
        # 4. TEACHER WORKLOAD LIST
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "workload":
            tg    = _build_tg()
            ct_mp = _ct_map()
            ws    = wb.create_sheet("Teacher Workload")

            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = "Teacher Workload List"
            c.fill = HDR_F; c.font = _font(True, 13, "FFFFFF")
            c.alignment = _align(); c.border = _border()
            ws.row_dimensions[1].height = 22

            for col, txt in enumerate(
                    ["Teacher", "Subject", "Class", "Periods/Week", "Total Periods"], 1):
                c = ws.cell(2, col, txt)
                c.fill = DAY_F; c.font = DAY_N
                c.alignment = _align(); c.border = _border()

            row = 3
            grand_total = 0
            for teacher in sorted(tg.keys()):
                tdata = tg[teacher]
                smry  = defaultdict(lambda: defaultdict(int))
                for d in range(len(days)):
                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            smry[e['subject']][e['class']] += 1

                total = sum(c for cd in smry.values() for c in cd.values())
                grand_total += total
                ctc = ct_mp.get(teacher, [])
                start_row = row

                for si, subj in enumerate(sorted(smry.keys())):
                    for cls, cnt in sorted(smry[subj].items()):
                        fill = SUM_F if row % 2 == 0 else WHT_F
                        c = ws.cell(row, 1, teacher if row == start_row else "")
                        c.fill = WRN_F if ctc else fill
                        c.font = _font(True if row == start_row else False, 9)
                        c.alignment = _align(); c.border = _border()

                        for col, val in enumerate([subj, cls, cnt], 2):
                            c2 = ws.cell(row, col, val)
                            c2.fill = fill
                            c2.alignment = _align(); c2.border = _border()
                            c2.font = _font(sz=9)

                        c5 = ws.cell(row, 5, total if row == start_row else "")
                        c5.fill = _fill("#d4e6f1") if row == start_row else fill
                        c5.font = _font(True if row == start_row else False, 9)
                        c5.alignment = _align(); c5.border = _border()
                        row += 1

                span = row - start_row
                if span > 1:
                    ws.merge_cells(start_row=start_row, start_column=1,
                                   end_row=row-1, end_column=1)

            for col, val in enumerate(["", "", "", "GRAND TOTAL", grand_total], 1):
                c = ws.cell(row, col, val)
                c.fill = HDR_F; c.font = _font(True, 10, "FFFFFF")
                c.alignment = _align(); c.border = _border()

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 16
            ws.column_dimensions["E"].width = 16

        # ─────────────────────────────────────────────────────────────────────
        # 5. ONE-SHEET TEACHERWISE
        #    TeacherName | DAY | P1 | P2 | … | Pn   (CLASS/SUBJECT per cell)
        # ─────────────────────────────────────────────────────────────────────
        elif mode == "one_sheet":
            tg = _build_tg()
            ws = wb.create_sheet("Teacherwise Timetable")
            ws.row_dimensions[1].height = 18

            ws.cell(1, 1, "Teacher")
            ws.cell(1, 2, "Day")
            for col in (1, 2):
                ws.cell(1, col).fill = HDR_F; ws.cell(1, col).font = HDR_N
                ws.cell(1, col).alignment = _align(); ws.cell(1, col).border = _border()
            for p in range(ppd):
                c = ws.cell(1, p+3, str(p+1))
                c.fill = HDR_F; c.font = HDR_N
                c.alignment = _align(); c.border = _border()

            row = 2
            for teacher in sorted(tg.keys()):
                tdata   = tg[teacher]
                t_start = row
                for d, dname in enumerate(days):
                    c = ws.cell(row, 1, teacher if d == 0 else "")
                    c.fill = WRN_F; c.alignment = _align()
                    c.font = _font(True if d == 0 else False, 9)
                    c.border = _border()

                    c2 = ws.cell(row, 2, dname)
                    c2.fill = DAY_F; c2.font = DAY_N
                    c2.alignment = _align(); c2.border = _border()

                    for p in range(ppd):
                        e = tdata[d][p] if d < len(tdata) else None
                        if e:
                            txt  = "{}/{}".format(e['class'], e['subject'])
                            fill = SUB_F if e.get('is_ct') else WHT_F
                        else:
                            txt  = ""; fill = FREE_F
                        c3 = ws.cell(row, p+3, txt)
                        c3.fill = fill; c3.alignment = _align()
                        c3.border = _border(); c3.font = _font(sz=8)
                    row += 1

                if len(days) > 1:
                    ws.merge_cells(start_row=t_start, start_column=1,
                                   end_row=row-1, end_column=1)

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 10
            for p in range(ppd):
                ws.column_dimensions[get_column_letter(p+3)].width = 18

        wb.save(filename)  # filename can be a file path or BytesIO

